#!/usr/bin/env python3
"""Gil's FlowDesk — Visual AI Workflow Editor Server (SQLite/PostgreSQL + Claude CLI)"""

import subprocess, json, os, re, uuid, sys, time, sqlite3, threading, hashlib, secrets, tempfile, shutil
try:
    import psycopg2
    import psycopg2.extras
    HAS_PSYCOPG2 = True
except ImportError:
    HAS_PSYCOPG2 = False
from datetime import datetime, timedelta
from http.server import HTTPServer, SimpleHTTPRequestHandler
from socketserver import ThreadingMixIn
from urllib.parse import urlparse, parse_qs

PORT = 8888
SESSION_NAME = "main"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(BASE_DIR, "config.json")

def load_config():
    """config.json에서 설정을 로드. 없으면 기본값 생성."""
    defaults = {
        "db_path": os.path.join(BASE_DIR, "canvas.db"),
        "port": 8888,
        "uploads_dir": os.path.join(BASE_DIR, "uploads")
    }
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            # 설정 파일의 값으로 defaults 업데이트
            for k, v in cfg.items():
                defaults[k] = v
        except Exception as e:
            print(f"[WARN] config.json 읽기 실패: {e}, 기본값 사용")
    else:
        # 최초 실행: 기본 config.json 생성
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(defaults, f, indent=2, ensure_ascii=False)
        print(f"[+] config.json 생성됨: {CONFIG_PATH}")
    return defaults

CONFIG = load_config()
# 환경변수가 있으면 config.json보다 우선 (Docker 배포 지원)
DB_TYPE = os.environ.get("DB_TYPE") or CONFIG.get("db_type", "sqlite")
DB_PATH = os.environ.get("DB_PATH") or CONFIG.get("db_path", os.path.join(BASE_DIR, "canvas.db"))
PORT = int(os.environ.get("PORT") or CONFIG.get("port", 8888))
UPLOADS_DIR = os.environ.get("UPLOADS_DIR") or CONFIG.get("uploads_dir", os.path.join(BASE_DIR, "uploads"))
# PostgreSQL 환경변수 오버라이드
if os.environ.get("DB_HOST"): CONFIG["db_host"] = os.environ["DB_HOST"]
if os.environ.get("DB_PORT"): CONFIG["db_port"] = int(os.environ["DB_PORT"])
if os.environ.get("DB_NAME"): CONFIG["db_name"] = os.environ["DB_NAME"]
if os.environ.get("DB_USER"): CONFIG["db_user"] = os.environ["DB_USER"]
if os.environ.get("DB_PASSWORD"): CONFIG["db_password"] = os.environ["DB_PASSWORD"]

if DB_TYPE == "postgresql" and not HAS_PSYCOPG2:
    print("[FATAL] db_type is 'postgresql' but psycopg2 is not installed.")
    print("        Install it: pip install psycopg2-binary")
    sys.exit(1)

def log(msg):
    ts = time.strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)

# ═══════════════════════════════════════
# DB Lock (동시 접속 방지)
# ═══════════════════════════════════════
import socket, platform, atexit

if DB_TYPE == "postgresql":
    LOCK_PATH = os.path.join(CONFIG.get("uploads_dir", BASE_DIR), ".server.lock")
else:
    LOCK_PATH = DB_PATH + ".lock"
LOCK_HEARTBEAT = 15  # 초

def _get_machine_id():
    hostname = socket.gethostname()
    user = os.environ.get("USER", os.environ.get("USERNAME", "unknown"))
    return f"{user}@{hostname}"

def _read_lock():
    """잠금 파일 읽기. {machine, pid, time} 반환 또는 None."""
    if not os.path.exists(LOCK_PATH):
        return None
    try:
        with open(LOCK_PATH, "r") as f:
            return json.load(f)
    except:
        return None

def _write_lock():
    """잠금 파일 생성/갱신."""
    lock_dir = os.path.dirname(LOCK_PATH)
    if lock_dir and not os.path.exists(lock_dir):
        os.makedirs(lock_dir, exist_ok=True)
    data = {
        "machine": _get_machine_id(),
        "pid": os.getpid(),
        "started": datetime.now().isoformat(),
        "heartbeat": datetime.now().isoformat()
    }
    with open(LOCK_PATH, "w") as f:
        json.dump(data, f)
    return data

def _remove_lock():
    """잠금 파일 제거."""
    try:
        if os.path.exists(LOCK_PATH):
            lock = _read_lock()
            # 내 잠금인 경우에만 제거
            if lock and lock.get("machine") == _get_machine_id() and lock.get("pid") == os.getpid():
                os.remove(LOCK_PATH)
                log("LOCK released")
    except:
        pass

def acquire_lock():
    """잠금 획득 시도. 실패 시 에러 메시지와 함께 종료."""
    me = _get_machine_id()
    lock = _read_lock()

    if lock:
        other = lock.get("machine", "?")
        hb = lock.get("heartbeat", "")

        # 같은 머신의 같은 PID → 이전 실행의 잔여 잠금 (무시)
        if other == me and lock.get("pid") == os.getpid():
            pass
        # 같은 머신의 다른 PID → 해당 프로세스가 살아있는지 확인
        elif other == me:
            try:
                os.kill(lock["pid"], 0)  # 프로세스 존재 확인
                print(f"\n╔══════════════════════════════════════════╗")
                print(f"║  ⚠️  이 PC에서 이미 서버가 실행 중입니다    ║")
                print(f"║  PID: {lock['pid']}                              ║")
                print(f"╚══════════════════════════════════════════╝")
                sys.exit(1)
            except OSError:
                log(f"LOCK stale (same machine, dead PID {lock['pid']}), overriding")
        else:
            # 다른 머신 → 하트비트 확인 (60초 초과하면 stale)
            try:
                last_hb = datetime.fromisoformat(hb)
                age = (datetime.now() - last_hb).total_seconds()
                if age < 60:
                    print(f"\n╔══════════════════════════════════════════════════╗")
                    print(f"║  ⛔ 다른 PC에서 사용 중 — 동시 접속 차단           ║")
                    print(f"║                                                  ║")
                    print(f"║  사용자: {other:<40s}║")
                    print(f"║  시작:   {lock.get('started','?'):<40s}║")
                    print(f"║  마지막: {hb:<40s}║")
                    print(f"║                                                  ║")
                    print(f"║  해당 PC에서 서버를 먼저 종료해주세요.              ║")
                    print(f"║  또는 잠금 파일을 수동 삭제:                       ║")
                    print(f"║  rm {LOCK_PATH:<44s}║")
                    print(f"╚══════════════════════════════════════════════════╝")
                    sys.exit(1)
                else:
                    log(f"LOCK stale ({other}, {age:.0f}s ago), overriding")
            except:
                log("LOCK corrupt, overriding")

    # 잠금 획득
    _write_lock()
    atexit.register(_remove_lock)
    log(f"LOCK acquired by {me} (PID {os.getpid()})")

    # 하트비트 스레드: 15초마다 잠금 파일 갱신
    def heartbeat():
        while True:
            time.sleep(LOCK_HEARTBEAT)
            try:
                lock = _read_lock()
                if lock and lock.get("pid") == os.getpid():
                    lock["heartbeat"] = datetime.now().isoformat()
                    with open(LOCK_PATH, "w") as f:
                        json.dump(lock, f)
            except:
                pass
    threading.Thread(target=heartbeat, daemon=True).start()

# ═══════════════════════════════════════
# Database (SQLite / PostgreSQL)
# ═══════════════════════════════════════
def _is_pg():
    return DB_TYPE == "postgresql"

def get_db():
    if _is_pg():
        conn = psycopg2.connect(
            host=CONFIG.get("db_host", "127.0.0.1"),
            port=CONFIG.get("db_port", 5432),
            dbname=CONFIG.get("db_name", "canvas_db"),
            user=CONFIG.get("db_user", "canvas"),
            password=CONFIG.get("db_password", ""),
            connect_timeout=10
        )
        conn.autocommit = True
        return conn
    else:
        # SQLite fallback
        db_dir = os.path.dirname(DB_PATH)
        if db_dir and not os.path.exists(db_dir):
            os.makedirs(db_dir, exist_ok=True)
        conn = sqlite3.connect(DB_PATH, check_same_thread=False, timeout=10)
        conn.row_factory = sqlite3.Row
        if '/mnt/' in DB_PATH or '\\\\' in DB_PATH:
            conn.execute("PRAGMA journal_mode=DELETE")
        else:
            conn.execute("PRAGMA journal_mode=WAL")
        return conn

def _pg_adapt_sql(sql):
    """Convert SQLite-style SQL to PostgreSQL-compatible SQL."""
    # Parameter placeholder: ? → %s
    adapted = sql.replace("?", "%s")
    # datetime('now','-30 days') → NOW() - INTERVAL '30 days'
    adapted = re.sub(
        r"datetime\(\s*'now'\s*,\s*'-(\d+)\s+days'\s*\)",
        r"NOW() - INTERVAL '\1 days'",
        adapted
    )
    return adapted

def init_db():
    db = get_db()
    if _is_pg():
        cur = db.cursor()
        pg_tables = [
            """CREATE TABLE IF NOT EXISTS project_folders (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL,
                icon TEXT DEFAULT '📂',
                color TEXT DEFAULT '',
                sort_order INTEGER DEFAULT 0,
                created TEXT NOT NULL
            )""",
            """CREATE TABLE IF NOT EXISTS projects (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                data TEXT NOT NULL,
                created TEXT NOT NULL,
                modified TEXT NOT NULL,
                favorite INTEGER DEFAULT 0,
                folder_id INTEGER
            )""",
            """CREATE TABLE IF NOT EXISTS temps (
                id SERIAL PRIMARY KEY,
                name TEXT,
                data TEXT NOT NULL,
                date TEXT NOT NULL,
                created TEXT NOT NULL
            )""",
            """CREATE TABLE IF NOT EXISTS memo_folders (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL,
                icon TEXT DEFAULT '📁',
                sort_order INTEGER DEFAULT 0,
                created TEXT NOT NULL
            )""",
            """CREATE TABLE IF NOT EXISTS memos (
                id SERIAL PRIMARY KEY,
                name TEXT,
                content TEXT DEFAULT '',
                folder_id INTEGER,
                is_temp INTEGER DEFAULT 1,
                created TEXT NOT NULL,
                modified TEXT NOT NULL,
                FOREIGN KEY (folder_id) REFERENCES memo_folders(id)
            )""",
            """CREATE TABLE IF NOT EXISTS executions (
                id TEXT PRIMARY KEY,
                project_id TEXT,
                node_id TEXT NOT NULL,
                node_name TEXT NOT NULL,
                input_raw TEXT,
                input_resolved TEXT,
                output TEXT,
                status TEXT DEFAULT 'running',
                chat_only INTEGER DEFAULT 1,
                started TEXT NOT NULL,
                finished TEXT,
                FOREIGN KEY (project_id) REFERENCES projects(id)
            )""",
            """CREATE TABLE IF NOT EXISTS conversations (
                id TEXT PRIMARY KEY,
                parent_exec_id TEXT,
                node_id TEXT NOT NULL,
                node_name TEXT NOT NULL,
                title TEXT,
                account_id INTEGER,
                created TEXT NOT NULL,
                FOREIGN KEY (parent_exec_id) REFERENCES executions(id)
            )""",
            """CREATE TABLE IF NOT EXISTS messages (
                id SERIAL PRIMARY KEY,
                conv_id TEXT NOT NULL,
                role TEXT NOT NULL,
                content TEXT NOT NULL,
                ts TEXT NOT NULL,
                FOREIGN KEY (conv_id) REFERENCES conversations(id)
            )""",
            """CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                salt TEXT NOT NULL,
                created TEXT NOT NULL
            )""",
            """CREATE TABLE IF NOT EXISTS sessions (
                token TEXT PRIMARY KEY,
                user_id INTEGER NOT NULL,
                username TEXT NOT NULL,
                created TEXT NOT NULL,
                expires TEXT NOT NULL
            )""",
            """CREATE TABLE IF NOT EXISTS trash (
                id SERIAL PRIMARY KEY,
                original_table TEXT NOT NULL,
                original_id TEXT NOT NULL,
                name TEXT,
                data TEXT NOT NULL,
                deleted_at TEXT NOT NULL
            )""",
            """CREATE TABLE IF NOT EXISTS system_settings (
                key TEXT PRIMARY KEY,
                value TEXT,
                updated TEXT
            )""",
            """CREATE TABLE IF NOT EXISTS claude_accounts (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL,
                credentials TEXT NOT NULL,
                active INTEGER DEFAULT 0,
                priority INTEGER DEFAULT 0,
                created TEXT NOT NULL
            )""",
            """CREATE TABLE IF NOT EXISTS gemini_accounts (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL,
                auth_type TEXT NOT NULL DEFAULT 'apikey',
                credentials TEXT NOT NULL,
                active INTEGER DEFAULT 0,
                priority INTEGER DEFAULT 0,
                created TEXT NOT NULL
            )""",
            """CREATE TABLE IF NOT EXISTS fav_folders (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL,
                parent_id INTEGER,
                sort_order INTEGER DEFAULT 0,
                color TEXT DEFAULT '',
                icon TEXT DEFAULT '⭐',
                created TEXT NOT NULL
            )""",
            """CREATE TABLE IF NOT EXISTS fav_items (
                id SERIAL PRIMARY KEY,
                folder_id INTEGER NOT NULL,
                memo_id INTEGER NOT NULL,
                sort_order INTEGER DEFAULT 0,
                created TEXT NOT NULL
            )""",
            """CREATE TABLE IF NOT EXISTS project_attachments (
                id SERIAL PRIMARY KEY,
                project_id TEXT NOT NULL,
                kind TEXT NOT NULL,
                target_id TEXT NOT NULL,
                sort_order INTEGER DEFAULT 0,
                created TEXT NOT NULL
            )""",
            """CREATE TABLE IF NOT EXISTS icons (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL,
                data TEXT NOT NULL,
                kind TEXT DEFAULT 'image',
                created TEXT NOT NULL
            )""",
        ]
        for ddl in pg_tables:
            cur.execute(ddl)
        # Indexes
        for idx_sql in [
            "CREATE INDEX IF NOT EXISTS idx_exec_node ON executions(node_id)",
            "CREATE INDEX IF NOT EXISTS idx_conv_node ON conversations(node_id)",
            "CREATE INDEX IF NOT EXISTS idx_msg_conv ON messages(conv_id)",
        ]:
            cur.execute(idx_sql)
        # Migration columns (PostgreSQL: check column existence before ALTER)
        alter_checks = [
            ("projects", "favorite", "ALTER TABLE projects ADD COLUMN favorite INTEGER DEFAULT 0"),
            ("projects", "folder_id", "ALTER TABLE projects ADD COLUMN folder_id INTEGER"),
            ("projects", "work_dir", "ALTER TABLE projects ADD COLUMN work_dir TEXT"),
            ("memos", "pinned", "ALTER TABLE memos ADD COLUMN pinned INTEGER DEFAULT 0"),
            ("memos", "color", "ALTER TABLE memos ADD COLUMN color TEXT DEFAULT ''"),
            ("memo_folders", "color", "ALTER TABLE memo_folders ADD COLUMN color TEXT DEFAULT ''"),
            ("memo_folders", "parent_id", "ALTER TABLE memo_folders ADD COLUMN parent_id INTEGER"),
            ("claude_accounts", "priority", "ALTER TABLE claude_accounts ADD COLUMN priority INTEGER DEFAULT 0"),
            ("conversations", "account_id", "ALTER TABLE conversations ADD COLUMN account_id INTEGER"),
            ("conversations", "project_id", "ALTER TABLE conversations ADD COLUMN project_id TEXT"),
            ("project_folders", "parent_id", "ALTER TABLE project_folders ADD COLUMN parent_id INTEGER"),
            ("messages", "chat_only", "ALTER TABLE messages ADD COLUMN chat_only INTEGER DEFAULT 1"),
            ("claude_accounts", "rate_limited_until", "ALTER TABLE claude_accounts ADD COLUMN rate_limited_until TEXT"),
            ("claude_accounts", "rate_limit_reason", "ALTER TABLE claude_accounts ADD COLUMN rate_limit_reason TEXT"),
            ("fav_items", "kind", "ALTER TABLE fav_items ADD COLUMN kind TEXT DEFAULT 'project'"),
            ("fav_items", "target_id", "ALTER TABLE fav_items ADD COLUMN target_id TEXT"),
            ("fav_folders", "kind", "ALTER TABLE fav_folders ADD COLUMN kind TEXT DEFAULT 'project'"),
        ]
        for tbl, col, alter_sql in alter_checks:
            cur.execute("""
                SELECT 1 FROM information_schema.columns
                WHERE table_name=%s AND column_name=%s
            """, (tbl, col))
            if not cur.fetchone():
                cur.execute(alter_sql)
        # NOT NULL 제거 — fav_items.memo_id를 nullable로 (project 항목 INSERT 가능하게)
        try:
            cur.execute("ALTER TABLE fav_items ALTER COLUMN memo_id DROP NOT NULL")
        except Exception:
            pass  # 이미 nullable이거나 컬럼 없음
        cur.close()
        db.close()
    else:
        # SQLite
        db.executescript("""
        CREATE TABLE IF NOT EXISTS project_folders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            icon TEXT DEFAULT '📂',
            color TEXT DEFAULT '',
            sort_order INTEGER DEFAULT 0,
            created TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS projects (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            data TEXT NOT NULL,
            created TEXT NOT NULL,
            modified TEXT NOT NULL,
            favorite INTEGER DEFAULT 0,
            folder_id INTEGER
        );
        CREATE TABLE IF NOT EXISTS temps (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            data TEXT NOT NULL,
            date TEXT NOT NULL,
            created TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS memo_folders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            icon TEXT DEFAULT '📁',
            sort_order INTEGER DEFAULT 0,
            created TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS memos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            content TEXT DEFAULT '',
            folder_id INTEGER,
            is_temp INTEGER DEFAULT 1,
            created TEXT NOT NULL,
            modified TEXT NOT NULL,
            FOREIGN KEY (folder_id) REFERENCES memo_folders(id)
        );
        CREATE TABLE IF NOT EXISTS executions (
            id TEXT PRIMARY KEY,
            project_id TEXT,
            node_id TEXT NOT NULL,
            node_name TEXT NOT NULL,
            input_raw TEXT,
            input_resolved TEXT,
            output TEXT,
            status TEXT DEFAULT 'running',
            chat_only INTEGER DEFAULT 1,
            started TEXT NOT NULL,
            finished TEXT,
            FOREIGN KEY (project_id) REFERENCES projects(id)
        );
        CREATE TABLE IF NOT EXISTS conversations (
            id TEXT PRIMARY KEY,
            parent_exec_id TEXT,
            node_id TEXT NOT NULL,
            node_name TEXT NOT NULL,
            title TEXT,
            account_id INTEGER,
            created TEXT NOT NULL,
            FOREIGN KEY (parent_exec_id) REFERENCES executions(id)
        );
        CREATE TABLE IF NOT EXISTS messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            conv_id TEXT NOT NULL,
            role TEXT NOT NULL,
            content TEXT NOT NULL,
            ts TEXT NOT NULL,
            FOREIGN KEY (conv_id) REFERENCES conversations(id)
        );
        CREATE INDEX IF NOT EXISTS idx_exec_node ON executions(node_id);
        CREATE INDEX IF NOT EXISTS idx_conv_node ON conversations(node_id);
        CREATE INDEX IF NOT EXISTS idx_msg_conv ON messages(conv_id);
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            salt TEXT NOT NULL,
            created TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS sessions (
            token TEXT PRIMARY KEY,
            user_id INTEGER NOT NULL,
            username TEXT NOT NULL,
            created TEXT NOT NULL,
            expires TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS trash (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            original_table TEXT NOT NULL,
            original_id TEXT NOT NULL,
            name TEXT,
            data TEXT NOT NULL,
            deleted_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS system_settings (
            key TEXT PRIMARY KEY,
            value TEXT,
            updated TEXT
        );
        CREATE TABLE IF NOT EXISTS claude_accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            credentials TEXT NOT NULL,
            active INTEGER DEFAULT 0,
            priority INTEGER DEFAULT 0,
            created TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS gemini_accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            auth_type TEXT NOT NULL DEFAULT 'apikey',
            credentials TEXT NOT NULL,
            active INTEGER DEFAULT 0,
            priority INTEGER DEFAULT 0,
            created TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS fav_folders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            parent_id INTEGER,
            sort_order INTEGER DEFAULT 0,
            color TEXT DEFAULT '',
            icon TEXT DEFAULT '⭐',
            created TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS fav_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            folder_id INTEGER NOT NULL,
            memo_id INTEGER NOT NULL,
            sort_order INTEGER DEFAULT 0,
            created TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS project_attachments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id TEXT NOT NULL,
            kind TEXT NOT NULL,
            target_id TEXT NOT NULL,
            sort_order INTEGER DEFAULT 0,
            created TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS icons (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            data TEXT NOT NULL,
            kind TEXT DEFAULT 'image',
            created TEXT NOT NULL
        );
        """)
        try:
            db.execute("ALTER TABLE projects ADD COLUMN favorite INTEGER DEFAULT 0")
        except sqlite3.OperationalError:
            pass
        for stmt in [
            "ALTER TABLE projects ADD COLUMN work_dir TEXT",
            "ALTER TABLE memos ADD COLUMN pinned INTEGER DEFAULT 0",
            "ALTER TABLE memos ADD COLUMN color TEXT DEFAULT ''",
            "ALTER TABLE memo_folders ADD COLUMN color TEXT DEFAULT ''",
            "ALTER TABLE memo_folders ADD COLUMN parent_id INTEGER",
            "ALTER TABLE claude_accounts ADD COLUMN priority INTEGER DEFAULT 0",
            "ALTER TABLE project_folders ADD COLUMN parent_id INTEGER",
            "ALTER TABLE messages ADD COLUMN chat_only INTEGER DEFAULT 1",
            "ALTER TABLE claude_accounts ADD COLUMN rate_limited_until TEXT",
            "ALTER TABLE claude_accounts ADD COLUMN rate_limit_reason TEXT",
            "ALTER TABLE fav_items ADD COLUMN kind TEXT DEFAULT 'project'",
            "ALTER TABLE fav_items ADD COLUMN target_id TEXT",
            "ALTER TABLE fav_folders ADD COLUMN kind TEXT DEFAULT 'project'",
            "ALTER TABLE conversations ADD COLUMN project_id TEXT",
        ]:
            try:
                db.execute(stmt)
            except sqlite3.OperationalError:
                pass
        db.commit()
        db.close()
    log(f"DB initialized ({DB_TYPE})")

init_db()
db_lock = threading.Lock()

# ═══════════════════════════════════════
# Password hashing & default user
# ═══════════════════════════════════════
def hash_password(password, salt=None):
    if salt is None:
        salt = secrets.token_hex(32)
    hashed = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt.encode('utf-8'), 100000)
    return hashed.hex(), salt

def _create_default_user():
    existing = db_exec("SELECT id FROM users WHERE username=?", ('gilhojong',), fetchone=True)
    if not existing:
        pw_hash, salt = hash_password('!!Il197119!!')
        db_exec("INSERT INTO users (username, password_hash, salt, created) VALUES (?,?,?,?)",
                ('gilhojong', pw_hash, salt, datetime.now().isoformat()))
        log("Default user created: gilhojong")

def db_exec(sql, params=(), fetch=False, fetchone=False):
    with db_lock:
        db = get_db()
        is_pg = _is_pg()
        try:
            if is_pg:
                adapted_sql = _pg_adapt_sql(sql)
                # For INSERT on SERIAL tables, add RETURNING id to get lastrowid
                need_returning = False
                if adapted_sql.strip().upper().startswith("INSERT") and not fetch and not fetchone:
                    # Tables with SERIAL id: temps, memo_folders, memos, messages
                    serial_tables = ("temps", "memo_folders", "memos", "messages", "claude_accounts",
                                     "gemini_accounts",
                                     "project_folders", "fav_folders", "fav_items", "project_attachments", "icons")
                    sql_upper = adapted_sql.upper()
                    if any(f"INTO {t.upper()}" in sql_upper for t in serial_tables):
                        if "RETURNING" not in sql_upper:
                            adapted_sql = adapted_sql.rstrip().rstrip(";") + " RETURNING id"
                            need_returning = True
                cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
                cur.execute(adapted_sql, params or ())
                if fetchone:
                    row = cur.fetchone()
                    return dict(row) if row else None
                if fetch:
                    return [dict(r) for r in cur.fetchall()]
                if need_returning:
                    row = cur.fetchone()
                    cur.close()
                    db.close()
                    return row["id"] if row else None
                cur.close()
                db.close()
                return None
            else:
                cur = db.execute(sql, params)
                if fetchone:
                    row = cur.fetchone()
                    return dict(row) if row else None
                if fetch:
                    return [dict(r) for r in cur.fetchall()]
                db.commit()
                return cur.lastrowid
        except Exception as e:
            log(f"DB ERROR: {e}\n  SQL: {sql[:200]}")
            raise
        finally:
            try:
                db.close()
            except:
                pass

# ═══════════════════════════════════════
# Manual memo initialization
# ═══════════════════════════════════════
def _create_manual_memo():
    """Create the manual memo if it doesn't exist yet."""
    MANUAL_NAME = "📖 Gil's FlowDesk 매뉴얼"
    manual_check = db_exec("SELECT id FROM memos WHERE name=?", (MANUAL_NAME,), fetchone=True)
    if manual_check:
        return
    manual_content = """# Gil's FlowDesk — JSON 구조 매뉴얼

## 프로젝트 JSON 구조

프로젝트는 다음 구조의 JSON으로 저장/내보내기/가져오기됩니다:

```json
{
  "cx": 0,           // 캔버스 X 오프셋
  "cy": 0,           // 캔버스 Y 오프셋
  "zoom": 1,         // 줌 레벨
  "cwd": "",         // 작업 디렉토리 경로
  "wfName": "",      // 프로젝트 이름
  "workflowId": "",  // 프로젝트 ID
  "nodes": [],       // 노드 배열
  "connections": [],  // 연결선 배열
  "canvasElements": [] // 캔버스 요소 (이미지 등)
}
```

## 노드 (nodes) 구조

```json
{
  "id": "n1234_1",        // 고유 ID
  "name": "Agent1",       // 노드 이름 ({{name}}으로 참조)
  "type": "agent",        // 타입: agent, memo, input, trigger
  "x": 100, "y": 100,    // 캔버스 위치
  "w": 280, "h": 240,    // 크기 (픽셀)
  "collapsed": false,     // 접힌 상태
  "chatOnly": true,       // true=대화전용(도구X), false=CLI모드(도구O)
  "inputTemplate": "",    // 입력 텍스트 ({{노드명}} 변수 사용 가능)
  "outputCapture": "",    // 실행 결과 (AI 응답)
  "systemPrompt": "",     // 시스템 프롬프트 (선택)
  "images": [],           // 첨부 이미지 경로 배열
  "status": "idle",       // idle, running, complete, error
  "history": []           // 실행 이력
}
```

## 노드 타입별 동작

| 타입 | 아이콘 | 실행 방식 |
|------|--------|-----------|
| agent | 🤖 | claude -p로 AI에 전송, 응답을 outputCapture에 저장 |
| memo | 📝 | inputTemplate을 그대로 outputCapture에 복사 (AI 호출 없음) |
| input | 📎 | memo와 동일 (데이터 입력용) |
| trigger | ⚡ | 즉시 complete, 하위 노드 실행 트리거 |

## 변수 시스템 ({{노드명}})

노드의 inputTemplate에 `{{다른노드이름}}`을 넣으면, 실행 시 해당 노드의 outputCapture 내용으로 자동 치환됩니다.

예시:
- 메모1 (type: memo): inputTemplate = "한국 체육과학 연구..."
- Agent1 (type: agent): inputTemplate = "다음 내용을 요약해: {{메모1}}"
  → 실행 시 {{메모1}}이 메모1의 내용으로 바뀌어 AI에 전송

## 연결선 (connections) 구조

```json
{
  "id": "conn_1",
  "sid": "n1234_1",  // 출발 노드 ID
  "tid": "n1234_2"   // 도착 노드 ID
}
```

연결선은 실행 순서를 결정합니다:
- 상위 노드가 모두 complete여야 하위 노드 실행 가능
- {{변수}} 치환과는 독립적 (변수명만 맞으면 연결 없이도 치환됨)

## 캔버스 요소 (canvasElements) 구조

```json
{
  "id": "ce_1",
  "type": "image",
  "x": 100, "y": 100,
  "w": 200, "h": 150,
  "src": "data:image/png;base64,...",
  "nodeType": "agent",     // 노드로 전환 시 설정
  "inputTemplate": "",
  "outputCapture": ""
}
```

## AI에게 프로젝트 생성 요청하기

이 JSON 구조를 AI에게 알려주면, 원하는 워크플로우를 JSON으로 생성받을 수 있습니다.

예시 프롬프트:
"다음 JSON 구조로 3단계 연구 분석 워크플로우를 만들어줘:
1. 메모 노드에 연구 데이터 입력
2. Agent1이 데이터 분석
3. Agent2가 분석 결과를 보고서로 작성
각 노드를 {{변수}}로 연결해줘."

생성된 JSON을 .json 파일로 저장 후 캔버스에 드래그앤드롭하면 바로 사용 가능합니다.

## 주요 API 엔드포인트

| Method | Path | 설명 |
|--------|------|------|
| POST | /api/node-exec | 노드 실행 (claude -p) |
| GET | /api/node-check?nodeId= | 실행 완료 확인 |
| POST | /api/project/save | 프로젝트 저장 |
| GET | /api/project/load?id= | 프로젝트 불러오기 |
| POST | /api/memo/save | 메모 저장 |
| GET | /api/memo/list | 메모 목록 |
| POST | /api/upload | 이미지 업로드 |

## 토큰 제한

- 한국어 기준 약 23,000자 이상의 입력은 타임아웃 가능
- 긴 텍스트는 메모를 분할하고 중간 요약 Agent를 넣어 처리
- 노드에 23,000자 초과 시 빨간 펄싱 경고 표시
"""
    # Create a manual folder
    manual_folder = db_exec("SELECT id FROM memo_folders WHERE name='📖 매뉴얼'", fetchone=True)
    if not manual_folder:
        fid = db_exec("INSERT INTO memo_folders (name, icon, color, created) VALUES (?,?,?,?)",
                      ("📖 매뉴얼", "📖", "#3B82F6", datetime.now().isoformat()))
    else:
        fid = manual_folder["id"]
    db_exec("INSERT INTO memos (name, content, folder_id, is_temp, pinned, created, modified) VALUES (?,?,?,?,?,?,?)",
            ("📖 Gil's FlowDesk 매뉴얼", manual_content, fid, 0, 1, datetime.now().isoformat(), datetime.now().isoformat()))
    log("Manual memo created")

_create_manual_memo()
_create_default_user()

# ═══════════════════════════════════════
# Claude credentials sync (DB → ~/.claude/.credentials.json)
# ═══════════════════════════════════════
def _sync_claude_credentials():
    """DB → ~/.claude/.credentials.json 동기화 (서버 시작 시).
    우선순위: claude_accounts(active=1) → 레거시 system_settings.claude_credentials."""
    try:
        # 1) New system: claude_accounts active row
        row = db_exec("SELECT credentials FROM claude_accounts WHERE active=1 LIMIT 1", fetchone=True)
        if row and row.get("credentials"):
            claude_dir = os.path.expanduser("~/.claude")
            os.makedirs(claude_dir, exist_ok=True)
            creds_path = os.path.join(claude_dir, ".credentials.json")
            with open(creds_path, "w", encoding="utf-8") as f:
                f.write(row["credentials"])
            try:
                os.chmod(creds_path, 0o600)
            except Exception:
                pass
            log("Claude credentials synced from active account")
            return
        # 2) Legacy: old system_settings key
        row = db_exec("SELECT value FROM system_settings WHERE key='claude_credentials'", fetchone=True)
        if row and row.get("value"):
            claude_dir = os.path.expanduser("~/.claude")
            os.makedirs(claude_dir, exist_ok=True)
            creds_path = os.path.join(claude_dir, ".credentials.json")
            if not os.path.exists(creds_path):
                with open(creds_path, "w", encoding="utf-8") as f:
                    f.write(row["value"])
                try:
                    os.chmod(creds_path, 0o600)
                except Exception:
                    pass
                log("Claude credentials synced from legacy system_settings")
    except Exception as e:
        log(f"Claude credentials sync failed: {e}")

# 서버 시작 시 항상 동기화 시도 (SQLite/PostgreSQL 공통)
_sync_claude_credentials()

# ═══════════════════════════════════════
# Per-account credential directories (multi-account simultaneous use)
# ═══════════════════════════════════════
def _get_account_dir(account_id):
    """각 계정의 CLAUDE_CONFIG_DIR 경로 반환. 없으면 생성."""
    base = os.path.join(tempfile.gettempdir(), "flowdesk-accts", str(account_id))
    os.makedirs(base, exist_ok=True)
    return base

def _sync_account_to_dir(account_id, credentials):
    """계정 credentials를 전용 디렉토리에 동기화.
    setup-token 토큰 (refreshToken 없음)은 파일 안 씀 (env var로만 사용)."""
    d = _get_account_dir(account_id)
    creds_path = os.path.join(d, ".credentials.json")
    # setup-token 여부 확인
    is_setup_token = False
    try:
        parsed = json.loads(credentials)
        oauth = parsed.get("claudeAiOauth", {})
        if oauth.get("accessToken") and not oauth.get("refreshToken"):
            is_setup_token = True
    except Exception:
        pass

    if is_setup_token:
        # 기존 credentials.json이 있으면 삭제 (setup-token이면 파일 쓰면 CLI가 혼동함)
        try:
            if os.path.exists(creds_path):
                os.remove(creds_path)
        except Exception:
            pass
    else:
        with open(creds_path, "w", encoding="utf-8") as f:
            f.write(credentials)
        try:
            os.chmod(creds_path, 0o600)
        except Exception:
            pass
    return d

def _sync_all_accounts():
    try:
        rows = db_exec("SELECT id, credentials FROM claude_accounts", fetch=True) or []
        for row in rows:
            _sync_account_to_dir(row["id"], row["credentials"])
        log(f"Synced {len(rows)} Claude accounts to temp dirs")
    except Exception as e:
        log(f"Sync accounts failed: {e}")

_sync_all_accounts()

# ═══════════════════════════════════════
# Gemini credentials sync (DB → ~/.gemini/)
# ═══════════════════════════════════════
def _write_gemini_creds_to_dir(target_dir, auth_type, credentials):
    """Gemini 계정 creds를 디렉토리에 기록.
    - auth_type='apikey': settings.json에 gemini-api-key 기록 (env GEMINI_API_KEY 도 설정함)
    - auth_type='oauth': oauth_creds.json + google_accounts.json + settings.json(oauth-personal)
    """
    try:
        os.makedirs(target_dir, exist_ok=True)
    except Exception:
        pass
    oauth_path = os.path.join(target_dir, "oauth_creds.json")
    google_accounts_path = os.path.join(target_dir, "google_accounts.json")
    settings_path = os.path.join(target_dir, "settings.json")
    if auth_type == "oauth":
        try:
            with open(oauth_path, "w", encoding="utf-8") as f:
                f.write(credentials)
            os.chmod(oauth_path, 0o600)
        except Exception as e:
            log(f"[GEMINI] write oauth_creds.json failed: {e}")
        # google_accounts.json: email 필요 (id_token에서 추출 시도)
        try:
            email = ""
            try:
                import base64
                parsed = json.loads(credentials)
                id_tok = parsed.get("id_token") or ""
                if id_tok and id_tok.count(".") >= 2:
                    payload = id_tok.split(".")[1]
                    payload += "=" * (-len(payload) % 4)
                    decoded = json.loads(base64.urlsafe_b64decode(payload).decode("utf-8"))
                    email = decoded.get("email", "")
            except Exception:
                pass
            with open(google_accounts_path, "w", encoding="utf-8") as f:
                json.dump({"active": email or "user@example.com", "old": []}, f)
        except Exception as e:
            log(f"[GEMINI] write google_accounts.json failed: {e}")
        # settings.json: 최신 gemini-cli 구조 (security.auth.selectedType)
        try:
            with open(settings_path, "w", encoding="utf-8") as f:
                json.dump({"security": {"auth": {"selectedType": "oauth-personal"}}}, f)
        except Exception:
            pass
    elif auth_type == "apikey":
        # apikey는 env var로 전달하지만, 최신 gemini-cli가 settings.json에서도 선택을 요구
        try:
            with open(settings_path, "w", encoding="utf-8") as f:
                json.dump({"security": {"auth": {"selectedType": "gemini-api-key"}}}, f)
        except Exception:
            pass
        # oauth_creds/google_accounts 있으면 제거
        for p in (oauth_path, google_accounts_path):
            try:
                if os.path.exists(p):
                    os.remove(p)
            except Exception:
                pass

def _sync_gemini_credentials():
    """DB active row → ~/.gemini/ 동기화 (서버 시작 시 active 계정 1개)."""
    try:
        row = db_exec(
            "SELECT auth_type, credentials FROM gemini_accounts WHERE active=1 LIMIT 1",
            fetchone=True,
        )
        if row:
            gemini_dir = os.path.expanduser("~/.gemini")
            _write_gemini_creds_to_dir(gemini_dir, row.get("auth_type") or "apikey", row.get("credentials") or "")
            log("Gemini credentials synced from active account")
    except Exception as e:
        log(f"Gemini credentials sync failed: {e}")

_sync_gemini_credentials()

def _get_gemini_account_dir(account_id):
    base = os.path.join(tempfile.gettempdir(), "flowdesk-gmini-accts", str(account_id))
    os.makedirs(base, exist_ok=True)
    return base

def _sync_gemini_account_to_dir(account_id, auth_type, credentials):
    """계정별 격리 디렉토리. HOME=<dir>, .gemini/ 하위에 기록."""
    d = _get_gemini_account_dir(account_id)
    gemini_sub = os.path.join(d, ".gemini")
    _write_gemini_creds_to_dir(gemini_sub, auth_type, credentials)
    return d

def _sync_all_gemini_accounts():
    try:
        rows = db_exec("SELECT id, auth_type, credentials FROM gemini_accounts", fetch=True) or []
        for row in rows:
            _sync_gemini_account_to_dir(row["id"], row.get("auth_type") or "apikey", row.get("credentials") or "")
        log(f"Synced {len(rows)} Gemini accounts to temp dirs")
    except Exception as e:
        log(f"Sync Gemini accounts failed: {e}")

_sync_all_gemini_accounts()

# ═══════════════════════════════════════
# Recovery slot TTL cleanup (이름 없는 임시 작업 2일 후 제거)
# ═══════════════════════════════════════
def _cleanup_stale_recovery_slots(days=2):
    """projects 테이블에서 id가 '__current_'로 시작하는 행 중 modified < now-Nd 인 행 삭제.
    이름 있는 프로젝트는 uuid 기반 id 라서 영향 없음."""
    try:
        cutoff = (datetime.now() - timedelta(days=days)).isoformat()
        rows = db_exec("SELECT id, modified FROM projects", fetch=True) or []
        removed = 0
        for r in rows:
            rid = str(r.get("id", ""))
            if not rid.startswith("__current_"):
                continue
            mod = r.get("modified") or ""
            if mod and mod < cutoff:
                db_exec("DELETE FROM projects WHERE id=?", (rid,))
                removed += 1
        if removed:
            log(f"[CLEANUP] 2일 경과 임시 복구 슬롯 {removed}개 제거")
    except Exception as e:
        log(f"[CLEANUP] recovery slots failed: {e}")

_cleanup_stale_recovery_slots(days=2)

# ═══════════════════════════════════════
# 프로젝트 작업 폴더 (Project = Folder) 정책 — 하드 고정
# ═══════════════════════════════════════
# 설계: 사용자가 경로로 고민하지 않도록 모든 경로는 아래 상수에서 자동 계산.
#   /synology/{YYYY}/{YYYYMMDD}_{이름}   ← 정식 프로젝트
#   /synology/_temp/{YYYYMMDD}_{이름|Untitled}__{8id}   ← 저장 안 한 임시 작업
#   /synology/.trash/_temp/{이름}__trashed-{ts}          ← 임시 작업의 휴지통
SYNOLOGY_CONTAINER_ROOT = "/synology"
TEMP_ROOT = os.path.join(SYNOLOGY_CONTAINER_ROOT, "_temp")
TRASH_ROOT = os.path.join(SYNOLOGY_CONTAINER_ROOT, ".trash", "_temp")
TEMP_TTL_DAYS = 2       # 임시 폴더 → 휴지통 이동
TRASH_TTL_DAYS = 5      # 휴지통 → 영구 삭제 (즉 총 2+3=5일 후 삭제)
PROJECT_SUB_PATTERN = "{year}"
PROJECT_FOLDER_PATTERN = "{date}_{name}"

# 호스트 prefix basename(e.g. "00_Gils_Project") — 사용자가 /synology/ 뒤에
# 이걸 중복해서 입력한 경우 자동으로 제거하기 위함.
def _host_prefix_basename():
    try:
        row = db_exec("SELECT value FROM system_settings WHERE key='ext_host_prefix'", fetchone=True)
        v = row.get("value") if row else None
        if v:
            base = os.path.basename(v.rstrip("/"))
            if base:
                return base
    except Exception:
        pass
    return "00_Gils_Project"

def _normalize_synology_path(path):
    """/synology/00_Gils_Project/… 식으로 호스트 prefix 를 컨테이너 경로에 중복
    지정한 경우 자동 교정. `(normalized, changed)` 반환."""
    if not path:
        return path, False
    original = path
    p = path.replace("\\", "/")
    # 이중/다중 슬래시 제거
    while "//" in p:
        p = p.replace("//", "/")
    # /synology/ 바로 뒤 호스트 basename 이 반복되면 전부 접기
    base = _host_prefix_basename()
    if base:
        needle = f"{SYNOLOGY_CONTAINER_ROOT}/{base}"
        while p == needle or p.startswith(needle + "/"):
            p = SYNOLOGY_CONTAINER_ROOT + p[len(needle):]
    p = p.rstrip("/") if p != "/" else p
    return p, (p != original)

def _seed_work_folder_defaults():
    """system_settings에 기본값 주입 + 과거 잘못 저장된 경로 일괄 정규화.
    (project_work_root / project_sub_pattern / project_folder_pattern 은 이제
    코드 상수로 강제되므로 DB 값은 참고용일 뿐이다.)"""
    defaults = {
        "project_work_root": SYNOLOGY_CONTAINER_ROOT,
        "project_sub_pattern": PROJECT_SUB_PATTERN,
        "project_folder_pattern": PROJECT_FOLDER_PATTERN,
        "ext_dsm_url": "https://gilhojong.synology.me:5001",
        "ext_quickconnect_id": "Gils-House-DB",
        "ext_container_prefix": SYNOLOGY_CONTAINER_ROOT,
        "ext_host_prefix": "/volume1/00_Gils_Project",
    }
    try:
        now = datetime.now().isoformat()
        # 기본값 주입 + 상수와 다른 경로 세팅은 강제로 상수 값으로 덮어씀
        FORCE_KEYS = {"project_work_root": SYNOLOGY_CONTAINER_ROOT,
                      "project_sub_pattern": PROJECT_SUB_PATTERN,
                      "project_folder_pattern": PROJECT_FOLDER_PATTERN}
        for k, v in defaults.items():
            row = db_exec("SELECT value FROM system_settings WHERE key=?", (k,), fetchone=True)
            if not row:
                db_exec("INSERT INTO system_settings (key, value, updated) VALUES (?,?,?)", (k, v, now))
            elif k in FORCE_KEYS and (row.get("value") or "") != FORCE_KEYS[k]:
                db_exec("UPDATE system_settings SET value=?, updated=? WHERE key=?",
                        (FORCE_KEYS[k], now, k))
                log(f"[MIGRATION] system_settings.{k} '{row.get('value')}' → '{FORCE_KEYS[k]}' (상수 강제)")
        # 이미 잘못된 경로로 저장된 projects.work_dir 전수 정규화
        try:
            rows = db_exec("SELECT id, work_dir FROM projects WHERE work_dir IS NOT NULL", fetch=True) or []
            fixed = 0
            for r in rows:
                old = r.get("work_dir") or ""
                new, changed = _normalize_synology_path(old)
                if changed:
                    db_exec("UPDATE projects SET work_dir=? WHERE id=?", (new, r["id"]))
                    log(f"[MIGRATION] project {r['id']} work_dir: {old} → {new}")
                    fixed += 1
            if fixed:
                log(f"[MIGRATION] {fixed}개 프로젝트 work_dir 정규화 완료")
        except Exception as e:
            log(f"[MIGRATION] projects work_dir 정규화 실패: {e}")
    except Exception as e:
        log(f"[WORKDIR] seed defaults failed: {e}")

_seed_work_folder_defaults()

def _sanitize_folder_name(name):
    """파일시스템 안전 폴더명.
    Windows SMB 호환을 위해 **끝 쪽의 점과 공백을 제거** (Windows/NTFS 는 후행 점 불가,
    Synology SMB 가 후행 점을 %2E 로 인코딩해서 표시하는 현상 방지)."""
    s = re.sub(r'[\\/:*?"<>|]', '_', (name or '')).strip()
    s = re.sub(r'\s+', ' ', s)
    # 양쪽 끝의 . 또는 공백 반복 제거 (중간 점은 유지)
    s = s.strip(' .')
    return s[:160] or 'Untitled'

# Windows 예약 이름 (파일시스템 호환 — NTFS/SMB 에서 쓸 수 없는 이름)
_RESERVED_WIN_NAMES = {
    "CON", "PRN", "AUX", "NUL",
    *{f"COM{i}" for i in range(0, 10)},
    *{f"LPT{i}" for i in range(0, 10)},
}
# 금지 문자 (Windows NTFS + SMB 충돌) + 실용적으로 거부할 문자 (%,&=#)
_FORBIDDEN_NAME_CHARS = set('\\/:*?"<>|')

def validate_project_name(name):
    """프로젝트 이름 검증. 유효하면 (True, None), 아니면 (False, 사용자용 오류 메시지).
    프론트와 서버 양쪽에서 호출되어 저장을 강제 차단."""
    if name is None:
        return (False, "이름이 비어있습니다.")
    raw = name
    trimmed = raw.strip()
    if not trimmed:
        return (False, "이름이 비어있거나 공백만 있습니다.")
    # 제어문자 (탭/개행/NULL 등)
    for ch in trimmed:
        if ord(ch) < 32:
            return (False, "탭·개행 등 제어문자는 사용할 수 없습니다.")
    # 앞뒤 점/공백 — Windows SMB 호환 문제
    if raw != raw.strip(" ."):
        return (False, "이름의 맨 앞이나 맨 뒤에 점(.) 또는 공백은 사용할 수 없습니다.\n(Windows·시놀로지 SMB 에서 폴더명이 %2E 로 깨집니다)")
    # 금지 문자
    bad = sorted({c for c in trimmed if c in _FORBIDDEN_NAME_CHARS})
    if bad:
        return (False, f"다음 문자는 폴더명에 쓸 수 없습니다: {' '.join(bad)}\n(허용 안 되는 기호: \\ / : * ? \" < > | )")
    # URL/HTML 에서 오해 살 수 있는 문자 — %
    if "%" in trimmed:
        return (False, "'%' 문자는 사용할 수 없습니다 (Windows 탐색기에서 URL 인코딩과 혼동).")
    # Windows 예약 이름 (date 접두사 제외한 깨끗한 이름 기준으로도 검사)
    base_for_reserved = re.sub(r'^\s*\d{4}[-\/\.]?\d{2}[-\/\.]?\d{2}[\s_\-]+', '', trimmed)
    for candidate in (trimmed, base_for_reserved):
        root = candidate.split('.', 1)[0].upper().strip()
        if root in _RESERVED_WIN_NAMES:
            return (False, f"'{candidate}' 은(는) Windows 예약 이름이라 사용할 수 없습니다.")
    # 길이 제한 (파일시스템 보수적으로)
    if len(trimmed) > 150:
        return (False, "이름이 너무 깁니다. 150자 이하로 줄여주세요.")
    return (True, None)

def _get_work_folder_settings():
    """정책이 하드 고정되었으므로 항상 상수 값 반환.
    (과거 DB 기반 설정은 _seed_work_folder_defaults 가 상수로 덮어씀.)"""
    return {
        "root": SYNOLOGY_CONTAINER_ROOT,
        "sub":  PROJECT_SUB_PATTERN,
        "folder": PROJECT_FOLDER_PATTERN,
    }

def _parse_date_from_name(name):
    """이름 앞부분 YYYYMMDD / YYYY-MM-DD / YYYY/MM/DD 감지 → (datetime, clean_name).
    감지 실패 시 (None, name)."""
    if not name:
        return (None, name)
    m = re.match(r'^\s*(\d{4})[-\/\.]?(\d{2})[-\/\.]?(\d{2})[\s_\-]+(.+?)\s*$', name)
    if not m:
        return (None, name.strip())
    try:
        dt = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        return (dt, m.group(4).strip())
    except ValueError:
        return (None, name.strip())

def _expand_pattern(pattern, name, now=None):
    """토큰 치환: {date} {year} {month} {day} {name}."""
    if not now:
        now = datetime.now()
    mapping = {
        "{date}": now.strftime("%Y%m%d"),
        "{year}": now.strftime("%Y"),
        "{month}": now.strftime("%m"),
        "{day}": now.strftime("%d"),
        "{name}": _sanitize_folder_name(name),
    }
    out = pattern or ""
    for k, v in mapping.items():
        out = out.replace(k, v)
    return out

def compute_work_dir(name, settings=None, date_override=None):
    """프로젝트 work_dir 절대경로 계산 — 정책 하드 고정.
    결과: /synology/{YYYY}/{YYYYMMDD}_{이름}
    이름 앞 YYYYMMDD_ 접두사 자동 감지, date_override 우선.
    """
    s = settings or _get_work_folder_settings()
    root = s.get("root") or SYNOLOGY_CONTAINER_ROOT
    parsed_dt, clean_name = _parse_date_from_name(name)
    effective = date_override or parsed_dt or datetime.now()
    sub = _expand_pattern(s.get("sub") or "", clean_name, effective)
    folder = _expand_pattern(s.get("folder") or "{name}", clean_name, effective) or _sanitize_folder_name(clean_name)
    parts = [root]
    if sub:
        parts.append(sub)
    parts.append(folder)
    path = os.path.normpath(os.path.join(*parts))
    # 안전망: 사용자가 어쩌다 중복 prefix를 넣었을 때도 정규화
    path, _ = _normalize_synology_path(path.replace("\\", "/"))
    return path

def ensure_work_dir(path):
    """폴더 생성(없으면). 실패 시 예외 메시지 반환. 성공 시 (True, path)."""
    if not path:
        return (False, "빈 경로")
    try:
        os.makedirs(path, exist_ok=True)
        return (True, path)
    except PermissionError:
        return (False, f"권한 없음: {path}")
    except Exception as e:
        return (False, str(e))

# ═══════════════════════════════════════
# 임시 프로젝트 폴더 (저장 안 한 작업물)
# ═══════════════════════════════════════
# 노드를 실행하는 순간 /synology/_temp/<slug>__<id>/ 에 폴더 생성.
# 매 실행마다 캔버스 JSON 을 snapshot.json 으로 기록해 복원 가능.
# 2일 경과 → /synology/.trash/_temp/ 로 이동, 추가 3일 후 영구 삭제.

# 프로젝트/세션별 임시 폴더 id → 디렉토리 경로 매핑 (프로세스 로컬)
_TEMP_DIR_CACHE = {}
_TEMP_CACHE_LOCK = threading.Lock()

def _temp_slug_from_name(name):
    safe = _sanitize_folder_name(name or "Untitled")
    date_str = datetime.now().strftime("%Y%m%d")
    return f"{date_str}_{safe}"

def get_or_create_temp_dir(session_key, name=None):
    """session_key(임시 프로젝트 식별자) 기반 폴더 반환. 없으면 생성."""
    if not session_key:
        session_key = f"anon_{uuid.uuid4().hex[:8]}"
    with _TEMP_CACHE_LOCK:
        cached = _TEMP_DIR_CACHE.get(session_key)
        if cached and os.path.isdir(cached):
            return cached
        slug = _temp_slug_from_name(name)
        short = session_key.replace("/", "_")[-8:] or uuid.uuid4().hex[:8]
        folder = os.path.join(TEMP_ROOT, f"{slug}__{short}")
        try:
            os.makedirs(folder, exist_ok=True)
            _TEMP_DIR_CACHE[session_key] = folder
            return folder
        except Exception as e:
            log(f"[TEMP] 폴더 생성 실패 {folder}: {e}")
            return None

def write_temp_snapshot(temp_dir, canvas_state):
    """임시 폴더에 캔버스 전체 스냅샷 JSON 기록. 복원용."""
    if not temp_dir or not os.path.isdir(temp_dir):
        return False
    try:
        snap_path = os.path.join(temp_dir, "snapshot.json")
        payload = {
            "savedAt": datetime.now().isoformat(),
            "kind": "temp_snapshot",
            "canvas": canvas_state,
        }
        with open(snap_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        log(f"[TEMP] snapshot 쓰기 실패 {temp_dir}: {e}")
        return False

def move_temp_to_trash(temp_dir):
    """임시 폴더를 /synology/.trash/_temp/ 로 이동. DB trash 테이블에 레코드 삽입."""
    if not temp_dir or not os.path.isdir(temp_dir):
        return False
    try:
        os.makedirs(TRASH_ROOT, exist_ok=True)
        base = os.path.basename(temp_dir.rstrip("/"))
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        target = os.path.join(TRASH_ROOT, f"{base}__trashed-{ts}")
        shutil.move(temp_dir, target)
        # DB 휴지통에도 포인터 레코드
        try:
            snap_path = os.path.join(target, "snapshot.json")
            snap_data = ""
            if os.path.exists(snap_path):
                with open(snap_path, "r", encoding="utf-8") as f:
                    snap_data = f.read()
            payload = json.dumps({"path": target, "original": temp_dir,
                                  "snapshot": snap_data[:1_000_000]}, ensure_ascii=False)
            db_exec("INSERT INTO trash (original_table, original_id, name, data, deleted_at) VALUES (?,?,?,?,?)",
                    ("temps_folder", base, base, payload, datetime.now().isoformat()))
        except Exception as e:
            log(f"[TEMP] DB trash 레코드 삽입 실패: {e}")
        # 캐시에서 제거
        with _TEMP_CACHE_LOCK:
            for k, v in list(_TEMP_DIR_CACHE.items()):
                if v == temp_dir:
                    _TEMP_DIR_CACHE.pop(k, None)
        log(f"[TEMP→TRASH] {temp_dir} → {target}")
        return True
    except Exception as e:
        log(f"[TEMP] trash 이동 실패 {temp_dir}: {e}")
        return False

def cleanup_temp_and_trash():
    """부팅 시 & 하루 1회: 오래된 임시/휴지통 정리.
    - /synology/_temp/<dir> mtime > TEMP_TTL_DAYS → 휴지통으로 이동
    - /synology/.trash/_temp/<dir> mtime > TRASH_TTL_DAYS → 영구 삭제
    """
    moved = 0
    purged = 0
    # 1) 임시 → 휴지통
    try:
        if os.path.isdir(TEMP_ROOT):
            cutoff_temp = time.time() - TEMP_TTL_DAYS * 86400
            for name in os.listdir(TEMP_ROOT):
                full = os.path.join(TEMP_ROOT, name)
                if not os.path.isdir(full):
                    continue
                try:
                    if os.path.getmtime(full) < cutoff_temp:
                        if move_temp_to_trash(full):
                            moved += 1
                except Exception:
                    continue
    except Exception as e:
        log(f"[CLEANUP] temp 스캔 실패: {e}")
    # 2) 휴지통 영구 삭제
    try:
        if os.path.isdir(TRASH_ROOT):
            cutoff_trash = time.time() - TRASH_TTL_DAYS * 86400
            for name in os.listdir(TRASH_ROOT):
                full = os.path.join(TRASH_ROOT, name)
                if not os.path.isdir(full):
                    continue
                try:
                    if os.path.getmtime(full) < cutoff_trash:
                        shutil.rmtree(full, ignore_errors=True)
                        try:
                            db_exec("DELETE FROM trash WHERE original_table='temps_folder' AND data LIKE ?",
                                    (f'%"path": "{full}"%',))
                        except Exception: pass
                        purged += 1
                        log(f"[TRASH→DELETE] {full}")
                except Exception:
                    continue
    except Exception as e:
        log(f"[CLEANUP] trash 스캔 실패: {e}")
    if moved or purged:
        log(f"[CLEANUP] 임시→휴지통 {moved}개, 휴지통→삭제 {purged}개")
    return {"moved": moved, "purged": purged}

# 부팅 시 1회 정리
try:
    cleanup_temp_and_trash()
except Exception as _e:
    log(f"[CLEANUP] 부팅 정리 실패: {_e}")

# 백그라운드 일일 정리 스레드
def _cleanup_daemon():
    while True:
        try:
            time.sleep(6 * 3600)  # 6시간마다
            cleanup_temp_and_trash()
        except Exception as e:
            log(f"[CLEANUP daemon] {e}")
            time.sleep(600)

threading.Thread(target=_cleanup_daemon, daemon=True).start()

# ═══════════════════════════════════════
# tmux
# ═══════════════════════════════════════
def run_tmux(*args):
    cmd = ["tmux"] + list(args)
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=5)
        ok = result.returncode == 0
        r = {"ok": ok, "stdout": result.stdout.strip(), "stderr": result.stderr.strip()}
        if not ok:
            log(f"  TMUX FAIL: {' '.join(args[:3])} → {result.stderr.strip()[:80]}")
        return r
    except Exception as e:
        return {"ok": False, "error": str(e)}

def get_session_info():
    info = {}
    r = run_tmux("list-sessions", "-F", "#{session_name}:#{session_windows}:#{session_attached}")
    if r["ok"] and r["stdout"]:
        info["sessions"] = []
        for line in r["stdout"].split("\n"):
            parts = line.split(":")
            if len(parts) >= 3:
                info["sessions"].append({"name": parts[0], "windows": int(parts[1]), "attached": parts[2] == "1"})
    r = run_tmux("list-panes", "-t", SESSION_NAME, "-F", "#{pane_index}:#{pane_width}x#{pane_height}:#{pane_current_command}:#{pane_current_path}")
    if r["ok"] and r["stdout"]:
        info["panes"] = []
        for line in r["stdout"].split("\n"):
            parts = line.split(":", 3)
            if len(parts) >= 4:
                info["panes"].append({"index": int(parts[0]), "size": parts[1], "command": parts[2], "path": parts[3]})
    r = run_tmux("list-windows", "-t", SESSION_NAME, "-F", "#{window_index}:#{window_name}:#{window_panes}:#{window_active}")
    if r["ok"] and r["stdout"]:
        info["windows"] = []
        for line in r["stdout"].split("\n"):
            parts = line.split(":")
            if len(parts) >= 4:
                info["windows"].append({"index": int(parts[0]), "name": parts[1], "panes": int(parts[2]), "active": parts[3] == "1"})
    return info

# ═══════════════════════════════════════
# Claude CLI
# ═══════════════════════════════════════
def _get_default_model(provider):
    """system_settings에서 provider별 기본 모델 조회."""
    try:
        key = f"default_{provider}_model"
        row = db_exec("SELECT value FROM system_settings WHERE key=?", (key,), fetchone=True)
        if row and row.get("value"):
            return row["value"]
    except Exception:
        pass
    return ""


def build_claude_cmd(prompt, opts=None):
    """claude CLI 명령 구성 — 고급 옵션 지원. prompt는 stdin으로 전달."""
    opts = opts or {}
    cmd = ["claude", "-p", "--dangerously-skip-permissions", "--effort", "max"]

    # 대화전용: Read만 허용 (이미지 인식용), 나머지 도구 차단
    chat_only = opts.get("chatOnly", True)
    if chat_only:
        cmd += ["--disallowedTools", "Bash,Write,Edit,Glob,Grep,Agent,NotebookEdit,WebFetch,WebSearch"]
        # Read는 허용 → 이미지 파일 인식 가능

    # 시스템 프롬프트
    sys_prompt = opts.get("systemPrompt", "")
    if sys_prompt:
        cmd += ["--append-system-prompt", sys_prompt]

    # 모델: 노드 지정 > 시스템 기본 > (미지정 → CLI 기본)
    model = opts.get("model") or _get_default_model("claude")
    if model:
        cmd += ["--model", model]

    # 폴백 모델
    cmd += ["--fallback-model", "sonnet"]

    # JSON 스키마 (구조화 출력)
    json_schema = opts.get("jsonSchema", "")
    if json_schema:
        cmd += ["--output-format", "json", "--json-schema", json_schema]

    # 최대 턴 수 (무한루프 방지)
    max_turns = opts.get("maxTurns", 0)
    if max_turns > 0:
        cmd += ["--max-turns", str(max_turns)]

    # 이미지 파일 경로가 있으면 프롬프트에 추가 텍스트 준비
    images = opts.get("images", [])
    final_prompt = prompt
    if images:
        img_instructions = "\n\n[첨부 이미지 파일 — Read 도구로 읽어서 분석하세요]:\n"
        img_instructions += "\n".join(f"- {img}" for img in images)
        final_prompt = prompt + img_instructions

    return cmd, final_prompt

# ═══════════════════════════════════════
# Gemini CLI (provider=gemini)
# ═══════════════════════════════════════
def build_gemini_cmd(prompt, opts=None):
    """gemini CLI 명령 구성 (v0.38.2+ 기준). prompt는 -p 인자로 직접 전달.

    Claude 대비 지원 사항:
      - model           → -m
      - systemPrompt    → 프롬프트 앞부분에 [시스템 지시사항] 블록으로 삽입
      - chatOnly        → True면 --approval-mode plan (읽기전용) / False면 yolo
      - jsonSchema      → 미지원, 프롬프트에 스키마 지시문 삽입으로 에뮬레이트
      - maxTurns        → 미지원 (무시)
      - images          → 프롬프트에 파일 경로 텍스트로 추가
      - outputJson      → -o json (stdout에서 .response 필드만 추출)
    Returns (cmd, final_prompt). final_prompt는 run_agent_safe가 기록용으로만 사용 (stdin 전달 안 함).
    """
    opts = opts or {}
    cmd = ["gemini"]
    # 모델: 노드 지정 > 시스템 기본 > 하드코딩 fallback
    model = opts.get("model") or _get_default_model("gemini") or "gemini-3.1-pro-preview"
    cmd += ["-m", model]
    # 승인 모드
    chat_only = opts.get("chatOnly", True)
    cmd += ["--approval-mode", "plan" if chat_only else "yolo"]
    # JSON 출력 기본 on (파싱 안정성)
    output_json = opts.get("outputJson", True)
    if output_json:
        cmd += ["-o", "json"]
    # 최종 프롬프트 조립
    final_prompt = prompt
    sys_prompt = opts.get("systemPrompt", "")
    if sys_prompt:
        final_prompt = f"[시스템 지시사항]\n{sys_prompt}\n\n[사용자 요청]\n{final_prompt}"
    images = opts.get("images", [])
    if images:
        final_prompt += "\n\n[첨부 이미지 파일]:\n" + "\n".join(f"- {img}" for img in images)
    json_schema = opts.get("jsonSchema", "")
    if json_schema:
        final_prompt = (
            "[중요] 아래 JSON 스키마에 완벽히 일치하는 JSON만 답변으로 출력하세요.\n"
            f"스키마:\n{json_schema}\n\n---\n\n{final_prompt}"
        )
    cmd += ["-p", final_prompt]
    return cmd, final_prompt


def parse_gemini_output(stdout, output_json=True):
    """gemini -o json stdout에서 .response 필드만 추출. 실패 시 원문."""
    if not output_json:
        return (stdout or "").strip()
    try:
        data = json.loads(stdout)
        return (data.get("response") or "").strip()
    except Exception:
        return (stdout or "").strip()


def get_gemini_env():
    """Gemini CLI 실행용 env. PATH만 공유 (계정 정보 없음)."""
    env = os.environ.copy()
    try:
        env["PATH"] = get_claude_env().get("PATH", env.get("PATH", ""))
    except Exception:
        pass
    return env


def pick_available_gemini_account():
    """active=1 계정 우선, 없으면 priority/id 가장 낮은 계정."""
    try:
        row = db_exec("SELECT id FROM gemini_accounts WHERE active=1 LIMIT 1", fetchone=True)
        if row:
            return int(row["id"])
        row = db_exec("SELECT id FROM gemini_accounts ORDER BY priority ASC, id ASC LIMIT 1", fetchone=True)
        if row:
            return int(row["id"])
    except Exception as e:
        log(f"[GEMINI] pick account failed: {e}")
    return None


def get_gemini_env_for_account(account_id):
    """계정별 격리 env: HOME=<per-account-dir>, apikey면 GEMINI_API_KEY 설정."""
    env = get_gemini_env()
    if not account_id:
        return env
    try:
        row = db_exec(
            "SELECT auth_type, credentials FROM gemini_accounts WHERE id=?",
            (int(account_id),), fetchone=True,
        )
        if not row:
            return env
        auth_type = row.get("auth_type") or "apikey"
        creds = row.get("credentials") or ""
        # HOME 격리: gemini CLI가 ~/.gemini/ 를 계정별 디렉토리에서 찾도록
        acct_dir = _sync_gemini_account_to_dir(int(account_id), auth_type, creds)
        env["HOME"] = acct_dir
        if auth_type == "apikey":
            key = creds.strip()
            # JSON으로 감싸 저장한 경우 파싱 시도
            try:
                parsed = json.loads(creds)
                if isinstance(parsed, dict):
                    key = parsed.get("apiKey") or parsed.get("api_key") or key
            except Exception:
                pass
            if key:
                env["GEMINI_API_KEY"] = key
    except Exception as e:
        log(f"[GEMINI] env for account {account_id} failed: {e}")
    return env


def run_gemini_safe(cmd_builder_fn, account_id=None, run_cwd=None, timeout=600):
    """Gemini CLI 호출. (stdout, used_account_id, fb_msg) 시그니처.
    account_id=None 이면 자동 선택. 레이트리밋 감지/로테이션은 미구현 (단순 실행).
    """
    cur_id = account_id
    if cur_id is None:
        cur_id = pick_available_gemini_account()
    env = get_gemini_env_for_account(cur_id) if cur_id else get_gemini_env()
    try:
        cmd, _ = cmd_builder_fn()
    except Exception as e:
        return ("", cur_id, f"gemini cmd build failed: {e}")
    log(f"[GEMINI] account={cur_id} model={cmd[cmd.index('-m')+1] if '-m' in cmd else '?'}")
    try:
        result = subprocess.run(
            cmd, capture_output=True, text=True,
            timeout=timeout, env=env, cwd=run_cwd,
            encoding="utf-8", errors="replace",
        )
        stdout = result.stdout or ""
        stderr = result.stderr or ""
        if result.returncode != 0 and not stdout.strip():
            tail = (stderr or "")[-300:]
            return ("", cur_id, f"gemini exit {result.returncode}: {tail}")
        parsed = parse_gemini_output(stdout, output_json=True)
        return (parsed, cur_id, "✓ gemini")
    except subprocess.TimeoutExpired:
        return ("", cur_id, "timeout")
    except FileNotFoundError:
        return ("", cur_id, "gemini CLI 미설치 (npm install -g @google/gemini-cli)")
    except Exception as e:
        log(f"[GEMINI] error: {e}")
        return ("", cur_id, str(e))


# ═══════════════════════════════════════
# Provider Dispatcher (agent 노드 provider 분기)
# ═══════════════════════════════════════
SUPPORTED_PROVIDERS = ("claude", "gemini")

def build_agent_cmd(provider, prompt, opts=None):
    if provider == "gemini":
        return build_gemini_cmd(prompt, opts)
    return build_claude_cmd(prompt, opts)

def run_agent_safe(provider, cmd_builder_fn, account_id=None, run_cwd=None, timeout=600):
    if provider == "gemini":
        return run_gemini_safe(cmd_builder_fn, account_id=account_id, run_cwd=run_cwd, timeout=timeout)
    return run_claude_safe(cmd_builder_fn, account_id, run_cwd=run_cwd, timeout=timeout)

# ═══════════════════════════════════════
# Claude login subprocess (web-based OAuth)
# ═══════════════════════════════════════
_claude_login_proc = None   # PID (int)
_claude_login_url = None
_claude_login_output = []
_claude_login_master_fd = None
_claude_login_lock = threading.Lock()

def _is_claude_proc_alive():
    global _claude_login_proc
    if not _claude_login_proc:
        return False
    try:
        wpid, _ = os.waitpid(_claude_login_proc, os.WNOHANG)
        if wpid == 0:
            return True
        return False
    except ChildProcessError:
        return False
    except Exception:
        return False

def get_claude_env():
    # nvm 또는 시스템 경로 자동 감지
    nvm_bin = os.path.expanduser("~/.nvm/versions/node/v22.22.2/bin")
    if not os.path.isdir(nvm_bin):
        # nvm 없으면 시스템 경로 사용
        import shutil
        claude_path = shutil.which("claude")
        nvm_bin = os.path.dirname(claude_path) if claude_path else "/usr/bin"
    env = os.environ.copy()
    env["PATH"] = nvm_bin + ":" + env.get("PATH", "")
    return env

# ─── Rate-limit 감지 + 자동 폴백 ────────────────────────────────────────
import re as _re_rl
RATE_LIMIT_PATTERNS = [
    # (정규식, 라벨, 쿨다운 시간(시간))
    (r"weekly\s+(usage\s+)?limit", "주간 한도", 24*7),
    (r"daily\s+(usage\s+)?limit", "일일 한도", 24),
    (r"5[\s-]?hour(?:ly)?\s+(usage\s+)?limit", "5시간 한도", 5),
    (r"hourly\s+(usage\s+)?limit", "1시간 한도", 1),
    (r"usage\s+limit\s+reached", "사용 한도", 5),
    (r"rate\s+limit\s+exceeded", "Rate limit", 1),
    (r"too\s+many\s+requests", "Too many requests", 1),
    (r"quota\s+exceeded", "할당량 초과", 24),
    (r"reached\s+your\s+(usage|message|conversation)\s+limit", "메시지 한도", 5),
]
def detect_rate_limit(text):
    """claude CLI 출력(stdout+stderr)을 검사해 rate limit 발견 시 (라벨, 쿨다운시간) 반환. 없으면 None."""
    if not text: return None
    low = text.lower()
    for pat, label, hours in RATE_LIMIT_PATTERNS:
        if _re_rl.search(pat, low):
            # "resets at HH:MM" 또는 "try again in N hours" 패턴이 있으면 그 시간 사용 시도
            m = _re_rl.search(r"try\s+again\s+in\s+(\d+)\s+hour", low)
            if m: hours = int(m.group(1))
            return (label, hours)
    return None

def mark_rate_limited(account_id, label, cooldown_hours):
    """계정에 cooldown 마킹"""
    if not account_id: return
    until = (datetime.now() + timedelta(hours=cooldown_hours)).isoformat()
    try:
        db_exec("UPDATE claude_accounts SET rate_limited_until=?, rate_limit_reason=? WHERE id=?",
                (until, label, int(account_id)))
        log(f"[RATE-LIMIT] 계정 {account_id} → {label} ({cooldown_hours}h cooldown until {until[:19]})")
    except Exception as e:
        log(f"[RATE-LIMIT] mark failed: {e}")

def is_account_available(account_id):
    """rate_limited_until 이 미래면 사용 불가"""
    if not account_id: return True
    try:
        row = db_exec("SELECT rate_limited_until FROM claude_accounts WHERE id=?", (int(account_id),), fetchone=True)
        if not row or not row.get("rate_limited_until"): return True
        until_str = row["rate_limited_until"]
        until = datetime.fromisoformat(until_str)
        if datetime.now() >= until:
            # 만료됨 → 자동 해제
            db_exec("UPDATE claude_accounts SET rate_limited_until=NULL, rate_limit_reason=NULL WHERE id=?", (int(account_id),))
            return True
        return False
    except Exception as e:
        log(f"[RATE-LIMIT] check failed: {e}")
        return True

def pick_available_account(exclude_ids=None):
    """rotation_mode 따라 사용 가능한 계정 1개 골라서 반환. 없으면 None."""
    exclude_ids = set(exclude_ids or [])
    try:
        mode_row = db_exec("SELECT value FROM system_settings WHERE key='claude_rotation_mode'", fetchone=True)
        mode = mode_row["value"] if mode_row and mode_row.get("value") else "round-robin"
        accounts = db_exec("SELECT id FROM claude_accounts ORDER BY priority ASC, id ASC", fetch=True) or []
        if not accounts: return None
        # manual 모드면 active만 시도
        if mode == "manual":
            active = db_exec("SELECT id FROM claude_accounts WHERE active=1 LIMIT 1", fetchone=True)
            if active and active["id"] not in exclude_ids and is_account_available(active["id"]):
                return active["id"]
            return None
        # round-robin / sequential: 사용 가능한 첫 계정
        for a in accounts:
            if a["id"] in exclude_ids: continue
            if is_account_available(a["id"]):
                return a["id"]
        return None
    except Exception as e:
        log(f"[RATE-LIMIT] pick failed: {e}")
        return None


def run_claude_safe(cmd_builder_fn, account_id, run_cwd=None, timeout=600, max_fallback=3):
    """Claude CLI 호출 + rate-limit 자동 폴백.
    cmd_builder_fn() → (cmd, final_prompt). 한도 감지되면 다른 계정으로 재시도.
    Returns (stdout, used_account_id, fallback_log)"""
    tried = set()
    last_err = None
    cur_id = account_id
    for attempt in range(max_fallback + 1):
        if not cur_id or cur_id in tried:
            cur_id = pick_available_account(exclude_ids=tried)
            if not cur_id:
                return ("", None, f"❌ 사용 가능한 계정 없음 ({last_err or '전부 한도 초과'})")
        tried.add(cur_id)
        env = get_claude_env_for_account(cur_id)
        try:
            cmd, final_prompt = cmd_builder_fn()
        except Exception as e:
            return ("", cur_id, f"cmd build failed: {e}")
        log(f"[CLAUDE] attempt={attempt+1} account={cur_id}")
        try:
            result = subprocess.run(cmd, input=final_prompt, capture_output=True, text=True,
                                    timeout=timeout, env=env, cwd=run_cwd)
            combined = (result.stdout or "") + "\n" + (result.stderr or "")
            rl = detect_rate_limit(combined)
            if rl:
                label, hours = rl
                mark_rate_limited(cur_id, label, hours)
                last_err = f"계정 {cur_id} → {label} ({hours}h)"
                log(f"[CLAUDE] rate-limited: {last_err} → 다른 계정 폴백")
                cur_id = None
                continue
            return (result.stdout.strip(), cur_id, f"✓ 계정 {cur_id}" + (f" (이전 {len(tried)-1}개 한도 초과로 폴백)" if len(tried)>1 else ""))
        except subprocess.TimeoutExpired:
            return ("", cur_id, "timeout")
        except Exception as e:
            log(f"[CLAUDE] error: {e}")
            return ("", cur_id, str(e))
    return ("", cur_id, f"max retry ({last_err})")


def get_claude_env_for_account(account_id=None):
    """특정 계정의 env 반환.
    - 전체 OAuth (refreshToken 있음): credentials.json 파일 사용
    - setup-token (refreshToken 없음): CLAUDE_CODE_OAUTH_TOKEN env var만 사용 (파일 없는 빈 dir)
    - account_id 없으면 active=1 계정으로 폴백
    """
    env = get_claude_env()
    # account_id 없으면 활성 계정으로 폴백
    if not account_id:
        try:
            active = db_exec("SELECT id FROM claude_accounts WHERE active=1 LIMIT 1", fetchone=True)
            if active and active.get("id"):
                account_id = active["id"]
        except Exception:
            pass
    if not account_id:
        return env
    try:
        row = db_exec("SELECT credentials FROM claude_accounts WHERE id=?", (int(account_id),), fetchone=True)
        if not row or not row.get("credentials"):
            return env
        try:
            data = json.loads(row["credentials"])
        except Exception:
            return env
        oauth = data.get("claudeAiOauth", {})
        token = oauth.get("accessToken", "")
        refresh = oauth.get("refreshToken", "")
        if token and not refresh:
            # setup-token 토큰 → env var만 사용, credentials.json 없는 빈 dir
            empty_dir = os.path.join(tempfile.gettempdir(), "flowdesk-envonly", str(account_id))
            os.makedirs(empty_dir, exist_ok=True)
            # 기존 credentials.json이 있으면 삭제 (과거 잘못 저장된 것 포함)
            try:
                stale = os.path.join(empty_dir, ".claude", ".credentials.json")
                if os.path.exists(stale):
                    os.remove(stale)
            except Exception:
                pass
            env["CLAUDE_CODE_OAUTH_TOKEN"] = token
            env["CLAUDE_CONFIG_DIR"] = empty_dir
            env["HOME"] = empty_dir
        else:
            # 전체 OAuth → credentials.json 파일 기반
            account_dir = _get_account_dir(int(account_id))
            env["CLAUDE_CONFIG_DIR"] = account_dir
            env["HOME"] = account_dir
    except Exception as e:
        log(f"Failed to set account env: {e}")
    return env

# ═══════════════════════════════════════
# HTTP Handler
# ═══════════════════════════════════════
class TmuxHandler(SimpleHTTPRequestHandler):

    def do_GET(self):
        parsed = urlparse(self.path)
        params = parse_qs(parsed.query)
        p = parsed.path

        # Auth check endpoint (no auth required)
        if p == "/api/auth/check":
            self._json(self._auth_check())
            return

        # Auth middleware for API routes
        if p.startswith("/api/") and p not in ("/api/auth/login", "/api/auth/check"):
            if not self._check_auth():
                self._json({"ok": False, "error": "unauthorized"}, 401)
                return

        if p == "/api/status": self._json(get_session_info())
        elif p == "/api/node-check": self._json(self._node_check(params))
        elif p == "/api/chat-check": self._json(self._chat_check(params))
        elif p == "/api/chat-history": self._json(self._chat_history(params))
        elif p == "/api/project/list": self._json(self._project_list())
        elif p == "/api/project/list-meta": self._json(self._project_list_meta())
        elif p == "/api/project/preview-work-dir": self._json(self._project_preview_work_dir(params))
        elif p == "/api/project/scan-unregistered": self._json(self._project_scan_unregistered(params))
        elif p == "/api/project/load": self._json(self._project_load(params))
        elif p == "/api/temp/list": self._json(self._temp_list())
        elif p == "/api/temp/load": self._json(self._temp_load(params))
        elif p == "/api/memo/list": self._json(self._memo_list(params))
        elif p == "/api/memo/get": self._json(self._memo_get(params))
        elif p == "/api/folder/list": self._json(self._folder_list())
        elif p == "/api/project-folder/list": self._json(self._project_folder_list())
        elif p == "/api/exec/list": self._json(self._exec_list(params))
        elif p == "/api/conv/list": self._json(self._conv_list(params))
        elif p == "/api/conv/messages": self._json(self._conv_messages(params))
        elif p == "/api/scratchpad/load": self._json(self._scratchpad_load())
        elif p == "/api/memo/pinned": self._json(self._memo_pinned())
        elif p == "/api/fav/folders": self._json(self._fav_folders(params))
        elif p == "/api/fav/items": self._json(self._fav_items(params))
        elif p == "/api/proj-attach/list": self._json(self._proj_attach_list(params))
        elif p == "/api/icons/list": self._json(self._icons_list())
        elif p == "/api/state/load": self._json(self._state_load(params))
        elif p == "/api/pane-content": self._json(self._pane_content(params))
        elif p == "/api/pane-prompt-check": self._json(self._pane_prompt(params))
        elif p == "/api/trash/list": self._json(self._trash_list())
        elif p == "/api/temp/folder-list": self._json(self._temp_folder_list())
        elif p == "/api/temp/folder-snapshot": self._json(self._temp_folder_snapshot(params))
        elif p == "/api/doc/list": self._json(self._doc_list(params))
        elif p == "/api/doc/versions": self._json(self._doc_versions(params))
        elif p == "/api/project/files": self._json(self._project_files(params))
        elif p == "/api/project/file":
            self._project_file_serve(params)
            return
        elif p == "/api/settings/get": self._json(self._settings_get())
        elif p == "/api/auth/health": self._json(self._auth_health(params))
        elif p == "/api/claude/accounts/list": self._json(self._claude_accounts_list())
        elif p == "/api/claude/login/status": self._json(self._claude_login_status())
        elif p == "/api/gemini/accounts/list": self._json(self._gemini_accounts_list())
        elif p == "/api/fs/browse": self._json(self._browse_path(params))
        elif p == "/api/fs/browse-system": self._json(self._fs_browse_system(params))
        elif p == "/api/fs/download":
            self._fs_download(params)
            return
        elif p.startswith("/uploads/"):
            # uploads 디렉토리가 외부 경로일 수 있으므로 직접 서빙
            file_path = os.path.join(UPLOADS_DIR, p[len("/uploads/"):])
            if os.path.isfile(file_path):
                import mimetypes
                mime = mimetypes.guess_type(file_path)[0] or 'application/octet-stream'
                self.send_response(200)
                self.send_header("Content-Type", mime)
                self.send_header("Content-Length", os.path.getsize(file_path))
                self.end_headers()
                with open(file_path, "rb") as f:
                    self.wfile.write(f.read())
            else:
                self.send_error(404, "File not found")
        elif p == "/":
            self.path = "/canvas.html"
            return SimpleHTTPRequestHandler.do_GET(self)
        else:
            return SimpleHTTPRequestHandler.do_GET(self)

    def do_POST(self):
        parsed = urlparse(self.path)
        content_len = int(self.headers.get("Content-Length", 0))
        body = json.loads(self.rfile.read(content_len)) if content_len > 0 else {}
        p = parsed.path

        # Auth login endpoint (no auth required)
        if p == "/api/auth/login":
            self._handle_auth_login(body)
            return
        if p == "/api/auth/logout":
            self._handle_auth_logout(body)
            return

        # Auth middleware for all other POST API routes
        if p.startswith("/api/"):
            if not self._check_auth():
                self._json({"ok": False, "error": "unauthorized"}, 401)
                return

        handlers = {
            "/api/node-exec": self._node_exec,
            "/api/chat": self._chat_send,
            "/api/chat/fork": self._chat_fork,
            "/api/project/save": self._project_save,
            "/api/project/adopt": self._project_adopt,
            "/api/media/download": self._media_download,
            "/api/media/extract-frame": self._media_extract_frame,
            "/api/youtube/search": self._youtube_search,
            "/api/media/subtitle": self._media_subtitle,
            "/api/project/delete": self._project_delete,
            "/api/state/save": self._state_save,
            "/api/temp/save": self._temp_save,
            "/api/temp/delete": self._temp_delete,
            "/api/memo/save": self._memo_save,
            "/api/memo/delete": self._memo_delete,
            "/api/folder/save": self._folder_save,
            "/api/folder/delete": self._folder_delete,
            "/api/scratchpad/save": self._scratchpad_save,
            "/api/memo/pin": self._memo_pin,
            "/api/fav/folder/save": self._fav_folder_save,
            "/api/fav/folder/delete": self._fav_folder_delete,
            "/api/fav/item/add": self._fav_item_add,
            "/api/fav/item/remove": self._fav_item_remove,
            "/api/project/date-folder": self._project_date_folder,
            "/api/proj-attach/add": self._proj_attach_add,
            "/api/proj-attach/remove": self._proj_attach_remove,
            "/api/icon/save": self._icon_save,
            "/api/icon/delete": self._icon_delete,
            "/api/project/star": self._project_star,
            "/api/project/move": self._project_move,
            "/api/project-folder/save": self._project_folder_save,
            "/api/project-folder/delete": self._project_folder_delete,
            "/api/upload": self._upload_file,
            "/api/pdf-split": self._pdf_split,
            "/api/list-pdf-images": self._list_pdf_images,
            "/api/reset-session": self._reset_session,
            "/api/add-pane": self._add_pane,
            "/api/setup-session": self._setup_session,
            "/api/split": self._split,
            "/api/send-command": self._send_command,
            "/api/preset": self._preset,
            "/api/layout": self._layout,
            "/api/kill-pane": self._kill_pane,
            "/api/new-window": self._new_window,
            "/api/select-window": self._select_window,
            "/api/trash/restore": self._trash_restore,
            "/api/trash/delete": self._trash_delete,
            "/api/trash/empty": self._trash_empty,
            "/api/settings/set": self._settings_set,
            "/api/settings/delete": self._settings_delete,
            "/api/claude/accounts/save": self._claude_account_save,
            "/api/claude/accounts/delete": self._claude_account_delete,
            "/api/claude/accounts/activate": self._claude_account_activate,
            "/api/claude/accounts/test": self._claude_account_test,
            "/api/claude/accounts/next": self._claude_next_account,
            "/api/claude/accounts/next-preview": self._claude_next_preview,
            "/api/claude/accounts/unlock": self._claude_account_unlock,
            "/api/claude/accounts/reorder": self._claude_accounts_reorder,
            "/api/claude/login/start": self._claude_login_start,
            "/api/claude/login/submit": self._claude_login_submit,
            "/api/claude/login/cancel": self._claude_login_cancel,
            "/api/gemini/accounts/save": self._gemini_account_save,
            "/api/gemini/accounts/delete": self._gemini_account_delete,
            "/api/gemini/accounts/activate": self._gemini_account_activate,
            "/api/gemini/accounts/test": self._gemini_account_test,
            "/api/fs/mkdir": self._fs_mkdir,
            "/api/fs/mkdir-system": self._fs_mkdir_system,
            "/api/fs/delete": self._fs_delete,
            "/api/fs/trash-list": self._fs_trash_list,
            "/api/fs/trash-restore": self._fs_trash_restore,
            "/api/fs/trash-delete": self._fs_trash_delete,
            "/api/fs/trash-empty": self._fs_trash_empty,
            "/api/temp/folder-restore": self._temp_folder_restore,
            "/api/temp/folder-promote": self._temp_folder_promote,
            "/api/temp/folder-delete": self._temp_folder_delete_now,
            "/api/temp/cleanup-now": self._temp_cleanup_now,
            "/api/sheet/import": self._sheet_import,
            "/api/sheet/export-xlsx": self._sheet_export_xlsx,
            "/api/doc/read": self._doc_read,
            "/api/doc/write": self._doc_write,
            "/api/project/file/upload": self._project_file_upload,
            "/api/project/mkdir": self._project_mkdir,
            "/api/project/copy-files": self._project_copy_files,
            "/api/project/move-files": self._project_move_files,
            "/api/doc/delete": self._doc_delete,
            "/api/doc/restore": self._doc_restore,
            "/api/doc/milestone": self._doc_milestone,
            "/api/doc/delete-version": self._doc_delete_version,
            "/api/doc/hwp-to-docx": self._doc_hwp_to_docx,
        }
        handler = handlers.get(p)
        if handler:
            self._json(handler(body))
        else:
            self._json({"ok": False, "error": "not found"}, 404)

    def do_OPTIONS(self):
        self.send_response(200)
        origin = self._get_cors_origin()
        self.send_header("Access-Control-Allow-Origin", origin)
        if origin != "*":
            self.send_header("Access-Control-Allow-Credentials", "true")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    # ── Node Execution (캔버스 워크플로우) ──

    def _node_exec(self, body):
        node_id = body.get("nodeId", "")
        node_name = body.get("nodeName", node_id)
        prompt = body.get("prompt", "")
        cwd = body.get("cwd", "")
        chat_only = body.get("chatOnly", True)
        project_id = body.get("projectId", "")
        account_id = body.get("accountId")
        if not node_id or not prompt:
            return {"ok": False, "error": "nodeId and prompt required"}

        exec_id = str(uuid.uuid4())[:8]
        out_file = f"/tmp/node_{node_id}_output.txt"
        done_file = f"/tmp/node_{node_id}_done"
        for f in [out_file, done_file]:
            if os.path.exists(f): os.remove(f)

        # DB에 실행 기록 생성 (project_id 비어있으면 NULL — FK 제약 회피)
        db_exec("INSERT INTO executions (id, project_id, node_id, node_name, input_raw, input_resolved, chat_only, started, status) VALUES (?,?,?,?,?,?,?,?,?)",
                (exec_id, project_id or None, node_id, node_name, body.get("inputRaw", ""), prompt, 1 if chat_only else 0, datetime.now().isoformat(), "running"))
        log(f"EXEC [{exec_id}] node={node_name} account={account_id} prompt={len(prompt)}자 chatOnly={chat_only}")
        if len(prompt) > 23000:
            log(f"EXEC [{exec_id}] WARNING: 프롬프트 {len(prompt)}자 — 23,000자 초과, 타임아웃 가능성 높음")

        def run():
            try:
                # run_cwd 결정 우선순위:
                # 1) body.cwd (명시적 지정)
                # 2) projects.work_dir (저장된 정식 프로젝트)
                # 3) /synology/_temp/<date>_<projname>__<id>/ (임시 프로젝트 — 저장 안 됨)
                run_cwd = None
                if cwd and os.path.isdir(cwd):
                    run_cwd = cwd
                else:
                    try:
                        wf_id = (body.get("projectId") or "").strip()
                        if wf_id:
                            wd_row = db_exec("SELECT work_dir FROM projects WHERE id=?", (wf_id,), fetchone=True)
                            if wd_row and wd_row.get("work_dir") and os.path.isdir(wd_row["work_dir"]):
                                run_cwd = wd_row["work_dir"]
                                log(f"EXEC [{exec_id}] project work_dir: {run_cwd}")
                        # 프로젝트 저장 안 되어 있으면 임시 폴더 사용
                        if not run_cwd:
                            temp_key = (body.get("tempKey") or wf_id or body.get("tabId") or f"anon_{node_id}")
                            temp_name = body.get("projectName") or "Untitled"
                            tdir = get_or_create_temp_dir(temp_key, temp_name)
                            if tdir:
                                run_cwd = tdir
                                log(f"EXEC [{exec_id}] temp cwd: {run_cwd}")
                                # 캔버스 스냅샷이 body.canvasState 로 넘어오면 기록
                                canvas_state = body.get("canvasState")
                                if canvas_state:
                                    write_temp_snapshot(tdir, canvas_state)
                    except Exception as e:
                        log(f"EXEC [{exec_id}] cwd resolve failed: {e}")
                        run_cwd = None
                provider = (body.get("provider") or "claude").lower()
                if provider not in SUPPORTED_PROVIDERS:
                    provider = "claude"
                def _build():
                    return build_agent_cmd(provider, prompt, {
                        "chatOnly": chat_only,
                        "systemPrompt": body.get("systemPrompt", ""),
                        "jsonSchema": body.get("jsonSchema", ""),
                        "maxTurns": body.get("maxTurns", 0),
                        "images": body.get("images", []),
                        "model": body.get("model", ""),
                    })
                output, used_acc, fb_msg = run_agent_safe(provider, _build, account_id, run_cwd=run_cwd, timeout=600)
                if used_acc and used_acc != account_id:
                    log(f"EXEC [{exec_id}] 계정 폴백: {account_id} → {used_acc}")
                    # node에 새 계정 영구 배정 (다음에도 같은 계정 쓰게)
                    # node_id는 frontend가 관리하므로 여기서 DB 업뎃 안 함, 응답에만 표시
                if not output:
                    log(f"EXEC [{exec_id}] FAILED: {fb_msg}")
                    output = f"(❌ 실행 실패: {fb_msg})"
                log(f"EXEC [{exec_id}] done! {len(output)}자 [{fb_msg}]")
                with open(out_file, "w", encoding="utf-8") as f: f.write(output)
                with open(done_file, "w") as f: f.write("done")
                db_exec("UPDATE executions SET output=?, status='complete', finished=? WHERE id=?",
                        (output, datetime.now().isoformat(), exec_id))
            except Exception as e:
                log(f"EXEC [{exec_id}] ERROR: {e}")
                err_msg = f"(오류: {e})"
                with open(out_file, "w", encoding="utf-8") as f: f.write(err_msg)
                with open(done_file, "w") as f: f.write("error")
                db_exec("UPDATE executions SET output=?, status='error', finished=? WHERE id=?",
                        (err_msg, datetime.now().isoformat(), exec_id))

        threading.Thread(target=run, daemon=True).start()
        return {"ok": True, "execId": exec_id, "nodeId": node_id}

    def _node_check(self, params):
        node_id = params.get("nodeId", [""])[0]
        done_file = f"/tmp/node_{node_id}_done"
        out_file = f"/tmp/node_{node_id}_output.txt"
        if os.path.exists(done_file):
            try:
                with open(out_file, "r", encoding="utf-8") as f: output = f.read().strip()
                return {"ok": True, "done": True, "output": output}
            except: return {"ok": True, "done": True, "output": ""}
        return {"ok": True, "done": False}

    # ── Chat (대화 세션) ──

    def _chat_send(self, body):
        conv_id = body.get("convId", "")
        message = body.get("message", "")
        cwd = body.get("cwd", "")
        chat_only = body.get("chatOnly", True)
        node_id = body.get("nodeId", "")
        node_name = body.get("nodeName", "")
        project_id = (body.get("projectId") or "").strip()
        if not message: return {"ok": False, "error": "message required"}

        # account_id 결정: body에 있으면 사용, 없으면 conversation에서 조회, 그것도 없으면 기본
        account_id = body.get("accountId")
        chat_provider = (body.get("provider") or "claude").lower()
        if chat_provider not in SUPPORTED_PROVIDERS:
            chat_provider = "claude"

        # 대화 세션 없으면 생성
        if not conv_id:
            conv_id = str(uuid.uuid4())[:8]
            # 새 conversation 생성 시 body의 accountId 사용 (없으면 provider별 자동 배정)
            if not account_id:
                try:
                    if chat_provider == "gemini":
                        account_id = pick_available_gemini_account()
                    else:
                        nxt = self._claude_next_account({})
                        if nxt.get("ok"): account_id = nxt.get("accountId")
                except: pass
            db_exec("INSERT INTO conversations (id, node_id, node_name, title, account_id, project_id, created) VALUES (?,?,?,?,?,?,?)",
                    (conv_id, node_id, node_name, message[:30], account_id, project_id or None, datetime.now().isoformat()))
        else:
            # 기존 conversation → 저장된 account_id 조회
            if not account_id:
                conv = db_exec("SELECT account_id FROM conversations WHERE id=?", (conv_id,), fetchone=True)
                if conv and conv.get("account_id"):
                    account_id = conv["account_id"]

        # 유저 메시지 저장 (chat_only 모드도 함께 — 각 메시지가 어느 모드로 처리됐는지 추적)
        db_exec("INSERT INTO messages (conv_id, role, content, ts, chat_only) VALUES (?,?,?,?,?)",
                (conv_id, "user", message, datetime.now().isoformat(), 1 if chat_only else 0))

        # 이전 대화 로드
        rows = db_exec("SELECT role, content FROM messages WHERE conv_id=? ORDER BY id", (conv_id,), fetch=True)
        prompt_parts = []
        for r in rows:
            prefix = "[User]" if r["role"] == "user" else "[Assistant]"
            prompt_parts.append(f"{prefix}: {r['content']}")
        full_prompt = "아래는 대화 기록입니다. 마지막 [User]에 답변하세요.\n\n" + "\n\n".join(prompt_parts)

        done_file = f"/tmp/chat_{conv_id}_done"
        if os.path.exists(done_file): os.remove(done_file)

        log(f"CHAT [{conv_id}] msg={len(message)}자 history={len(rows)}")

        def run():
            try:
                # cwd 결정: 1) body.cwd → 2) project_id 의 work_dir → 3) 임시 폴더
                run_cwd = cwd if cwd and os.path.isdir(cwd) else None
                if not run_cwd:
                    # conversations 레코드에서 project_id 재조회 (기존 대화는 업데이트 안 했으므로)
                    pid = project_id
                    if not pid:
                        try:
                            cr = db_exec("SELECT project_id FROM conversations WHERE id=?", (conv_id,), fetchone=True)
                            if cr and cr.get("project_id"):
                                pid = cr["project_id"]
                        except Exception: pass
                    if pid:
                        try:
                            wd_row = db_exec("SELECT work_dir FROM projects WHERE id=?", (pid,), fetchone=True)
                            if wd_row and wd_row.get("work_dir") and os.path.isdir(wd_row["work_dir"]):
                                run_cwd = wd_row["work_dir"]
                                log(f"CHAT [{conv_id}] project work_dir: {run_cwd}")
                        except Exception: pass
                    # 그래도 없으면 임시 폴더 (대화 id 를 세션 키로)
                    if not run_cwd:
                        temp_key = pid or conv_id
                        temp_name = body.get("projectName") or node_name or "Chat"
                        tdir = get_or_create_temp_dir(temp_key, temp_name)
                        if tdir:
                            run_cwd = tdir
                            log(f"CHAT [{conv_id}] temp cwd: {run_cwd}")
                provider = (body.get("provider") or "claude").lower()
                if provider not in SUPPORTED_PROVIDERS:
                    provider = "claude"
                def _build():
                    return build_agent_cmd(provider, full_prompt, {
                        "chatOnly": chat_only,
                        "systemPrompt": body.get("systemPrompt", ""),
                        "images": body.get("images", []),
                        "model": body.get("model", ""),
                    })
                reply, used_acc, fb_msg = run_agent_safe(provider, _build, account_id, run_cwd=run_cwd, timeout=600)
                if used_acc and used_acc != account_id:
                    log(f"CHAT [{conv_id}] 계정 폴백 {account_id} → {used_acc}")
                    # 대화에 새 계정 영구 배정 (다음 메시지부터 같은 계정 사용)
                    try: db_exec("UPDATE conversations SET account_id=? WHERE id=?", (used_acc, conv_id))
                    except: pass
                if not reply:
                    reply = f"(❌ 응답 실패: {fb_msg})"
                log(f"CHAT [{conv_id}] reply={len(reply)}자 [{fb_msg}]")
                db_exec("INSERT INTO messages (conv_id, role, content, ts, chat_only) VALUES (?,?,?,?,?)",
                        (conv_id, "assistant", reply, datetime.now().isoformat(), 1 if chat_only else 0))
                with open(done_file, "w") as f: f.write(reply)
            except Exception as e:
                log(f"CHAT [{conv_id}] ERROR: {e}")
                db_exec("INSERT INTO messages (conv_id, role, content, ts, chat_only) VALUES (?,?,?,?,?)",
                        (conv_id, "assistant", f"(오류: {e})", datetime.now().isoformat(), 1 if chat_only else 0))
                with open(done_file, "w") as f: f.write(f"(오류: {e})")

        threading.Thread(target=run, daemon=True).start()
        return {"ok": True, "convId": conv_id}

    def _chat_check(self, params):
        conv_id = params.get("convId", [""])[0]
        done_file = f"/tmp/chat_{conv_id}_done"
        if os.path.exists(done_file):
            with open(done_file, "r", encoding="utf-8") as f: reply = f.read().strip()
            return {"ok": True, "done": True, "reply": reply}
        return {"ok": True, "done": False}

    def _chat_history(self, params):
        conv_id = params.get("convId", [""])[0]
        if not conv_id: return {"ok": True, "messages": []}
        rows = db_exec("SELECT role, content, ts FROM messages WHERE conv_id=? ORDER BY id", (conv_id,), fetch=True)
        return {"ok": True, "messages": rows}

    def _chat_fork(self, body):
        """실행 결과에서 대화 분기 생성"""
        exec_id = body.get("execId", "")
        node_id = body.get("nodeId", "")
        node_name = body.get("nodeName", "")
        if not exec_id and not node_id:
            return {"ok": False, "error": "execId or nodeId required"}

        conv_id = str(uuid.uuid4())[:8]
        title = f"{node_name} 분기"

        # 실행 기록에서 input/output 가져와서 초기 메시지로
        if exec_id:
            ex = db_exec("SELECT * FROM executions WHERE id=?", (exec_id,), fetchone=True)
            if ex:
                node_id = ex["node_id"]
                node_name = ex["node_name"]
                title = f"{node_name} @{ex['started'][:16]}"
        # 부모 노드의 account_id를 상속 (body에 있으면 사용)
        account_id = body.get("accountId")
        project_id = (body.get("projectId") or "").strip() or None
        db_exec("INSERT INTO conversations (id, parent_exec_id, node_id, node_name, title, account_id, project_id, created) VALUES (?,?,?,?,?,?,?,?)",
                (conv_id, exec_id, node_id, node_name, title, account_id, project_id, datetime.now().isoformat()))

        # 실행 기록의 input/output을 초기 대화로 삽입
        if exec_id:
            ex = db_exec("SELECT * FROM executions WHERE id=?", (exec_id,), fetchone=True)
            if ex:
                if ex.get("input_resolved"):
                    db_exec("INSERT INTO messages (conv_id, role, content, ts) VALUES (?,?,?,?)",
                            (conv_id, "user", ex["input_resolved"], ex["started"]))
                if ex.get("output"):
                    db_exec("INSERT INTO messages (conv_id, role, content, ts) VALUES (?,?,?,?)",
                            (conv_id, "assistant", ex["output"], ex.get("finished", ex["started"])))
        log(f"FORK conv={conv_id} from exec={exec_id} node={node_name}")
        return {"ok": True, "convId": conv_id}

    # ── Execution History ──

    def _exec_list(self, params):
        node_id = params.get("nodeId", [""])[0]
        if node_id:
            rows = db_exec("SELECT id, node_name, status, started, finished, substr(input_resolved,1,50) as input_preview, substr(output,1,50) as output_preview FROM executions WHERE node_id=? ORDER BY started DESC LIMIT 50", (node_id,), fetch=True)
        else:
            rows = db_exec("SELECT id, node_name, status, started, finished, substr(input_resolved,1,50) as input_preview, substr(output,1,50) as output_preview FROM executions ORDER BY started DESC LIMIT 100", fetch=True)
        return {"ok": True, "executions": rows}

    # ── Conversation List ──

    def _conv_list(self, params):
        node_id = params.get("nodeId", [""])[0]
        if node_id:
            rows = db_exec("SELECT c.id, c.title, c.node_name, c.created, c.parent_exec_id, (SELECT COUNT(*) FROM messages WHERE conv_id=c.id) as msg_count FROM conversations c WHERE c.node_id=? ORDER BY c.created DESC LIMIT 50", (node_id,), fetch=True)
        else:
            rows = db_exec("SELECT c.id, c.title, c.node_name, c.created, c.parent_exec_id, (SELECT COUNT(*) FROM messages WHERE conv_id=c.id) as msg_count FROM conversations c ORDER BY c.created DESC LIMIT 100", fetch=True)
        return {"ok": True, "conversations": rows}

    def _conv_messages(self, params):
        conv_id = params.get("convId", [""])[0]
        rows = db_exec("SELECT id, role, content, ts, chat_only FROM messages WHERE conv_id=? ORDER BY id", (conv_id,), fetch=True)
        return {"ok": True, "messages": rows}

    # ── Project CRUD ──

    def _project_save(self, body):
        """프로젝트 저장 — 경로 정책 하드 고정.
        정책:
          - 신규 프로젝트: body.workDir 완전 무시하고 이름+날짜로 재계산 (잔상 버그 차단)
          - 기존 프로젝트: 저장된 work_dir 유지 (이름 변경해도 폴더는 그대로)
          - work_dir 내부에 project.json 실파일 dump (DB + 파일시스템 양방향 진실)
          - 임시 폴더(/synology/_temp/...)에서 진행하던 작업이면 정식 폴더로 승격/이동
        """
        pid = body.get("id", str(uuid.uuid4())[:8])
        name = body.get("name", "Untitled")
        now = datetime.now().isoformat()
        # 이름 검증 — 금지 문자/후행 점·공백/예약명 있으면 저장 차단
        ok_name, name_err = validate_project_name(name)
        if not ok_name:
            log(f"[PROJECT_SAVE] 이름 거부: '{name}' — {name_err}")
            return {"ok": False, "error": name_err, "errorKind": "invalid_name"}
        existing = db_exec("SELECT id, work_dir FROM projects WHERE id=?", (pid,), fetchone=True)

        # 날짜 오버라이드: body.saveDate = YYYY-MM-DD / YYYYMMDD / YYYY/MM/DD
        date_override = None
        date_str = (body.get("saveDate") or "").strip()
        if date_str:
            for fmt in ("%Y-%m-%d", "%Y%m%d", "%Y/%m/%d"):
                try:
                    date_override = datetime.strptime(date_str, fmt)
                    break
                except ValueError:
                    continue

        # work_dir 결정
        prev_work_dir = existing.get("work_dir") if existing else None
        if existing and prev_work_dir:
            # 기존 프로젝트 덮어쓰기 → 폴더 유지 (이름이 바뀌어도 물리 폴더는 안 움직임)
            work_dir = _normalize_synology_path(prev_work_dir)[0]
        else:
            work_dir = compute_work_dir(name, date_override=date_override)

        # 임시 폴더에서 정식 폴더로 승격 (있으면 이동)
        # 1) body.promoteFromTemp 명시 지정, 또는 2) body.tempKey 로 캐시된 폴더 자동 감지
        promoted_from = (body.get("promoteFromTemp") or "").strip()
        if not promoted_from:
            tkey = (body.get("tempKey") or "").strip()
            if tkey:
                with _TEMP_CACHE_LOCK:
                    cached = _TEMP_DIR_CACHE.get(tkey)
                if cached and os.path.isdir(cached):
                    promoted_from = cached
        if promoted_from and os.path.isdir(promoted_from) and promoted_from.startswith(TEMP_ROOT):
            try:
                if os.path.exists(work_dir) and os.listdir(work_dir):
                    # 이미 정식 폴더에 내용이 있으면 임시 폴더를 그 안의 _from_temp_<ts>/ 로 병합
                    merge_name = f"_from_temp_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                    shutil.move(promoted_from, os.path.join(work_dir, merge_name))
                    log(f"[PROMOTE] 임시 폴더 내용 병합: {promoted_from} → {work_dir}/{merge_name}")
                else:
                    parent = os.path.dirname(work_dir)
                    os.makedirs(parent, exist_ok=True)
                    if os.path.exists(work_dir):
                        shutil.rmtree(work_dir)
                    shutil.move(promoted_from, work_dir)
                    log(f"[PROMOTE] 임시 폴더 → 정식: {promoted_from} → {work_dir}")
            except Exception as e:
                log(f"[PROMOTE] 임시 폴더 승격 실패: {e}")

        # 실제 폴더 생성
        ok_mk, path_or_err = ensure_work_dir(work_dir)
        if not ok_mk:
            log(f"[PROJECT_SAVE] 폴더 생성 실패: {path_or_err} (프로젝트는 DB 저장, 폴더는 나중에 수동)")

        # body 정리 — workDir 은 이제 서버가 결정한 값으로 덮어씀
        body["workDir"] = work_dir
        body["id"] = pid
        body["name"] = name
        data = json.dumps(body, ensure_ascii=False)

        if existing:
            db_exec("UPDATE projects SET name=?, data=?, modified=?, work_dir=? WHERE id=?",
                    (name, data, now, work_dir, pid))
        else:
            db_exec("INSERT INTO projects (id, name, data, created, modified, work_dir) VALUES (?,?,?,?,?,?)",
                    (pid, name, data, now, now, work_dir))

        # work_dir/project.json 실파일 기록 (실패해도 DB 저장은 유지)
        json_written = False
        if ok_mk:
            try:
                proj_json_path = os.path.join(work_dir, "project.json")
                with open(proj_json_path, "w", encoding="utf-8") as f:
                    json.dump(body, f, ensure_ascii=False, indent=2)
                json_written = True
            except Exception as e:
                log(f"[PROJECT_SAVE] project.json 쓰기 실패: {e}")

        log(f"PROJECT save [{pid}] {name} work_dir={work_dir} folder={ok_mk} json={json_written}")
        return {"ok": True, "id": pid, "workDir": work_dir,
                "folderCreated": ok_mk, "jsonWritten": json_written}

    def _project_load(self, params):
        pid = params.get("id", [""])[0]
        row = db_exec("SELECT * FROM projects WHERE id=?", (pid,), fetchone=True)
        if not row: return {"ok": False, "error": "not found"}
        proj = json.loads(row["data"])
        # work_dir 보강 (data에 없고 컬럼에만 있을 때)
        if not proj.get("workDir") and row.get("work_dir"):
            proj["workDir"] = row["work_dir"]
        return {"ok": True, "project": proj}

    def _project_preview_work_dir(self, params):
        """새 프로젝트 이름 주면 자동 생성될 work_dir 경로 미리보기.
        선택적 date=YYYY-MM-DD or YYYYMMDD 파라미터로 날짜 오버라이드."""
        name = params.get("name", [""])[0]
        if not name:
            return {"ok": False, "error": "name required"}
        date_override = None
        date_str = (params.get("date", [""])[0] or "").strip()
        if date_str:
            for fmt in ("%Y-%m-%d", "%Y%m%d", "%Y/%m/%d"):
                try:
                    date_override = datetime.strptime(date_str, fmt)
                    break
                except ValueError:
                    continue
        path = compute_work_dir(name, date_override=date_override)
        exists = os.path.isdir(path)
        # 파싱된 날짜 / 정리된 이름도 같이 반환 (UI 피드백용)
        parsed_dt, clean_name = _parse_date_from_name(name)
        effective = date_override or parsed_dt or datetime.now()
        # 이름 검증 결과도 함께 반환 — 프론트가 즉시 경고 표시
        ok_name, name_err = validate_project_name(name)
        return {"ok": True, "path": path, "exists": exists,
                "effectiveDate": effective.strftime("%Y-%m-%d"),
                "detectedFromName": bool(parsed_dt and not date_override),
                "cleanName": clean_name,
                "nameValid": ok_name,
                "nameError": name_err or ""}

    def _project_adopt(self, body):
        """기존 폴더를 프로젝트로 편입.
        - folder/project.json 있으면 → 그 상태를 그대로 복원 (노드·연결·캔버스요소·옵션 모두)
        - 없으면 → 빈 플로우로 시작
        body: {folderPath, name?}"""
        folder = (body.get("folderPath") or "").strip()
        if not folder or not os.path.isdir(folder):
            return {"ok": False, "error": "유효한 폴더 경로가 아닙니다"}
        # 이미 편입된 폴더 검사
        exists_row = db_exec("SELECT id, name FROM projects WHERE work_dir=?", (folder,), fetchone=True)
        if exists_row:
            return {"ok": True, "already": True, "id": exists_row["id"], "name": exists_row["name"]}
        name = (body.get("name") or os.path.basename(folder.rstrip("/")) or "Imported").strip()
        ok_name, name_err = validate_project_name(name)
        if not ok_name:
            return {"ok": False, "error": name_err, "errorKind": "invalid_name"}
        pid = str(uuid.uuid4())[:8]
        now = datetime.now().isoformat()

        # project.json 있으면 그 상태를 그대로 복원
        proj_json_path = os.path.join(folder, "project.json")
        restored = False
        payload = None
        if os.path.exists(proj_json_path):
            try:
                with open(proj_json_path, "r", encoding="utf-8") as f:
                    payload = json.load(f)
                # 새 프로젝트로 편입하므로 id/workDir/name 은 새 것으로 덮어씀
                payload["id"] = pid
                payload["name"] = name
                payload["workDir"] = folder
                payload["cwd"] = folder
                restored = True
                log(f"[PROJECT_ADOPT] project.json 복원: {folder} · 노드 {len(payload.get('nodes') or [])}개 · 연결 {len(payload.get('connections') or [])}개 · 캔버스요소 {len(payload.get('canvasElements') or [])}개")
            except Exception as e:
                log(f"[PROJECT_ADOPT] project.json 파싱 실패 ({e}) — 빈 플로우로 시작")
                payload = None
        if payload is None:
            payload = {"id": pid, "name": name, "workDir": folder,
                       "nodes": [], "connections": [], "canvasElements": [], "cwd": folder}

        data = json.dumps(payload, ensure_ascii=False)
        db_exec(
            "INSERT INTO projects (id, name, data, created, modified, work_dir) VALUES (?,?,?,?,?,?)",
            (pid, name, data, now, now, folder),
        )
        # 새 pid로 project.json 도 갱신 (일관성)
        if restored:
            try:
                with open(proj_json_path, "w", encoding="utf-8") as f:
                    json.dump(payload, f, ensure_ascii=False, indent=2)
            except Exception as e:
                log(f"[PROJECT_ADOPT] project.json 재기록 실패: {e}")
        log(f"[PROJECT_ADOPT] {pid} {name} ← {folder} (restored={restored})")
        return {"ok": True, "id": pid, "name": name, "workDir": folder, "restored": restored,
                "nodeCount": len((payload.get("nodes") or [])),
                "connCount": len((payload.get("connections") or []))}

    def _project_scan_unregistered(self, params=None):
        """설정된 work_root 하위에서 아직 DB에 없는 폴더들 반환 (최근 3년까지)."""
        s = _get_work_folder_settings()
        root = s.get("root") or "/synology"
        if not os.path.isdir(root):
            return {"ok": False, "error": f"작업 루트가 없습니다: {root}", "root": root}
        # 이미 등록된 work_dir 세트
        registered = set()
        try:
            rows = db_exec("SELECT work_dir FROM projects WHERE work_dir IS NOT NULL", fetch=True) or []
            for r in rows:
                if r.get("work_dir"):
                    registered.add(os.path.normpath(r["work_dir"]))
        except Exception:
            pass
        candidates = []
        # sub 패턴에 {year}가 있으면 년도 폴더 → 그 하위를 스캔, 없으면 root 바로 하위
        has_year = "{year}" in (s.get("sub") or "")
        try:
            if has_year:
                for yr in sorted(os.listdir(root), reverse=True)[:5]:
                    yr_path = os.path.join(root, yr)
                    if not os.path.isdir(yr_path): continue
                    if not re.match(r'^\d{4}$', yr): continue
                    for fn in sorted(os.listdir(yr_path), reverse=True):
                        full = os.path.normpath(os.path.join(yr_path, fn))
                        if os.path.isdir(full) and not fn.startswith('.') and full not in registered:
                            candidates.append({"path": full, "name": fn, "year": yr})
            else:
                for fn in sorted(os.listdir(root), reverse=True):
                    full = os.path.normpath(os.path.join(root, fn))
                    if os.path.isdir(full) and not fn.startswith('.') and full not in registered:
                        candidates.append({"path": full, "name": fn})
        except PermissionError:
            return {"ok": False, "error": f"권한 없음: {root}"}
        except Exception as e:
            return {"ok": False, "error": str(e)}
        return {"ok": True, "folders": candidates[:100], "root": root}

    # ══════════════ 미디어 다운로드 (YouTube/Instagram/TikTok 등) ══════════════
    def _media_extract_frame(self, body):
        """해당 시각 전후 짧은 세그먼트만 yt-dlp 로 로컬에 다운받은 뒤 ffmpeg 로 프레임 추출.
        -g 방식은 구글비디오 HTTP 헤더·DASH 분리 등으로 ffmpeg 에서 실패 잦음 → 세그먼트 우회.
        body: {url, timestamp (초 단위, 실수)}
        return: {ok, url (/uploads/images/xxx.png), path, size}"""
        url = (body.get("url") or "").strip()
        ts = float(body.get("timestamp") or 0)
        if not url:
            return {"ok": False, "error": "url 필요"}
        if ts < 0: ts = 0

        img_dir = os.path.join(UPLOADS_DIR, "images")
        os.makedirs(img_dir, exist_ok=True)
        out_name = f"ytframe_{int(time.time()*1000)}_{int(ts)}s.png"
        out_path = os.path.join(img_dir, out_name)

        # 임시 세그먼트 다운 디렉토리
        temp_dir = tempfile.mkdtemp(prefix="ytframe_")
        seg_template = os.path.join(temp_dir, "seg.%(ext)s")
        # 더 여유 있는 범위 (앞 2초, 뒤 4초 — 총 6초)로 키프레임 확보 확률 증가
        seg_start = max(0, int(ts) - 2)
        seg_end = int(ts) + 4
        sections = f"*{seg_start}-{seg_end}"

        # 포맷 후보: format 18 (360p progressive mp4, 가장 호환) 우선 → 720p → best
        # --force-keyframes-at-cuts 는 재인코딩 유발하여 실패 잦음 → 제거
        attempts = [
            # 1차: 포맷 18 진행성 mp4, 가장 안정
            ["-f", "18/best[height<=360][ext=mp4]/best[ext=mp4]",
             "--download-sections", sections],
            # 2차: 720p 까지 허용, 세그먼트 여전히 사용
            ["-f", "best[height<=720][ext=mp4]/best[height<=720]/best",
             "--download-sections", sections],
            # 3차: 최후 — 세그먼트 플래그 없이 full 다운로드 (작은 영상만 실용)
            ["-f", "18/best[height<=360][ext=mp4]/best[ext=mp4]"],
        ]
        r = None
        seg_file = None
        used_full_video = False  # 3차 시도(세그먼트 플래그 없음)면 True
        errors_collected = []
        for idx, args in enumerate(attempts):
            try:
                r = subprocess.run(
                    ["yt-dlp"] + args +
                    ["--no-warnings", "--no-playlist", "-o", seg_template, url],
                    capture_output=True, text=True, timeout=120,
                    encoding="utf-8", errors="replace",
                )
            except subprocess.TimeoutExpired:
                errors_collected.append(f"시도 {idx+1}: yt-dlp 타임아웃")
                continue
            except FileNotFoundError:
                shutil.rmtree(temp_dir, ignore_errors=True)
                return {"ok": False, "error": "yt-dlp 미설치"}
            # 결과 파일 찾기 — .part 가 남았을 수도
            try:
                for fn in sorted(os.listdir(temp_dir)):
                    full = os.path.join(temp_dir, fn)
                    if not os.path.isfile(full): continue
                    if not fn.startswith("seg."): continue
                    if fn.endswith(".part"): continue  # 미완성 파일 무시
                    if os.path.getsize(full) < 1024: continue  # 1KB 미만은 실패
                    seg_file = full
                    break
            except Exception: pass
            if seg_file:
                used_full_video = (idx == 2)  # 3차 시도는 세그먼트 플래그 없음
                log(f"[YT_FRAME] 시도 {idx+1} 성공 → {seg_file} ({os.path.getsize(seg_file)} bytes, full={used_full_video})")
                break
            # 이 시도 실패 → stderr 수집
            err_snippet = ((r.stderr if r else '') or (r.stdout if r else '') or '').strip()[-400:]
            errors_collected.append(f"시도 {idx+1}: {err_snippet}")
            # temp_dir 비우기 (부분 파일 제거)
            try:
                for fn in os.listdir(temp_dir):
                    try: os.remove(os.path.join(temp_dir, fn))
                    except Exception: pass
            except Exception: pass

        if not seg_file:
            shutil.rmtree(temp_dir, ignore_errors=True)
            combined = "\n---\n".join(errors_collected)[-800:]
            return {"ok": False, "error": f"세그먼트 다운로드 실패 — 3차 폴백까지 모두 실패.\n\n{combined or 'yt-dlp가 파일을 생성하지 못함'}"}
        # full-video 모드(3차 시도)였다면 local_ts = 실제 ts (0부터 시작이 아님)
        # 그 외는 ts - seg_start

        # 세그먼트 모드면 ts - seg_start, 풀-비디오 모드면 ts 그대로
        local_ts = max(0, ts if used_full_video else (ts - seg_start))
        try:
            r2 = subprocess.run(
                ["ffmpeg", "-y",
                 "-ss", f"{local_ts:.2f}",
                 "-i", seg_file,
                 "-frames:v", "1",
                 "-q:v", "2",
                 "-update", "1",
                 out_path],
                capture_output=True, timeout=30,
            )
        except subprocess.TimeoutExpired:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return {"ok": False, "error": "ffmpeg 타임아웃 (30초)"}
        except FileNotFoundError:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return {"ok": False, "error": "ffmpeg 미설치"}
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

        if not os.path.isfile(out_path) or os.path.getsize(out_path) == 0:
            err_raw = (r2.stderr or b"").decode("utf-8", errors="replace")
            # ffmpeg 버전 헤더 · configure 라인 제거 → 진짜 오류만 남김
            meaningful = []
            for line in err_raw.splitlines():
                l = line.strip()
                if not l: continue
                if l.startswith(("ffmpeg version","built with","configuration:","  lib","Input #","Stream #")): continue
                meaningful.append(l)
            err_msg = "\n".join(meaningful[-5:])[-400:] or "알 수 없는 오류"
            return {"ok": False, "error": f"프레임 추출 실패: {err_msg}"}
        log(f"[EXTRACT_FRAME] {url} @ {ts:.2f}s → {out_path}")
        return {"ok": True, "url": f"/uploads/images/{out_name}",
                "path": out_path, "size": os.path.getsize(out_path),
                "timestamp": ts}

    def _media_download(self, body):
        """URL → 프로젝트 work_dir 하위 videos/ 또는 images/ 로 자동 분류 저장.
        body:
          url         필수 — 다운로드 대상 URL
          projectId   프로젝트 id (work_dir 조회)
          format      'mp4' | 'webm' | 'best'  (기본 mp4)
          quality     'best' | '1080' | '720' | '480' | '360'  (기본 best)
          audioOnly   bool — true 면 mp3 추출 (영상/화질 옵션 무시)
          subtitles   bool — 자막 포함 (있으면 같이 다운)
        """
        url = (body.get("url") or "").strip()
        project_id = (body.get("projectId") or "").strip()
        fmt = (body.get("format") or "mp4").strip().lower()
        quality = str(body.get("quality") or "best").strip().lower()
        audio_only = bool(body.get("audioOnly"))
        subtitles = bool(body.get("subtitles"))
        if fmt not in ("mp4", "webm", "best"): fmt = "mp4"
        if quality not in ("best", "1080", "720", "480", "360"): quality = "best"
        if not url:
            return {"ok": False, "error": "URL이 필요합니다"}

        # 프로젝트 work_dir 조회 — 미저장 프로젝트는 거부 (파일 미아 방지)
        work_dir = None
        if project_id:
            try:
                row = db_exec("SELECT work_dir FROM projects WHERE id=?", (project_id,), fetchone=True)
                if row and row.get("work_dir"):
                    work_dir = row["work_dir"]
            except Exception:
                pass
        if not work_dir:
            return {"ok": False, "error": "프로젝트를 먼저 저장해주세요 (다운로드 파일은 프로젝트 폴더에 저장됩니다)",
                    "errorKind": "no_project"}
        try:
            os.makedirs(work_dir, exist_ok=True)
        except Exception as e:
            return {"ok": False, "error": f"작업 폴더 생성 실패: {e}"}

        # 임시 하위 폴더에 먼저 다운로드
        temp_id = str(uuid.uuid4())[:8]
        temp_dir = os.path.join(work_dir, "_dl_tmp", temp_id)
        try:
            os.makedirs(temp_dir, exist_ok=True)
        except Exception as e:
            return {"ok": False, "error": f"임시 폴더 생성 실패: {e}"}

        # yt-dlp 명령 구성
        output_template = os.path.join(temp_dir, "%(title).80s.%(ext)s")
        cmd = [
            "yt-dlp",
            "-o", output_template,
            "--no-playlist-reverse",
            "--no-warnings",
            "--ignore-errors",
        ]
        if audio_only:
            cmd += ["-x", "--audio-format", "mp3", "--audio-quality", "0"]
        else:
            # 해상도 제한 구성
            res_cap = "" if quality == "best" else f"[height<={quality}]"
            if fmt == "mp4":
                # mp4 우선, m4a 오디오 → 머지해서 mp4 로
                cmd += [
                    "-f", f"bv*[ext=mp4]{res_cap}+ba[ext=m4a]/b[ext=mp4]{res_cap}/bv*{res_cap}+ba/b{res_cap}",
                    "--merge-output-format", "mp4",
                ]
            elif fmt == "webm":
                cmd += [
                    "-f", f"bv*[ext=webm]{res_cap}+ba[ext=webm]/b[ext=webm]{res_cap}/bv*{res_cap}+ba/b{res_cap}",
                    "--merge-output-format", "webm",
                ]
            else:  # best
                if res_cap:
                    cmd += ["-f", f"bv*{res_cap}+ba/b{res_cap}"]
            # 해상도 정렬 우선순위
            if quality != "best":
                cmd += ["-S", f"res:{quality},codec:h264,ext:mp4:m4a"]
            else:
                cmd += ["-S", "res,codec:h264,ext:mp4:m4a"]
        if subtitles:
            cmd += ["--write-subs", "--write-auto-subs", "--sub-langs", "ko,en", "--convert-subs", "srt"]
        cmd.append(url)
        log(f"[DL] {url} → {temp_dir}  fmt={fmt} quality={quality} audio={audio_only}")
        try:
            result = subprocess.run(
                cmd, capture_output=True, text=True,
                timeout=600, encoding="utf-8", errors="replace",
            )
            stderr_tail = (result.stderr or "")[-800:]
            stdout_tail = (result.stdout or "")[-800:]
        except subprocess.TimeoutExpired:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return {"ok": False, "error": "타임아웃 (10분)"}
        except FileNotFoundError:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return {"ok": False, "error": "yt-dlp가 설치되지 않았습니다. Dockerfile 재빌드 필요 (pip install yt-dlp)."}
        except Exception as e:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return {"ok": False, "error": f"실행 오류: {e}"}

        # 다운로드된 파일 스캔 → videos/ images/ 로 분류 이동
        video_exts = {".mp4", ".webm", ".mkv", ".mov", ".avi", ".flv", ".m4v"}
        image_exts = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp", ".heic"}
        other_exts = {".mp3", ".m4a", ".wav", ".ogg", ".opus"}  # 오디오면 audio/
        videos_dir = os.path.join(work_dir, "videos")
        images_dir = os.path.join(work_dir, "images")
        audio_dir  = os.path.join(work_dir, "audio")
        moved = []
        try:
            for fn in os.listdir(temp_dir):
                src = os.path.join(temp_dir, fn)
                if not os.path.isfile(src):
                    continue
                ext = os.path.splitext(fn)[1].lower()
                if ext in video_exts:
                    os.makedirs(videos_dir, exist_ok=True)
                    dst_dir, kind = videos_dir, "video"
                elif ext in image_exts:
                    os.makedirs(images_dir, exist_ok=True)
                    dst_dir, kind = images_dir, "image"
                elif ext in other_exts:
                    os.makedirs(audio_dir, exist_ok=True)
                    dst_dir, kind = audio_dir, "audio"
                else:
                    # 기타 메타 파일(썸네일 .jpg, 자막 .vtt 등) 스킵 — 이미 jpg는 위에서 이미지로 처리됨
                    continue
                dst = os.path.join(dst_dir, fn)
                if os.path.exists(dst):
                    base, e = os.path.splitext(fn)
                    dst = os.path.join(dst_dir, f"{base}_{int(time.time())}{e}")
                try:
                    shutil.move(src, dst)
                    moved.append({
                        "path": dst,
                        "name": os.path.basename(dst),
                        "kind": kind,
                        "size": os.path.getsize(dst),
                    })
                except Exception as me:
                    log(f"[DL] 이동 실패 {src}: {me}")
        except Exception as e:
            log(f"[DL] 스캔 실패: {e}")

        # 임시 폴더 정리
        shutil.rmtree(temp_dir, ignore_errors=True)
        try:
            parent = os.path.join(work_dir, "_dl_tmp")
            if os.path.isdir(parent) and not os.listdir(parent):
                os.rmdir(parent)
        except Exception:
            pass

        if not moved:
            err = stderr_tail or stdout_tail or "다운로드된 파일이 없습니다"
            return {"ok": False, "error": err, "raw": (result.stderr or '')[-400:]}

        return {"ok": True, "files": moved, "count": len(moved),
                "workDir": work_dir,
                "videoCount": sum(1 for m in moved if m["kind"] == "video"),
                "imageCount": sum(1 for m in moved if m["kind"] == "image"),
                "audioCount": sum(1 for m in moved if m["kind"] == "audio")}

    # ═══════════════════════════════════════
    # 🎥 YouTube 검색 (Data API v3) — hot_score 계산 포함
    # ═══════════════════════════════════════
    def _youtube_search(self, body):
        """YouTube Data API v3 검색.
        body: {query, maxResults=30, order='viewCount', duration='any', publishedAfter=''}
        return: {ok, items:[{video_id, title, channel, channel_id, views, likes, comments,
                  duration_sec, published, thumbnail, daily_views, engagement, hot_score, ...}]}
        """
        # 우선순위: DB(설정 UI로 저장한 값) → env(.env 배포 폴백)
        api_key = ""
        try:
            r = db_exec("SELECT value FROM system_settings WHERE key='youtube_api_key'", fetchone=True)
            if r and r.get("value"): api_key = r["value"].strip()
        except Exception: pass
        if not api_key:
            api_key = os.environ.get("YOUTUBE_API_KEY", "").strip()
        if not api_key:
            return {"ok": False, "error": "YouTube API 키가 설정되지 않았습니다. ⚙️ 설정 → 🎬 YouTube Data API 키에서 등록하세요.",
                    "errorKind": "no_api_key"}
        try:
            from googleapiclient.discovery import build
            from googleapiclient.errors import HttpError
        except ImportError:
            return {"ok": False, "error": "google-api-python-client 미설치 — Dockerfile 재빌드 필요"}

        query = (body.get("query") or "").strip()
        if not query:
            return {"ok": False, "error": "검색어를 입력하세요"}
        max_results = min(int(body.get("maxResults") or 30), 200)
        order = body.get("order") or "viewCount"  # relevance, date, rating, viewCount, title
        duration = body.get("duration") or "any"  # any, short, medium, long
        published_after = (body.get("publishedAfter") or "").strip()  # ISO8601 or ""

        try:
            yt = build("youtube", "v3", developerKey=api_key, cache_discovery=False)
            # 1) search.list — video id 만 먼저 얻음 (여러 페이지)
            video_ids = []
            page_token = None
            per_page = 50  # API 최대
            while len(video_ids) < max_results:
                remaining = max_results - len(video_ids)
                search_params = {
                    "q": query,
                    "part": "id",
                    "type": "video",
                    "maxResults": min(per_page, remaining),
                    "order": order,
                    "videoDuration": duration,
                }
                if page_token: search_params["pageToken"] = page_token
                if published_after: search_params["publishedAfter"] = published_after
                resp = yt.search().list(**search_params).execute()
                for item in resp.get("items", []):
                    vid = item.get("id", {}).get("videoId")
                    if vid: video_ids.append(vid)
                page_token = resp.get("nextPageToken")
                if not page_token: break

            if not video_ids:
                return {"ok": True, "items": [], "totalCount": 0}

            # 2) videos.list — 메트릭 배치 조회 (50개씩)
            items = []
            for i in range(0, len(video_ids), 50):
                chunk = video_ids[i:i+50]
                vresp = yt.videos().list(
                    id=",".join(chunk),
                    part="snippet,contentDetails,statistics",
                ).execute()
                for v in vresp.get("items", []):
                    snip = v.get("snippet", {})
                    stat = v.get("statistics", {})
                    cd = v.get("contentDetails", {})
                    vid = v.get("id","")
                    views = int(stat.get("viewCount", 0) or 0)
                    likes = int(stat.get("likeCount", 0) or 0)
                    comments = int(stat.get("commentCount", 0) or 0)
                    dur_sec = self._parse_iso8601_duration(cd.get("duration", "PT0S"))
                    published = snip.get("publishedAt", "")
                    # 업로드 이후 일수
                    days = 1
                    try:
                        pub_dt = datetime.fromisoformat(published.replace("Z", "+00:00"))
                        days = max(1, (datetime.now(pub_dt.tzinfo) - pub_dt).days)
                    except Exception: pass
                    daily_views = views // days if days else views
                    engagement = ((likes + comments) / views * 100) if views else 0
                    hot_score = round(daily_views * (1 + engagement/10) / 1000, 1)
                    thumbs = snip.get("thumbnails", {})
                    thumb = (thumbs.get("high") or thumbs.get("medium") or thumbs.get("default") or {}).get("url", "")
                    items.append({
                        "video_id": vid,
                        "url": f"https://www.youtube.com/watch?v={vid}",
                        "title": snip.get("title",""),
                        "channel_name": snip.get("channelTitle",""),
                        "channel_id": snip.get("channelId",""),
                        "published": published,
                        "days": days,
                        "views": views,
                        "likes": likes,
                        "comments": comments,
                        "duration_sec": dur_sec,
                        "duration_str": self._format_duration(dur_sec),
                        "daily_views": daily_views,
                        "engagement": round(engagement, 2),
                        "hot_score": hot_score,
                        "thumbnail": thumb,
                        "description": (snip.get("description","") or "")[:200],
                    })
            # 핫스코어 내림차순 기본 정렬
            items.sort(key=lambda x: x["hot_score"], reverse=True)
            return {"ok": True, "items": items, "totalCount": len(items),
                    "query": query, "order": order}
        except HttpError as e:
            err_detail = str(e)
            if "quotaExceeded" in err_detail:
                return {"ok": False, "error": "YouTube API 할당량 초과 (일일 10,000 units). 내일 다시 시도하거나 새 API 키를 만드세요."}
            return {"ok": False, "error": f"YouTube API 오류: {err_detail[:300]}"}
        except Exception as e:
            return {"ok": False, "error": f"검색 실패: {e}"}

    def _parse_iso8601_duration(self, s):
        """'PT4M13S' → 253초"""
        m = re.match(r'PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?', s or '')
        if not m: return 0
        h = int(m.group(1) or 0)
        mi = int(m.group(2) or 0)
        se = int(m.group(3) or 0)
        return h*3600 + mi*60 + se

    def _format_duration(self, sec):
        if not sec: return ""
        h, r = divmod(int(sec), 3600)
        m, s = divmod(r, 60)
        return f"{h}:{m:02d}:{s:02d}" if h else f"{m}:{s:02d}"

    # ═══════════════════════════════════════
    # 📝 자막 생성 — YouTube native → faster-whisper 폴백
    # ═══════════════════════════════════════
    def _media_subtitle(self, body):
        """비디오 URL → SRT 파일 생성.
        body: {url, projectId, method='auto' (auto|youtube|whisper), videoTitle?}
        return: {ok, srtPath (/videos/title.srt), preview (첫 500자), source, language}
        """
        url = (body.get("url") or "").strip()
        project_id = (body.get("projectId") or "").strip()
        method = (body.get("method") or "auto").lower()
        want_title = (body.get("videoTitle") or "").strip()
        if not url or not project_id:
            return {"ok": False, "error": "url 과 projectId 필요"}

        # 프로젝트 work_dir
        try:
            row = db_exec("SELECT work_dir FROM projects WHERE id=?", (project_id,), fetchone=True)
        except Exception as e:
            return {"ok": False, "error": f"DB 오류: {e}"}
        if not row or not row.get("work_dir"):
            return {"ok": False, "error": "프로젝트를 먼저 저장해주세요"}
        videos_dir = os.path.join(row["work_dir"], "videos")
        os.makedirs(videos_dir, exist_ok=True)

        # 영상 ID 및 제목 추출
        ytm = re.search(r'(?:youtube\.com/(?:watch\?v=|embed/|shorts/)|youtu\.be/)([A-Za-z0-9_-]{10,14})', url)
        yt_id = ytm.group(1) if ytm else None
        safe_title = _sanitize_folder_name(want_title or (yt_id or "subtitle"))
        srt_path = os.path.join(videos_dir, f"{safe_title}.srt")

        segments = None
        language = None
        source = None

        # 1) YouTube 네이티브 자막 시도
        if method in ("auto", "youtube") and yt_id:
            try:
                from youtube_transcript_api import YouTubeTranscriptApi
                for lang in ["ko", "en", "ja"]:
                    try:
                        data = YouTubeTranscriptApi.get_transcript(yt_id, languages=[lang])
                        segments = [{"start": d.get("start",0), "duration": d.get("duration",0),
                                     "text": (d.get("text","") or "").replace("\n"," ")} for d in data]
                        language = lang; source = "youtube-native"
                        break
                    except Exception:
                        continue
            except ImportError:
                pass

        # 2) 실패 or method=whisper → faster-whisper
        if segments is None and method in ("auto", "whisper"):
            try:
                segments, language = self._whisper_transcribe(url, videos_dir)
                source = "whisper"
            except Exception as e:
                return {"ok": False, "error": f"자막 생성 실패: {e}"}

        if not segments:
            return {"ok": False, "error": "자막을 가져올 수 없습니다. YouTube 자막 없음·Whisper 미설치·영상 접근 불가 중 하나."}

        # SRT 파일 쓰기
        try:
            self._write_srt(srt_path, segments)
        except Exception as e:
            return {"ok": False, "error": f"SRT 쓰기 실패: {e}"}

        # 프리뷰 텍스트
        preview = " ".join(s.get("text","") for s in segments)[:500]
        log(f"[SUBTITLE] {source} ({language}) → {srt_path} ({len(segments)}개 세그먼트)")
        return {"ok": True,
                "srtPath": os.path.join("videos", f"{safe_title}.srt").replace("\\","/"),
                "srtAbsPath": srt_path,
                "preview": preview, "segmentCount": len(segments),
                "source": source, "language": language,
                "title": safe_title}

    def _resolve_local_media_path(self, url):
        """URL이 내부 API 경로(/api/project/file?projectId=X&path=Y)거나 로컬 파일 경로면
        실제 파일시스템 절대 경로를 반환. 외부 URL이면 None."""
        if not url: return None
        u = str(url).strip()
        # 1) 절대 파일 경로 (Linux/Windows)
        if os.path.isabs(u) and os.path.isfile(u):
            return u
        # 2) 내부 API: /api/project/file?projectId=X&path=Y (혹은 전체 URL 형태)
        try:
            from urllib.parse import urlparse, parse_qs, unquote
            parsed = urlparse(u)
            path = parsed.path or u
            if "/api/project/file" in path:
                qs = parse_qs(parsed.query or u.split("?",1)[-1] if "?" in u else "")
                pid = (qs.get("projectId") or [""])[0]
                rel = unquote((qs.get("path") or [""])[0])
                if pid and rel:
                    try:
                        row = db_exec("SELECT work_dir FROM projects WHERE id=?", (pid,), fetchone=True)
                    except Exception:
                        row = None
                    if row and row.get("work_dir"):
                        full = os.path.join(row["work_dir"], rel.replace("\\","/"))
                        if os.path.isfile(full):
                            return full
        except Exception as e:
            log(f"[RESOLVE_LOCAL] 실패: {e}")
        return None

    def _whisper_transcribe(self, url, workdir):
        """faster-whisper 로 전사. 오디오 먼저 추출.
        URL이 내부 API (/api/project/file?...) 또는 로컬 경로면 ffmpeg로 직접 추출,
        외부 URL (YouTube 등)이면 yt-dlp 사용.
        return: (segments list, language)"""
        try:
            from faster_whisper import WhisperModel
        except ImportError:
            raise Exception("faster-whisper 미설치 — Dockerfile 재빌드 필요")

        # 로컬 파일 경로 해석 시도
        local_path = self._resolve_local_media_path(url)

        tmp_dir = tempfile.mkdtemp(prefix="whisper_")
        audio_tpl = os.path.join(tmp_dir, "audio.%(ext)s")
        try:
            audio_file = None
            if local_path and os.path.isfile(local_path):
                # 로컬 파일 → ffmpeg 로 직접 오디오 추출
                audio_file = os.path.join(tmp_dir, "audio.mp3")
                r = subprocess.run(
                    ["ffmpeg", "-y", "-i", local_path, "-vn",
                     "-acodec", "libmp3lame", "-q:a", "5", audio_file],
                    capture_output=True, text=True, timeout=300,
                    encoding="utf-8", errors="replace",
                )
                if r.returncode != 0 or not os.path.isfile(audio_file):
                    raise Exception(f"ffmpeg 오디오 추출 실패: {(r.stderr or r.stdout or '')[-300:]}")
                log(f"[WHISPER] 로컬 파일 audio 추출 완료: {local_path} → {audio_file}")
            else:
                # 외부 URL → yt-dlp
                r = subprocess.run(
                    ["yt-dlp", "-x", "--audio-format", "mp3", "--audio-quality", "5",
                     "-o", audio_tpl, "--no-warnings", "--no-playlist", url],
                    capture_output=True, text=True, timeout=300,
                    encoding="utf-8", errors="replace",
                )
                for fn in os.listdir(tmp_dir):
                    if fn.startswith("audio."):
                        audio_file = os.path.join(tmp_dir, fn); break
                if not audio_file:
                    raise Exception(f"yt-dlp 오디오 추출 실패: {(r.stderr or r.stdout or '')[-300:]}")

            # Whisper 모델 로드 (캐시 디렉토리 사용)
            model_size = os.environ.get("WHISPER_MODEL", "small")
            cache = os.environ.get("WHISPER_CACHE_DIR", "/app/whisper-cache")
            log(f"[WHISPER] 로딩 모델={model_size} cache={cache}")
            model = WhisperModel(model_size, device="cpu", compute_type="int8",
                                 download_root=cache)
            segs_iter, info = model.transcribe(audio_file, beam_size=5, language=None)
            segments = []
            for seg in segs_iter:
                segments.append({"start": seg.start, "duration": seg.end - seg.start,
                                 "text": (seg.text or "").strip()})
            return segments, info.language
        finally:
            shutil.rmtree(tmp_dir, ignore_errors=True)

    def _write_srt(self, path, segments):
        """세그먼트 → SRT 포맷 파일"""
        def fmt_t(sec):
            h, r = divmod(int(sec), 3600)
            m, s = divmod(r, 60)
            ms = int((sec - int(sec)) * 1000)
            return f"{h:02d}:{m:02d}:{s:02d},{ms:03d}"
        with open(path, "w", encoding="utf-8") as f:
            for i, seg in enumerate(segments, 1):
                start = seg.get("start", 0)
                dur = seg.get("duration", 2)
                end = start + dur
                f.write(f"{i}\n{fmt_t(start)} --> {fmt_t(end)}\n{seg.get('text','').strip()}\n\n")

    # ═══════════════════════════════════════
    # 📊 Sheet 노드 — xlsx/csv/tsv 가져오기·내보내기
    # ═══════════════════════════════════════
    def _sheet_import(self, body):
        """파일 base64 받아서 {sheets: [{name, columns, rows, mergesFlattened}]} 반환.
        body:
          filename   원본 파일명 (확장자로 포맷 판별)
          data       base64 인코딩된 파일 내용
          sheetName  (xlsx 전용) 특정 시트만 원할 때 — 없으면 모든 시트
        """
        import base64 as _b64, csv as _csv, io as _io
        filename = (body.get("filename") or "").strip()
        b64 = body.get("data") or ""
        want_sheet = (body.get("sheetName") or "").strip()
        if not filename or not b64:
            return {"ok": False, "error": "filename/data 필요"}
        try:
            if "," in b64 and b64.startswith("data:"):
                b64 = b64.split(",", 1)[1]
            raw = _b64.b64decode(b64)
        except Exception as e:
            return {"ok": False, "error": f"base64 디코드 실패: {e}"}
        ext = os.path.splitext(filename)[1].lower()

        # CSV / TSV — stdlib 만으로
        if ext in (".csv", ".tsv", ".txt"):
            delim = "\t" if ext == ".tsv" else ","
            try:
                text = raw.decode("utf-8-sig")
            except UnicodeDecodeError:
                try: text = raw.decode("cp949")
                except Exception: text = raw.decode("utf-8", errors="replace")
            reader = _csv.reader(_io.StringIO(text), delimiter=delim)
            all_rows = list(reader)
            if not all_rows:
                return {"ok": False, "error": "파일이 비어있습니다"}
            header = all_rows[0]
            # 빈 끝 열 제거
            while header and not (header[-1] or "").strip():
                header.pop()
            columns = [{"id": f"c{i+1}", "name": (h or f"열{i+1}"), "type": "text", "width": 140}
                       for i, h in enumerate(header)]
            col_ids = [c["id"] for c in columns]
            rows = []
            for rrow in all_rows[1:]:
                r = {}
                for i, cid in enumerate(col_ids):
                    r[cid] = rrow[i] if i < len(rrow) else ""
                rows.append(r)
            return {"ok": True, "sheets": [{
                "name": os.path.splitext(os.path.basename(filename))[0],
                "columns": columns, "rows": rows, "mergesFlattened": 0,
            }], "fileType": ext.lstrip(".")}

        # XLSX — openpyxl
        if ext in (".xlsx", ".xlsm"):
            try:
                import openpyxl
            except ImportError:
                return {"ok": False, "error": "openpyxl 미설치 — Dockerfile 재빌드 필요"}
            try:
                wb = openpyxl.load_workbook(_io.BytesIO(raw), data_only=True, read_only=False)
            except Exception as e:
                return {"ok": False, "error": f"xlsx 파싱 실패: {e}"}
            result_sheets = []
            target_names = [want_sheet] if want_sheet else wb.sheetnames
            for name in target_names:
                if name not in wb.sheetnames:
                    continue
                ws = wb[name]
                merges_flattened = 0
                # 병합 영역 값을 좌상단에서 가져와 모든 셀에 복사 (평탄화)
                try:
                    ranges = list(ws.merged_cells.ranges)
                    for rng in ranges:
                        top_left = ws.cell(row=rng.min_row, column=rng.min_col).value
                        ws.unmerge_cells(str(rng))
                        for r in range(rng.min_row, rng.max_row + 1):
                            for c in range(rng.min_col, rng.max_col + 1):
                                ws.cell(row=r, column=c).value = top_left
                        merges_flattened += 1
                except Exception as e:
                    log(f"[SHEET_IMPORT] 병합 평탄화 오류 (무시): {e}")
                # 실제 데이터 영역
                rows_data = list(ws.iter_rows(values_only=True))
                if not rows_data:
                    continue
                # 마지막 전체 None 행/열 트림
                while rows_data and all(v is None or v == "" for v in rows_data[-1]):
                    rows_data.pop()
                if not rows_data:
                    continue
                header = list(rows_data[0])
                # 끝 쪽 None 열 트림
                max_cols = len(header)
                while max_cols > 0 and (header[max_cols-1] is None or header[max_cols-1] == ""):
                    max_cols -= 1
                header = header[:max_cols]
                columns = [{"id": f"c{i+1}",
                            "name": (str(h) if h is not None and str(h).strip() else f"열{i+1}"),
                            "type": "text", "width": 140}
                           for i, h in enumerate(header)]
                col_ids = [c["id"] for c in columns]
                rows = []
                for rrow in rows_data[1:]:
                    r = {}
                    for i, cid in enumerate(col_ids):
                        v = rrow[i] if i < len(rrow) else None
                        if v is None:
                            r[cid] = ""
                        elif isinstance(v, (int, float)):
                            r[cid] = v
                        elif hasattr(v, "isoformat"):
                            r[cid] = v.isoformat()
                        else:
                            r[cid] = str(v)
                    rows.append(r)
                result_sheets.append({
                    "name": name, "columns": columns, "rows": rows,
                    "mergesFlattened": merges_flattened,
                })
            if not result_sheets:
                return {"ok": False, "error": "파싱된 시트가 없습니다"}
            return {"ok": True, "sheets": result_sheets, "fileType": "xlsx",
                    "allSheetNames": wb.sheetnames}
        return {"ok": False, "error": f"지원하지 않는 확장자: {ext}"}

    def _sheet_export_xlsx(self, body):
        """sheet {columns, rows} 받아서 xlsx 바이트를 base64 로 반환."""
        import base64 as _b64, io as _io
        sheet = body.get("sheet") or {}
        name = (body.get("name") or "Sheet1")[:31]
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return {"ok": False, "error": "openpyxl 미설치 — Dockerfile 재빌드 필요"}
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = name
            cols = sheet.get("columns") or []
            rows = sheet.get("rows") or []
            # 헤더
            for ci, c in enumerate(cols, start=1):
                cell = ws.cell(row=1, column=ci, value=c.get("name") or f"열{ci}")
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="EEEEEE")
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[cell.column_letter].width = (c.get("width") or 120) / 7
            # 데이터
            for ri, row in enumerate(rows, start=2):
                for ci, c in enumerate(cols, start=1):
                    v = row.get(c["id"], "")
                    ws.cell(row=ri, column=ci, value=v)
            buf = _io.BytesIO()
            wb.save(buf)
            return {"ok": True, "filename": f"{name}.xlsx",
                    "data": _b64.b64encode(buf.getvalue()).decode("ascii")}
        except Exception as e:
            return {"ok": False, "error": f"xlsx 생성 실패: {e}"}

    # ═══════════════════════════════════════
    # 📄 Document 노드 — DOCX/HWP 파일 I/O + Time Machine 버전관리
    # ═══════════════════════════════════════
    # 파일 레이아웃:
    #   {work_dir}/documents/{basename}.{ext}               ← 현재
    #   {work_dir}/documents/.versions/
    #     {basename}.v{N}.{YYYYMMDD_HHMMSS}.{ext}           ← 자동 스냅샷
    #     {basename}.M.{tag}.{YYYYMMDD_HHMMSS}.{ext}        ← 마일스톤
    # Document 노드가 받아들일 수 있는 모든 파일 확장자
    # - 편집 가능 (inline 에디터): docx, doc, hwp, hwpx, odt, rtf, txt, md
    # - 파일 카드만 (read-only reference): pptx, xlsx, pdf, zip, psd, ai 등 거의 모든 일반 파일
    _DOC_SUPPORTED_EXTS = (
        # 편집 가능
        ".docx", ".doc", ".hwp", ".hwpx", ".odt", ".rtf", ".txt", ".md",
        # 오피스·문서 (읽기 전용 참조)
        ".pptx", ".ppt", ".xlsx", ".xls", ".csv", ".tsv", ".odp", ".ods", ".key", ".pdf", ".epub",
        # 그래픽·디자인
        ".psd", ".ai", ".sketch", ".fig", ".xd", ".eps", ".indd",
        # 코드·데이터
        ".json", ".xml", ".yaml", ".yml", ".ini", ".toml", ".log", ".sql", ".py", ".js", ".ts", ".html", ".css", ".java", ".cpp", ".c", ".h", ".go", ".rs", ".rb", ".php", ".sh",
        # 아카이브
        ".zip", ".7z", ".rar", ".tar", ".gz", ".bz2", ".xz",
        # 미디어 (File 카드 용도 — 캔버스 CE 와는 별개, 참조로만)
        ".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".svg", ".tiff", ".heic",
        ".mp4", ".webm", ".mov", ".m4v", ".mkv", ".avi",
        ".mp3", ".m4a", ".wav", ".ogg", ".opus", ".flac", ".aac",
        # 기타
        ".ics", ".vcf", ".torrent",
    )

    def _doc_project_dir(self, project_id):
        """project_id → documents 디렉토리 절대경로"""
        if not project_id:
            return None, "projectId 필요"
        try:
            row = db_exec("SELECT work_dir FROM projects WHERE id=?", (project_id,), fetchone=True)
        except Exception as e:
            return None, f"DB 오류: {e}"
        if not row or not row.get("work_dir"):
            return None, "프로젝트를 먼저 저장해주세요 (문서는 프로젝트 폴더에 저장됩니다)"
        work_dir = row["work_dir"]
        doc_dir = os.path.join(work_dir, "documents")
        os.makedirs(doc_dir, exist_ok=True)
        return doc_dir, None

    def _doc_validate_filename(self, name):
        """파일명 검증 (경로조작 방지)"""
        if not name or "/" in name or "\\" in name or ".." in name:
            return False
        if name.startswith("."):
            return False
        ext = os.path.splitext(name)[1].lower()
        if ext not in self._DOC_SUPPORTED_EXTS:
            return False
        return True

    def _doc_list(self, params):
        """프로젝트의 documents/ 내 모든 문서 리스트"""
        pid = params.get("projectId", [""])[0]
        doc_dir, err = self._doc_project_dir(pid)
        if err: return {"ok": False, "error": err}
        items = []
        try:
            for name in sorted(os.listdir(doc_dir)):
                if name.startswith("."): continue
                full = os.path.join(doc_dir, name)
                if not os.path.isfile(full): continue
                ext = os.path.splitext(name)[1].lower()
                if ext not in self._DOC_SUPPORTED_EXTS: continue
                try:
                    st = os.stat(full)
                    items.append({
                        "name": name,
                        "path": os.path.join("documents", name).replace("\\", "/"),
                        "size": st.st_size,
                        "mtime": st.st_mtime,
                        "ext": ext.lstrip("."),
                    })
                except Exception: continue
        except Exception as e:
            return {"ok": False, "error": str(e)}
        return {"ok": True, "docDir": doc_dir, "items": items}

    def _doc_read(self, body):
        """파일 바이너리 읽어서 base64 로 반환.
        body: {projectId, filename}
        returns: {ok, data(base64), size, mtime, path}"""
        import base64 as _b64
        pid = (body.get("projectId") or "").strip()
        filename = (body.get("filename") or "").strip()
        doc_dir, err = self._doc_project_dir(pid)
        if err: return {"ok": False, "error": err}
        if not self._doc_validate_filename(filename):
            return {"ok": False, "error": "유효하지 않은 파일명"}
        full = os.path.join(doc_dir, filename)
        if not os.path.isfile(full):
            return {"ok": False, "error": "파일 없음", "errorKind": "not_found"}
        try:
            with open(full, "rb") as f: raw = f.read()
            st = os.stat(full)
            return {"ok": True, "data": _b64.b64encode(raw).decode("ascii"),
                    "size": st.st_size, "mtime": st.st_mtime,
                    "filename": filename,
                    "path": os.path.join("documents", filename).replace("\\", "/")}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _doc_snapshot(self, doc_dir, filename, reason="auto", milestone_tag=None):
        """현재 파일을 .versions/ 에 스냅샷. 실패해도 silent."""
        full = os.path.join(doc_dir, filename)
        if not os.path.isfile(full): return None
        vdir = os.path.join(doc_dir, ".versions")
        try: os.makedirs(vdir, exist_ok=True)
        except Exception: return None
        base, ext = os.path.splitext(filename)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        if milestone_tag:
            safe_tag = re.sub(r'[^\w\-_가-힣]', '_', milestone_tag)[:40] or 'milestone'
            snap_name = f"{base}.M.{safe_tag}.{ts}{ext}"
        else:
            # 자동 버전 번호 계산
            existing = [f for f in os.listdir(vdir) if f.startswith(base+".v") and f.endswith(ext)]
            n = len(existing) + 1
            snap_name = f"{base}.v{n:03d}.{ts}{ext}"
        snap_path = os.path.join(vdir, snap_name)
        try:
            shutil.copy2(full, snap_path)
            log(f"[DOC_SNAPSHOT] {reason}: {filename} → .versions/{snap_name}")
            return snap_name
        except Exception as e:
            log(f"[DOC_SNAPSHOT] 실패 {filename}: {e}")
            return None

    def _doc_thin_versions(self, doc_dir, filename):
        """Time Machine 스타일 씨이어링 — 마일스톤은 건드리지 않음.
        - <24h : 전부 유지
        - <7d  : 시간당 1개
        - <30d : 하루 1개
        - <180d: 주당 1개
        - >=180d: 월당 1개
        """
        vdir = os.path.join(doc_dir, ".versions")
        if not os.path.isdir(vdir): return
        base, ext = os.path.splitext(filename)
        now = time.time()
        candidates = []
        try:
            for name in os.listdir(vdir):
                if not name.startswith(base+"."): continue
                if ".M." in name: continue   # 마일스톤 제외
                if not name.endswith(ext): continue
                full = os.path.join(vdir, name)
                if not os.path.isfile(full): continue
                try:
                    mt = os.path.getmtime(full)
                    candidates.append((mt, full))
                except Exception: continue
        except Exception: return
        # 오래된 것부터 그룹핑 버킷 기준으로 1개만 남김
        candidates.sort(key=lambda x: x[0])
        kept_buckets = set()
        to_delete = []
        for mt, path in candidates:
            age = now - mt
            if age < 86400:  # 24h
                bucket = None  # 전부 유지
            elif age < 7*86400:
                bucket = ("h", int(mt // 3600))
            elif age < 30*86400:
                bucket = ("d", int(mt // 86400))
            elif age < 180*86400:
                bucket = ("w", int(mt // (7*86400)))
            else:
                bucket = ("m", int(mt // (30*86400)))
            if bucket is None: continue
            if bucket in kept_buckets:
                to_delete.append(path)
            else:
                kept_buckets.add(bucket)
        pruned = 0
        for p in to_delete:
            try: os.remove(p); pruned += 1
            except Exception: pass
        if pruned: log(f"[DOC_THIN] {filename}: {pruned}개 자동 정리")

    def _project_mkdir(self, body):
        """프로젝트 폴더 안에 임의 하위 폴더 생성.
        body: {projectId, folderName (상대경로 허용, 예: 'refs' 또는 'archive/2024')}"""
        pid = (body.get("projectId") or "").strip()
        folder = (body.get("folderName") or "").strip()
        if not pid or not folder:
            return {"ok": False, "error": "projectId, folderName 필요"}
        # 상위 이동(..) / 절대경로 차단
        folder = folder.replace("\\", "/").strip("/")
        if ".." in folder.split("/") or folder.startswith("/"):
            return {"ok": False, "error": "유효하지 않은 경로"}
        try:
            row = db_exec("SELECT work_dir FROM projects WHERE id=?", (pid,), fetchone=True)
        except Exception as e:
            return {"ok": False, "error": f"DB 오류: {e}"}
        if not row or not row.get("work_dir"):
            return {"ok": False, "error": "프로젝트 미저장"}
        target = os.path.join(row["work_dir"], folder)
        if not os.path.abspath(target).startswith(os.path.abspath(row["work_dir"])):
            return {"ok": False, "error": "경로 침범"}
        try:
            os.makedirs(target, exist_ok=True)
            return {"ok": True, "path": target, "relPath": folder}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _project_copy_files(self, body):
        """프로젝트 폴더 내 파일들을 다른 목적지로 복사.
        body: {projectId (source), files (relpaths), targetProjectId?, targetAbsPath?, subfolder?}
        - targetProjectId 우선 (해당 프로젝트 work_dir 로)
        - 아니면 targetAbsPath (/volume1/... 절대 경로)
        - subfolder 지정하면 그 밑으로 (없으면 루트)
        return: {ok, copied:[{src,dst}], failed:[]}"""
        return self._project_copy_or_move(body, move=False)

    def _project_move_files(self, body):
        """동일하지만 원본 제거 (move)."""
        return self._project_copy_or_move(body, move=True)

    def _project_copy_or_move(self, body, move=False):
        import shutil as _sh
        pid = (body.get("projectId") or "").strip()
        files = body.get("files") or []
        target_pid = (body.get("targetProjectId") or "").strip()
        target_abs = (body.get("targetAbsPath") or "").strip()
        subfolder = (body.get("subfolder") or "").strip().replace("\\","/").strip("/")
        if not pid or not files:
            return {"ok": False, "error": "projectId, files 필요"}
        # 소스 work_dir
        try:
            row = db_exec("SELECT work_dir FROM projects WHERE id=?", (pid,), fetchone=True)
        except Exception as e:
            return {"ok": False, "error": f"DB 오류: {e}"}
        if not row or not row.get("work_dir"):
            return {"ok": False, "error": "소스 프로젝트 미저장"}
        src_root = row["work_dir"]
        # 목적지 결정
        if target_pid:
            try:
                trow = db_exec("SELECT work_dir FROM projects WHERE id=?", (target_pid,), fetchone=True)
            except Exception as e:
                return {"ok": False, "error": f"타겟 DB 오류: {e}"}
            if not trow or not trow.get("work_dir"):
                return {"ok": False, "error": "타겟 프로젝트 미저장"}
            dst_root = trow["work_dir"]
        elif target_abs:
            dst_root = target_abs
        else:
            return {"ok": False, "error": "타겟 경로 필요 (targetProjectId 또는 targetAbsPath)"}
        if subfolder:
            if ".." in subfolder.split("/"): return {"ok": False, "error": "유효하지 않은 subfolder"}
            dst_root = os.path.join(dst_root, subfolder)
        try: os.makedirs(dst_root, exist_ok=True)
        except Exception as e: return {"ok": False, "error": f"타겟 폴더 생성 실패: {e}"}
        copied = []; failed = []
        for rel in files:
            rel = (rel or "").replace("\\","/").lstrip("/")
            if ".." in rel.split("/"):
                failed.append({"file":rel,"error":"경로 침범"}); continue
            src = os.path.join(src_root, rel)
            if not os.path.isfile(src):
                failed.append({"file":rel,"error":"원본 없음"}); continue
            fname = os.path.basename(rel)
            dst = os.path.join(dst_root, fname)
            # 중복 시 (2), (3) ... 번호
            if os.path.exists(dst):
                stem, ext = os.path.splitext(fname)
                for i in range(2, 1000):
                    cand = os.path.join(dst_root, f"{stem} ({i}){ext}")
                    if not os.path.exists(cand): dst = cand; break
            try:
                if move: _sh.move(src, dst)
                else: _sh.copy2(src, dst)
                copied.append({"src":rel,"dst":os.path.relpath(dst, dst_root)})
            except Exception as e:
                failed.append({"file":rel,"error":str(e)})
        return {"ok": True, "copied": copied, "failed": failed,
                "dstRoot": dst_root, "moved": move}

    def _project_file_upload(self, body):
        """로컬 파일을 프로젝트 폴더로 직접 저장 (싱크 스타일).
        body: {projectId, filename, data(base64), category? ('documents'|'images'|'videos'|'audio'|'attachments')}
        return: {ok, filename, path(절대), url(브라우저용), relPath(프로젝트 기준 상대)}
        프로젝트 미저장 시 → _temp 폴더 사용."""
        import base64 as _b64
        pid = (body.get("projectId") or "").strip()
        filename = (body.get("filename") or "").strip()
        b64 = body.get("data") or ""
        category = (body.get("category") or "").strip()
        if not filename or not b64:
            return {"ok": False, "error": "filename 과 data 필요"}
        # 카테고리 자동 판정 (확장자)
        if not category:
            ext = filename.rsplit(".",1)[-1].lower() if "." in filename else ""
            if ext in ("jpg","jpeg","png","gif","webp","bmp","svg","ico"): category = "images"
            elif ext in ("mp4","mov","webm","mkv","avi","m4v"): category = "videos"
            elif ext in ("mp3","wav","m4a","flac","ogg","aac"): category = "audio"
            elif ext in ("docx","hwp","hwpx","pdf","xlsx","txt","md","csv","json","html","rtf","odt","pptx"): category = "documents"
            else: category = "attachments"
        # 프로젝트 work_dir 확보
        work_dir = None
        if pid:
            try:
                row = db_exec("SELECT work_dir FROM projects WHERE id=?", (pid,), fetchone=True)
                if row and row.get("work_dir"): work_dir = row["work_dir"]
            except Exception as e:
                log(f"[PFU] DB 조회 실패: {e}")
        if not work_dir:
            # _temp 폴더 사용 (저장 안 된 프로젝트)
            try:
                temp_key = body.get("tempKey") or ""
                work_dir = self._get_or_create_temp_dir_pf(temp_key)
            except Exception:
                work_dir = None
        if not work_dir:
            return {"ok": False, "error": "프로젝트 폴더 확보 실패"}
        target_dir = os.path.join(work_dir, category)
        try: os.makedirs(target_dir, exist_ok=True)
        except Exception as e: return {"ok": False, "error": f"디렉토리 생성 실패: {e}"}
        # 파일명 sanitize + 중복 시 자동 번호
        base = re.sub(r'[^\w\s.\-_가-힣()]', '_', filename).strip().strip('.')
        if not base: base = "file"
        target = os.path.join(target_dir, base)
        if os.path.exists(target):
            stem, ext = os.path.splitext(base)
            for i in range(2, 1000):
                candidate = os.path.join(target_dir, f"{stem} ({i}){ext}")
                if not os.path.exists(candidate):
                    target = candidate; base = f"{stem} ({i}){ext}"; break
        # 저장
        try:
            if "," in b64 and b64.startswith("data:"): b64 = b64.split(",",1)[1]
            data = _b64.b64decode(b64)
            with open(target, "wb") as f: f.write(data)
            st = os.stat(target)
        except Exception as e:
            return {"ok": False, "error": f"저장 실패: {e}"}
        rel_path = f"{category}/{base}"
        # 브라우저 표시용 URL (이미지/비디오 <src> 지원)
        url = f"/api/project/file?projectId={pid}&path={category}%2F{base}" if pid else ""
        log(f"[PFU] {filename} → {target} ({st.st_size} bytes, category={category})")
        return {"ok": True, "filename": base, "path": target,
                "url": url, "relPath": rel_path, "category": category, "size": st.st_size}

    def _get_or_create_temp_dir_pf(self, temp_key):
        """_temp 폴더 안에 현재 세션용 임시 디렉토리 반환 (없으면 생성)."""
        try:
            base = self._resolve_temp_base() if hasattr(self, "_resolve_temp_base") else "/synology/_temp"
        except Exception:
            base = "/synology/_temp"
        os.makedirs(base, exist_ok=True)
        key = (temp_key or f"tk_{int(time.time()*1000)}").strip()
        d = os.path.join(base, key)
        os.makedirs(d, exist_ok=True)
        return d

    def _doc_write(self, body):
        """파일 저장 + 저장 전 자동 스냅샷 + 씨이어링.
        body: {projectId, filename, data(base64), clientMtime?, force?}
        - clientMtime 이 있고 디스크 mtime 이 더 크면 외부 수정 감지 → 자동 스냅샷 후 진행 (force=true)
          또는 {errorKind:'external_modified'} 리턴 (force=false)
        """
        import base64 as _b64
        pid = (body.get("projectId") or "").strip()
        filename = (body.get("filename") or "").strip()
        b64 = body.get("data") or ""
        client_mtime = body.get("clientMtime")
        force = bool(body.get("force"))
        doc_dir, err = self._doc_project_dir(pid)
        if err: return {"ok": False, "error": err}
        if not self._doc_validate_filename(filename):
            return {"ok": False, "error": "유효하지 않은 파일명"}
        full = os.path.join(doc_dir, filename)
        # 외부 수정 감지
        if os.path.isfile(full) and client_mtime is not None:
            try:
                actual_mt = os.path.getmtime(full)
                if actual_mt > float(client_mtime) + 1:  # 1s tolerance
                    if not force:
                        return {"ok": False, "errorKind": "external_modified",
                                "error": "외부에서 수정된 흔적이 있습니다. 덮어쓰시겠어요?",
                                "diskMtime": actual_mt, "clientMtime": client_mtime}
                    # force=true → 저장 직전 외부 버전 스냅샷 보관
                    self._doc_snapshot(doc_dir, filename, reason="pre-overwrite-external")
            except Exception: pass
        # 저장 전 현재 상태 스냅샷 (파일이 이미 있으면)
        if os.path.isfile(full):
            self._doc_snapshot(doc_dir, filename, reason="pre-save")
        # 실제 쓰기
        try:
            if "," in b64 and b64.startswith("data:"):
                b64 = b64.split(",", 1)[1]
            raw = _b64.b64decode(b64)
            with open(full, "wb") as f: f.write(raw)
            st = os.stat(full)
            # 씨이어링
            try: self._doc_thin_versions(doc_dir, filename)
            except Exception as e: log(f"[DOC_THIN] {e}")
            log(f"[DOC_WRITE] {filename} ({st.st_size} bytes)")
            return {"ok": True, "size": st.st_size, "mtime": st.st_mtime,
                    "path": os.path.join("documents", filename).replace("\\", "/")}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _doc_versions(self, params):
        """파일의 버전 목록 (최신순). 자동 + 마일스톤 구분."""
        pid = params.get("projectId", [""])[0]
        filename = params.get("filename", [""])[0]
        doc_dir, err = self._doc_project_dir(pid)
        if err: return {"ok": False, "error": err}
        if not self._doc_validate_filename(filename):
            return {"ok": False, "error": "유효하지 않은 파일명"}
        vdir = os.path.join(doc_dir, ".versions")
        base, ext = os.path.splitext(filename)
        items = []
        if os.path.isdir(vdir):
            try:
                for name in os.listdir(vdir):
                    if not name.startswith(base+"."): continue
                    if not name.endswith(ext): continue
                    full = os.path.join(vdir, name)
                    if not os.path.isfile(full): continue
                    is_milestone = ".M." in name
                    tag = ""
                    if is_milestone:
                        # {base}.M.{tag}.{ts}{ext}
                        inner = name[len(base)+3:-len(ext)]  # "{tag}.{ts}"
                        parts = inner.rsplit(".", 1)
                        if len(parts) == 2: tag = parts[0]
                    try:
                        st = os.stat(full)
                        items.append({
                            "name": name,
                            "kind": "milestone" if is_milestone else "auto",
                            "tag": tag,
                            "size": st.st_size,
                            "mtime": st.st_mtime,
                        })
                    except Exception: continue
            except Exception as e:
                return {"ok": False, "error": str(e)}
        items.sort(key=lambda x: x["mtime"], reverse=True)
        # 현재 파일 정보도 추가
        current = None
        full = os.path.join(doc_dir, filename)
        if os.path.isfile(full):
            try:
                st = os.stat(full)
                current = {"size": st.st_size, "mtime": st.st_mtime}
            except Exception: pass
        return {"ok": True, "current": current, "versions": items}

    def _doc_restore(self, body):
        """선택 버전을 현재 파일로 복원. 복원 전 현재 상태 자동 스냅샷.
        body: {projectId, filename, versionName}"""
        pid = (body.get("projectId") or "").strip()
        filename = (body.get("filename") or "").strip()
        ver = (body.get("versionName") or "").strip()
        doc_dir, err = self._doc_project_dir(pid)
        if err: return {"ok": False, "error": err}
        if not self._doc_validate_filename(filename):
            return {"ok": False, "error": "유효하지 않은 파일명"}
        if "/" in ver or "\\" in ver or ".." in ver:
            return {"ok": False, "error": "유효하지 않은 버전명"}
        vdir = os.path.join(doc_dir, ".versions")
        ver_path = os.path.join(vdir, ver)
        if not os.path.isfile(ver_path):
            return {"ok": False, "error": "버전 파일 없음"}
        full = os.path.join(doc_dir, filename)
        # 복원 직전 현재 상태 자동 보관
        if os.path.isfile(full):
            self._doc_snapshot(doc_dir, filename, reason="pre-restore")
        try:
            shutil.copy2(ver_path, full)
            st = os.stat(full)
            log(f"[DOC_RESTORE] {filename} ← {ver}")
            return {"ok": True, "size": st.st_size, "mtime": st.st_mtime}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _doc_milestone(self, body):
        """현재 파일을 마일스톤 스냅샷으로 보관.
        body: {projectId, filename, tag, note?}"""
        pid = (body.get("projectId") or "").strip()
        filename = (body.get("filename") or "").strip()
        tag = (body.get("tag") or "").strip() or "milestone"
        doc_dir, err = self._doc_project_dir(pid)
        if err: return {"ok": False, "error": err}
        if not self._doc_validate_filename(filename):
            return {"ok": False, "error": "유효하지 않은 파일명"}
        snap = self._doc_snapshot(doc_dir, filename, reason="milestone", milestone_tag=tag)
        if not snap:
            return {"ok": False, "error": "스냅샷 실패 — 파일이 없거나 권한 문제"}
        return {"ok": True, "versionName": snap}

    def _doc_delete_version(self, body):
        """특정 버전 파일 삭제 (마일스톤도 허용)."""
        pid = (body.get("projectId") or "").strip()
        filename = (body.get("filename") or "").strip()
        ver = (body.get("versionName") or "").strip()
        doc_dir, err = self._doc_project_dir(pid)
        if err: return {"ok": False, "error": err}
        if not self._doc_validate_filename(filename):
            return {"ok": False, "error": "유효하지 않은 파일명"}
        if "/" in ver or "\\" in ver or ".." in ver:
            return {"ok": False, "error": "유효하지 않은 버전명"}
        p = os.path.join(doc_dir, ".versions", ver)
        if not os.path.isfile(p):
            return {"ok": False, "error": "파일 없음"}
        try:
            os.remove(p)
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _doc_delete(self, body):
        """현재 문서 삭제 (휴지통으로 이동 — .trash-doc/).
        body: {projectId, filename}"""
        pid = (body.get("projectId") or "").strip()
        filename = (body.get("filename") or "").strip()
        doc_dir, err = self._doc_project_dir(pid)
        if err: return {"ok": False, "error": err}
        if not self._doc_validate_filename(filename):
            return {"ok": False, "error": "유효하지 않은 파일명"}
        full = os.path.join(doc_dir, filename)
        if not os.path.isfile(full):
            return {"ok": False, "error": "파일 없음"}
        # 삭제 전 마지막 스냅샷 (안전)
        self._doc_snapshot(doc_dir, filename, reason="pre-delete", milestone_tag="deleted")
        try:
            trash_dir = os.path.join(doc_dir, ".trash-doc")
            os.makedirs(trash_dir, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            target = os.path.join(trash_dir, f"{ts}_{filename}")
            shutil.move(full, target)
            log(f"[DOC_DELETE] {filename} → .trash-doc/")
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _doc_hwp_to_docx(self, body):
        """LibreOffice 로 HWP → DOCX 변환. libreoffice 가 설치되어 있을 때만 동작.
        body: {projectId, filename} — filename 이 .hwp 면 같은 베이스의 .docx 생성"""
        pid = (body.get("projectId") or "").strip()
        filename = (body.get("filename") or "").strip()
        doc_dir, err = self._doc_project_dir(pid)
        if err: return {"ok": False, "error": err}
        if not filename.lower().endswith((".hwp", ".hwpx")):
            return {"ok": False, "error": "HWP/HWPX 만 변환 가능"}
        src = os.path.join(doc_dir, filename)
        if not os.path.isfile(src):
            return {"ok": False, "error": "원본 없음"}
        # libreoffice 확인
        from shutil import which as _which
        soffice = _which("libreoffice") or _which("soffice")
        if not soffice:
            return {"ok": False, "error": "libreoffice 미설치 — Dockerfile 재빌드 후 재시도"}
        try:
            result = subprocess.run(
                [soffice, "--headless", "--convert-to", "docx", "--outdir", doc_dir, src],
                capture_output=True, text=True, timeout=120,
            )
            base = os.path.splitext(filename)[0]
            out = os.path.join(doc_dir, base + ".docx")
            if not os.path.isfile(out):
                return {"ok": False, "error": f"변환 실패: {result.stderr[:400] or result.stdout[:400]}"}
            log(f"[DOC_HWP2DOCX] {filename} → {base}.docx")
            return {"ok": True, "filename": base + ".docx"}
        except subprocess.TimeoutExpired:
            return {"ok": False, "error": "변환 타임아웃 (2분)"}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    # ═══════════════════════════════════════
    # 📂 프로젝트 폴더 실시간 파일 리스트 + 파일 서빙
    # ═══════════════════════════════════════
    _PROJ_FILE_GROUPS = {
        'documents': {
            'label':'📄 문서',
            'exts':{'docx','doc','hwp','hwpx','odt','rtf','txt','md','pdf','pptx','ppt','xlsx','xls','csv','odp','ods','key','tsv'},
        },
        'images': {
            'label':'🖼 이미지',
            'exts':{'jpg','jpeg','png','gif','webp','bmp','svg','ico','tiff','heic','heif'},
        },
        'videos': {
            'label':'🎬 동영상',
            'exts':{'mp4','webm','ogv','mov','m4v','avi','mkv','flv'},
        },
        'audio': {
            'label':'🎵 오디오',
            'exts':{'mp3','m4a','wav','ogg','opus','aac','flac'},
        },
        'other': {'label':'📎 기타', 'exts': set()},
    }
    _PROJ_HIDDEN_DIRS = {'.versions','.trash-doc','_dl_tmp','.locks','.trash','_temp','_dl_tmp'}
    _PROJ_HIDDEN_FILES = {'project.json', '.DS_Store', 'Thumbs.db', 'desktop.ini'}

    def _classify_ext(self, ext):
        e=(ext or '').lower()
        for key, g in self._PROJ_FILE_GROUPS.items():
            if key=='other': continue
            if e in g['exts']: return key
        return 'other'

    def _project_files(self, params):
        """프로젝트 폴더 파일 리스트 그룹화 반환."""
        pid = params.get("projectId", [""])[0]
        show_system = params.get("showSystem", ["0"])[0] == "1"
        max_depth = int(params.get("maxDepth", ["3"])[0] or "3")
        try:
            row = db_exec("SELECT work_dir, name FROM projects WHERE id=?", (pid,), fetchone=True)
        except Exception as e:
            return {"ok": False, "error": f"DB 오류: {e}"}
        if not row or not row.get("work_dir"):
            return {"ok": False, "error": "프로젝트를 먼저 저장해주세요"}
        root = row["work_dir"]
        if not os.path.isdir(root):
            return {"ok": False, "error": f"폴더 없음: {root}"}

        groups = {k: {'label': v['label'], 'files': []} for k,v in self._PROJ_FILE_GROUPS.items()}
        all_entries = []
        try:
            for dirpath, dirnames, filenames in os.walk(root):
                # 시스템 폴더 스킵
                if not show_system:
                    dirnames[:] = [d for d in dirnames if d not in self._PROJ_HIDDEN_DIRS and not d.startswith('.')]
                rel_root = os.path.relpath(dirpath, root).replace('\\','/')
                depth = 0 if rel_root == '.' else rel_root.count('/') + 1
                if depth > max_depth:
                    dirnames[:] = []
                    continue
                for fn in filenames:
                    if not show_system:
                        if fn in self._PROJ_HIDDEN_FILES: continue
                        if fn.startswith('.'): continue
                    full = os.path.join(dirpath, fn)
                    try:
                        st = os.stat(full)
                    except Exception: continue
                    ext = os.path.splitext(fn)[1].lower().lstrip('.')
                    rel = os.path.relpath(full, root).replace('\\','/')
                    kind = self._classify_ext(ext)
                    entry = {
                        'name': fn, 'path': rel,
                        'size': st.st_size, 'mtime': st.st_mtime, 'ext': ext,
                        'kind': kind,
                    }
                    groups[kind]['files'].append(entry)
                    all_entries.append(entry)
        except Exception as e:
            return {"ok": False, "error": str(e)}

        # 각 그룹 최신순 정렬
        for g in groups.values():
            g['files'].sort(key=lambda f: f['mtime'], reverse=True)

        # 변경 감지 해시 (이름+mtime+size)
        sig = "\n".join(f"{e['path']}|{e['mtime']}|{e['size']}" for e in sorted(all_entries, key=lambda x:x['path']))
        phash = hashlib.md5(sig.encode('utf-8')).hexdigest()[:16]

        return {
            "ok": True, "root": root, "groups": groups, "hash": phash,
            "totalFiles": len(all_entries),
            "projectName": row.get("name",""),
        }

    def _project_file_serve(self, params):
        """프로젝트 폴더 내 파일을 Content-Type 맞춰 서빙 (이미지·비디오 <img src> 용).
        쿼리: projectId, path (프로젝트 루트 기준 상대)"""
        pid = params.get("projectId", [""])[0]
        rel = params.get("path", [""])[0]
        if not pid or not rel:
            self.send_error(400, "projectId and path required")
            return
        # 경로 조작 방지
        if '..' in rel.split('/') or rel.startswith('/'):
            self.send_error(400, "invalid path")
            return
        try:
            row = db_exec("SELECT work_dir FROM projects WHERE id=?", (pid,), fetchone=True)
        except Exception:
            row = None
        if not row or not row.get("work_dir"):
            self.send_error(404, "project not found")
            return
        root = row["work_dir"]
        full = os.path.join(root, rel.replace('\\','/'))
        # canonical 경로 재확인 (심볼릭 링크 · .. 완전 방지)
        if not os.path.abspath(full).startswith(os.path.abspath(root)):
            self.send_error(403, "forbidden")
            return
        if not os.path.isfile(full):
            self.send_error(404, "file not found")
            return
        try:
            import mimetypes
            mime, _ = mimetypes.guess_type(full)
            if not mime: mime = 'application/octet-stream'
            size = os.path.getsize(full)
            self.send_response(200)
            self.send_header("Content-Type", mime)
            self.send_header("Content-Length", str(size))
            # 캐시: mtime 기반 ETag (브라우저가 변경 감지 가능)
            etag = f'"{int(os.path.getmtime(full))}-{size}"'
            self.send_header("ETag", etag)
            self.send_header("Cache-Control", "private, max-age=60")
            # 인라인 표시 (이미지/비디오/PDF 등)
            self.end_headers()
            with open(full, "rb") as f:
                while True:
                    chunk = f.read(65536)
                    if not chunk: break
                    self.wfile.write(chunk)
        except Exception as e:
            log(f"[PROJECT_FILE] 서빙 오류 {full}: {e}")
            try: self.send_error(500, str(e))
            except Exception: pass

    def _project_list(self):
        rows = db_exec("SELECT id, name, modified, favorite, folder_id FROM projects ORDER BY favorite DESC, modified DESC LIMIT 200", fetch=True) or []
        # __current__ 류는 Python에서 필터링 (LIKE ESCAPE 호환성 이슈 회피)
        rows = [r for r in rows if not (r.get("id") or "").startswith("__current")]
        log(f"[PROJECT_LIST] {len(rows)}개 반환")
        return {"ok": True, "projects": rows}

    def _project_list_meta(self):
        """메타데이터만 (data 필드 제외, 가벼움) — 노드 수도 함께"""
        rows = db_exec("SELECT id, name, modified, created, favorite, folder_id, work_dir FROM projects ORDER BY favorite DESC, modified DESC LIMIT 200", fetch=True) or []
        # __current__ 류는 Python에서 필터링
        rows = [r for r in rows if not (r.get("id") or "").startswith("__current")]
        log(f"[PROJECT_LIST_META] {len(rows)}개 반환")
        # 노드 수만 추가 추출
        for r in rows:
            try:
                d = db_exec("SELECT data FROM projects WHERE id=?", (r["id"],), fetchone=True)
                if d:
                    pj = json.loads(d["data"])
                    r["nodeCount"] = len(pj.get("nodes", []))
                    r["connCount"] = len(pj.get("connections", []))
                else:
                    r["nodeCount"] = 0
                    r["connCount"] = 0
            except:
                r["nodeCount"] = 0
                r["connCount"] = 0
        return {"ok": True, "projects": rows}

    def _project_star(self, body):
        pid = body.get("id", "")
        fav = 1 if body.get("favorite") else 0
        if not pid: return {"ok": False, "error": "id required"}
        db_exec("UPDATE projects SET favorite=? WHERE id=?", (fav, pid))
        return {"ok": True}

    def _project_move(self, body):
        pid = body.get("id", "")
        folder_id = body.get("folderId")
        if not pid: return {"ok": False, "error": "id required"}
        db_exec("UPDATE projects SET folder_id=? WHERE id=?", (folder_id, pid))
        return {"ok": True}

    # ── Project Folders ──

    def _project_folder_list(self):
        rows = db_exec("SELECT * FROM project_folders ORDER BY sort_order, name", fetch=True)
        return {"ok": True, "folders": rows}

    def _project_folder_save(self, body):
        fid = body.get("id")
        name = body.get("name", "새 폴더")
        icon = body.get("icon", "📂")
        color = body.get("color", "")
        parent_id = body.get("parentId")
        if parent_id in ("", 0, None): parent_id = None
        now = datetime.now().isoformat()
        if fid:
            # 부분 업데이트 — body에 있는 필드만 갱신
            updates, params = [], []
            if "name" in body: updates.append("name=?"); params.append(name)
            if "icon" in body: updates.append("icon=?"); params.append(icon)
            if "color" in body: updates.append("color=?"); params.append(color)
            if "parentId" in body: updates.append("parent_id=?"); params.append(parent_id)
            if updates:
                params.append(fid)
                db_exec(f"UPDATE project_folders SET {', '.join(updates)} WHERE id=?", tuple(params))
            return {"ok": True, "id": fid}
        else:
            new_id = db_exec("INSERT INTO project_folders (name, icon, color, parent_id, created) VALUES (?,?,?,?,?)", (name, icon, color, parent_id, now))
            return {"ok": True, "id": new_id}

    def _project_folder_delete(self, body):
        fid = body.get("id", "")
        if fid:
            # 하위 폴더는 부모 NULL로 (최상단으로 이동), 프로젝트도 폴더 NULL로
            db_exec("UPDATE project_folders SET parent_id=NULL WHERE parent_id=?", (fid,))
            db_exec("UPDATE projects SET folder_id=NULL WHERE folder_id=?", (fid,))
            db_exec("DELETE FROM project_folders WHERE id=?", (fid,))
        return {"ok": True}

    # ── 날짜 폴더 자동 생성 (YYYY > YYYYMMDD_xxx 계층) ──
    def _project_date_folder(self, body):
        """오늘 날짜의 폴더 자동 생성 (없으면). 연도 폴더 → 일자 폴더 계층"""
        from datetime import datetime
        today = datetime.now()
        year = today.strftime("%Y")
        date = today.strftime("%Y%m%d")
        suffix = (body.get("suffix") or "").strip()
        date_folder_name = f"{date}_{suffix}" if suffix else date
        # 1) 연도 폴더 찾기/생성
        year_row = db_exec("SELECT id FROM project_folders WHERE name=? AND parent_id IS NULL", (year,), fetchone=True)
        if year_row:
            year_id = year_row["id"]
        else:
            year_id = db_exec("INSERT INTO project_folders (name, icon, color, parent_id, created) VALUES (?,?,?,?,?)",
                              (year, "📅", "#3B82F6", None, today.isoformat()))
        # 2) 일자 폴더 찾기/생성 (연도 폴더 안에)
        date_row = db_exec("SELECT id FROM project_folders WHERE name=? AND parent_id=?", (date_folder_name, year_id), fetchone=True)
        if date_row:
            date_id = date_row["id"]
        else:
            date_id = db_exec("INSERT INTO project_folders (name, icon, color, parent_id, created) VALUES (?,?,?,?,?)",
                              (date_folder_name, "📂", "", year_id, today.isoformat()))
        return {"ok": True, "yearFolderId": year_id, "dateFolderId": date_id, "dateFolderName": date_folder_name}

    # ── 프로젝트 첨부 (메모, 대화) ──
    def _proj_attach_list(self, params):
        pid = params.get("projectId", [""])[0]
        if not pid: return {"ok": True, "attachments": []}
        rows = db_exec("SELECT id, kind, target_id, sort_order FROM project_attachments WHERE project_id=? ORDER BY sort_order, id", (pid,), fetch=True)
        # enrich with target metadata
        out = []
        for r in rows:
            item = {"id": r["id"], "kind": r["kind"], "targetId": r["target_id"], "sortOrder": r["sort_order"]}
            try:
                if r["kind"] == "memo":
                    m = db_exec("SELECT name, substr(content,1,80) as preview FROM memos WHERE id=?", (int(r["target_id"]),), fetchone=True)
                    if m: item["title"] = m.get("name") or "무제"; item["preview"] = m.get("preview") or ""
                elif r["kind"] == "conversation":
                    c = db_exec("SELECT title, node_name FROM conversations WHERE id=?", (r["target_id"],), fetchone=True)
                    if c: item["title"] = c.get("title") or "대화"; item["preview"] = c.get("node_name") or ""
            except Exception as e:
                pass
            out.append(item)
        return {"ok": True, "attachments": out}

    def _proj_attach_add(self, body):
        pid = body.get("projectId")
        kind = body.get("kind")
        target_id = body.get("targetId")
        if not pid or not kind or target_id is None:
            return {"ok": False, "error": "projectId, kind, targetId required"}
        # 중복 방지
        existing = db_exec("SELECT id FROM project_attachments WHERE project_id=? AND kind=? AND target_id=?",
                           (pid, kind, str(target_id)), fetchone=True)
        if existing: return {"ok": True, "id": existing["id"], "duplicate": True}
        new_id = db_exec("INSERT INTO project_attachments (project_id, kind, target_id, created) VALUES (?,?,?,?)",
                         (pid, kind, str(target_id), datetime.now().isoformat()))
        return {"ok": True, "id": new_id}

    def _proj_attach_remove(self, body):
        iid = body.get("id")
        if iid:
            db_exec("DELETE FROM project_attachments WHERE id=?", (int(iid),))
        return {"ok": True}

    # ── 아이콘 라이브러리 (커스텀 업로드 이미지) ──
    def _icons_list(self):
        rows = db_exec("SELECT id, name, data, kind, created FROM icons ORDER BY id DESC LIMIT 200", fetch=True)
        return {"ok": True, "icons": rows or []}

    def _icon_save(self, body):
        name = body.get("name", "icon")
        data = body.get("data", "")  # data URL
        kind = body.get("kind", "image")
        if not data: return {"ok": False, "error": "data required"}
        # 5MB 제한
        if len(data) > 6_500_000: return {"ok": False, "error": "icon too large (>5MB base64)"}
        new_id = db_exec("INSERT INTO icons (name, data, kind, created) VALUES (?,?,?,?)",
                         (name, data, kind, datetime.now().isoformat()))
        return {"ok": True, "id": new_id}

    def _icon_delete(self, body):
        iid = body.get("id")
        if iid: db_exec("DELETE FROM icons WHERE id=?", (int(iid),))
        return {"ok": True}

    # ── Temp Saves (날짜 기준 자동 백업) ──

    def _temp_save(self, body):
        """일별 임시 저장 — 같은 날짜에 추가 저장 가능 (여러 개)"""
        from datetime import datetime
        today = datetime.now().strftime("%Y-%m-%d")
        now = datetime.now().isoformat()
        name = body.get("name", "")
        data = json.dumps(body, ensure_ascii=False)
        db_exec("INSERT INTO temps (name, data, date, created) VALUES (?,?,?,?)", (name, data, today, now))
        # 30일 이상 된 temp 자동 정리
        if DB_TYPE == "postgresql":
            db_exec("DELETE FROM temps WHERE created::timestamp < NOW() - INTERVAL '30 days'")
        else:
            db_exec("DELETE FROM temps WHERE created < datetime('now','-30 days')")
        return {"ok": True}

    def _temp_list(self):
        """temp 목록 (메타데이터만, 최신순)"""
        rows = db_exec("SELECT id, name, date, created FROM temps ORDER BY created DESC LIMIT 50", fetch=True)
        for r in rows:
            try:
                d = db_exec("SELECT data FROM temps WHERE id=?", (r["id"],), fetchone=True)
                if d:
                    pj = json.loads(d["data"])
                    r["nodeCount"] = len(pj.get("nodes", []))
                    r["connCount"] = len(pj.get("connections", []))
            except:
                r["nodeCount"] = 0
                r["connCount"] = 0
        return {"ok": True, "temps": rows}

    def _temp_load(self, params):
        tid = params.get("id", [""])[0]
        row = db_exec("SELECT data FROM temps WHERE id=?", (tid,), fetchone=True)
        if not row: return {"ok": False, "error": "not found"}
        return {"ok": True, "state": json.loads(row["data"])}

    def _temp_delete(self, body):
        tid = body.get("id", "")
        if not tid: return {"ok": False, "error": "id required"}
        row = db_exec("SELECT id, name, data, date, created FROM temps WHERE id=?", (tid,), fetchone=True)
        if row:
            trash_data = json.dumps(dict(row), ensure_ascii=False, default=str)
            db_exec("INSERT INTO trash (original_table, original_id, name, data, deleted_at) VALUES (?,?,?,?,?)",
                    ("temps", str(tid), row.get("name", ""), trash_data, datetime.now().isoformat()))
        db_exec("DELETE FROM temps WHERE id=?", (tid,))
        return {"ok": True}

    # ── Memo & Folders ──

    def _memo_list(self, params):
        """메모 목록 (folder_id 또는 is_temp 필터)"""
        folder_id = params.get("folderId", [None])[0]
        is_temp = params.get("isTemp", [None])[0]
        sql = "SELECT id, name, folder_id, is_temp, pinned, color, substr(content,1,100) as preview, created, modified FROM memos WHERE name!='__scratchpad__'"
        cond = []
        args = []
        if folder_id is not None:
            if folder_id == "null":
                cond.append("folder_id IS NULL")
            else:
                cond.append("folder_id=?")
                args.append(folder_id)
        if is_temp is not None:
            cond.append("is_temp=?")
            args.append(int(is_temp))
        if cond: sql += " AND " + " AND ".join(cond)
        sql += " ORDER BY modified DESC LIMIT 200"
        rows = db_exec(sql, tuple(args), fetch=True)
        return {"ok": True, "memos": rows}

    def _memo_get(self, params):
        mid = params.get("id", [""])[0]
        row = db_exec("SELECT * FROM memos WHERE id=?", (mid,), fetchone=True)
        if not row: return {"ok": False, "error": "not found"}
        return {"ok": True, "memo": row}

    def _memo_save(self, body):
        """⚠ CRITICAL: 부분 업데이트 — body에 명시된 필드만 갱신.
        과거: name/content가 body에 없으면 ""로 덮어써서 데이터 소실됨 (드래그 이동 시).
        지금: body에 있는 필드만 UPDATE."""
        mid = body.get("id")
        now = datetime.now().isoformat()
        if mid:
            # 안전: 기존 행 백업 (UPDATE 전 휴지통에 스냅샷)
            try:
                existing_row = db_exec("SELECT id, name, content, folder_id, is_temp, pinned, color, created, modified FROM memos WHERE id=?", (mid,), fetchone=True)
                if existing_row and ("name" in body or "content" in body):
                    # name/content를 갱신하는 경우만 백업 (드래그 이동 등 메타만 바뀌면 백업 불필요)
                    snap = json.dumps(dict(existing_row), ensure_ascii=False, default=str)
                    db_exec("INSERT INTO trash (original_table, original_id, name, data, deleted_at) VALUES (?,?,?,?,?)",
                            ("memos_pre_update", str(mid), existing_row.get("name", ""), snap, now))
            except Exception as e:
                log(f"[MEMO_SAVE] backup failed (non-fatal): {e}")
            # 부분 UPDATE — body에 있는 필드만
            updates, params = ["modified=?"], [now]
            if "name" in body:
                updates.append("name=?"); params.append(body.get("name") or "")
            if "content" in body:
                updates.append("content=?"); params.append(body.get("content") or "")
            if "folderId" in body:
                fid = body.get("folderId")
                updates.append("folder_id=?"); params.append(fid if fid else None)
            if "isTemp" in body:
                updates.append("is_temp=?"); params.append(1 if body.get("isTemp") else 0)
            if "pinned" in body:
                updates.append("pinned=?"); params.append(1 if body.get("pinned") else 0)
            if "color" in body:
                updates.append("color=?"); params.append(body.get("color") or "")
            params.append(mid)
            db_exec(f"UPDATE memos SET {', '.join(updates)} WHERE id=?", tuple(params))
            log(f"[MEMO_SAVE] partial update id={mid} fields={list(body.keys())}")
            return {"ok": True, "id": mid}
        else:
            # 신규 생성 — folder 자동 처리
            name = body.get("name", "") or f"메모{int(time.time())}"
            content = body.get("content", "")
            is_temp = 1 if body.get("isTemp", True) else 0
            folder_id = body.get("folderId")
            if not folder_id and not is_temp:
                today = datetime.now().strftime("%Y-%m-%d")
                existing = db_exec("SELECT id FROM memo_folders WHERE name=?", (today,), fetchone=True)
                if existing:
                    folder_id = existing["id"]
                else:
                    folder_id = db_exec("INSERT INTO memo_folders (name, icon, color, created) VALUES (?,?,?,?)",
                                       (today, "📅", "", now))
            new_id = db_exec("INSERT INTO memos (name, content, folder_id, is_temp, created, modified) VALUES (?,?,?,?,?,?)",
                             (name, content, folder_id, is_temp, now, now))
            return {"ok": True, "id": new_id}

    def _memo_delete(self, body):
        mid = body.get("id", "")
        if not mid: return {"ok": False, "error": "id required"}
        row = db_exec("SELECT id, name, content, folder_id, is_temp, pinned, color, created, modified FROM memos WHERE id=?", (mid,), fetchone=True)
        if row:
            trash_data = json.dumps(dict(row), ensure_ascii=False, default=str)
            db_exec("INSERT INTO trash (original_table, original_id, name, data, deleted_at) VALUES (?,?,?,?,?)",
                    ("memos", str(mid), row.get("name", ""), trash_data, datetime.now().isoformat()))
        db_exec("DELETE FROM memos WHERE id=?", (mid,))
        return {"ok": True}

    def _folder_list(self):
        rows = db_exec("SELECT f.id, f.name, f.icon, f.sort_order, f.color, f.parent_id, f.created, (SELECT COUNT(*) FROM memos WHERE folder_id=f.id) as memo_count FROM memo_folders f ORDER BY sort_order, name", fetch=True)
        return {"ok": True, "folders": rows}

    def _folder_save(self, body):
        fid = body.get("id")
        now = datetime.now().isoformat()
        if fid:
            # Partial update: only update fields explicitly provided in body.
            updates = []
            params = []
            if "name" in body:
                updates.append("name=?"); params.append(body.get("name") or "새 폴더")
            if "icon" in body:
                updates.append("icon=?"); params.append(body.get("icon") or "📁")
            if "color" in body:
                updates.append("color=?"); params.append(body.get("color") or "")
            if "parentId" in body:
                pid = body.get("parentId")
                updates.append("parent_id=?"); params.append(pid if pid else None)
            if updates:
                params.append(fid)
                db_exec(f"UPDATE memo_folders SET {', '.join(updates)} WHERE id=?", tuple(params))
            return {"ok": True, "id": fid}
        else:
            name = body.get("name", "새 폴더")
            icon = body.get("icon", "📁")
            color = body.get("color", "")
            parent_id = body.get("parentId")
            if parent_id in ("", 0): parent_id = None
            new_id = db_exec("INSERT INTO memo_folders (name, icon, color, parent_id, created) VALUES (?,?,?,?,?)", (name, icon, color, parent_id, now))
            return {"ok": True, "id": new_id}

    def _folder_delete(self, body):
        fid = body.get("id", "")
        if fid:
            db_exec("UPDATE memos SET folder_id=NULL WHERE folder_id=?", (fid,))
            db_exec("DELETE FROM memo_folders WHERE id=?", (fid,))
        return {"ok": True}

    def _project_delete(self, body):
        pid = body.get("id", "")
        if not pid: return {"ok": False, "error": "id required"}
        row = db_exec("SELECT id, name, data, created, modified, favorite, folder_id FROM projects WHERE id=?", (pid,), fetchone=True)
        if row:
            trash_data = json.dumps(dict(row), ensure_ascii=False, default=str)
            db_exec("INSERT INTO trash (original_table, original_id, name, data, deleted_at) VALUES (?,?,?,?,?)",
                    ("projects", str(pid), row.get("name", ""), trash_data, datetime.now().isoformat()))
        db_exec("DELETE FROM projects WHERE id=?", (pid,))
        return {"ok": True}

    # ── Trash (휴지통) ──

    def _trash_list(self):
        rows = db_exec("SELECT id, original_table, original_id, name, deleted_at FROM trash ORDER BY deleted_at DESC LIMIT 100", fetch=True)
        return {"ok": True, "items": rows}

    def _trash_restore(self, body):
        tid = body.get("id")
        if not tid: return {"ok": False, "error": "id required"}
        row = db_exec("SELECT * FROM trash WHERE id=?", (tid,), fetchone=True)
        if not row: return {"ok": False, "error": "not found"}
        data = json.loads(row["data"])
        table = row["original_table"]
        if table == "projects":
            db_exec("INSERT INTO projects (id, name, data, created, modified, favorite, folder_id) VALUES (?,?,?,?,?,?,?)",
                    (data.get("id"), data.get("name",""), data.get("data","{}"), data.get("created",""), data.get("modified",""), data.get("favorite",0), data.get("folder_id")))
        elif table == "temps":
            db_exec("INSERT INTO temps (name, data, date, created) VALUES (?,?,?,?)",
                    (data.get("name",""), data.get("data","{}"), data.get("date",""), data.get("created","")))
        elif table == "memos":
            db_exec("INSERT INTO memos (name, content, folder_id, is_temp, pinned, color, created, modified) VALUES (?,?,?,?,?,?,?,?)",
                    (data.get("name",""), data.get("content",""), data.get("folder_id"), data.get("is_temp",1), data.get("pinned",0), data.get("color",""), data.get("created",""), data.get("modified","")))
        db_exec("DELETE FROM trash WHERE id=?", (tid,))
        return {"ok": True}

    def _trash_delete(self, body):
        tid = body.get("id")
        if tid: db_exec("DELETE FROM trash WHERE id=?", (tid,))
        return {"ok": True}

    def _trash_empty(self, body):
        db_exec("DELETE FROM trash")
        return {"ok": True}

    # ── 임시 프로젝트 폴더 (저장 안 한 작업물) ──

    def _temp_folder_list(self):
        """임시 폴더(살아있는 것) + 휴지통 폴더(유예 중) 목록."""
        def _scan(root, kind):
            items = []
            if not os.path.isdir(root):
                return items
            try:
                for name in sorted(os.listdir(root)):
                    full = os.path.join(root, name)
                    if not os.path.isdir(full) or name.startswith("."):
                        continue
                    try:
                        mt = os.path.getmtime(full)
                        size_bytes = 0
                        file_count = 0
                        for dirpath, _dirs, files in os.walk(full):
                            for fn in files:
                                try:
                                    size_bytes += os.path.getsize(os.path.join(dirpath, fn))
                                    file_count += 1
                                except Exception: pass
                        snap = None
                        snap_path = os.path.join(full, "snapshot.json")
                        if os.path.exists(snap_path):
                            try:
                                with open(snap_path, "r", encoding="utf-8") as f:
                                    j = json.load(f)
                                c = j.get("canvas") or {}
                                snap = {
                                    "hasSnapshot": True,
                                    "nodeCount": len((c.get("nodes") or [])),
                                    "connectionCount": len((c.get("connections") or [])),
                                    "savedAt": j.get("savedAt"),
                                    "name": c.get("wfName") or c.get("name"),
                                }
                            except Exception:
                                snap = {"hasSnapshot": True, "error": "스냅샷 파싱 실패"}
                        items.append({
                            "kind": kind,
                            "name": name,
                            "path": full,
                            "mtime": datetime.fromtimestamp(mt).isoformat(),
                            "ageHours": round((time.time() - mt) / 3600, 1),
                            "sizeBytes": size_bytes,
                            "fileCount": file_count,
                            "snapshot": snap,
                        })
                    except Exception:
                        continue
            except PermissionError:
                pass
            return items
        alive = _scan(TEMP_ROOT, "temp")
        trashed = _scan(TRASH_ROOT, "trash")
        return {"ok": True, "alive": alive, "trashed": trashed,
                "tempTtlDays": TEMP_TTL_DAYS, "trashTtlDays": TRASH_TTL_DAYS}

    def _temp_folder_snapshot(self, params):
        """임시/휴지통 폴더의 snapshot.json 원본 반환 (복원 미리보기)."""
        path = (params.get("path", [""])[0] or "").strip()
        if not path or not os.path.isdir(path):
            return {"ok": False, "error": "경로 없음"}
        # 보안: TEMP_ROOT 또는 TRASH_ROOT 하위만 허용
        norm = os.path.abspath(path)
        if not (norm.startswith(os.path.abspath(TEMP_ROOT)) or
                norm.startswith(os.path.abspath(TRASH_ROOT))):
            return {"ok": False, "error": "접근 거부"}
        snap_path = os.path.join(path, "snapshot.json")
        if not os.path.exists(snap_path):
            return {"ok": False, "error": "snapshot.json 없음"}
        try:
            with open(snap_path, "r", encoding="utf-8") as f:
                j = json.load(f)
            return {"ok": True, "snapshot": j}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _temp_folder_restore(self, body):
        """휴지통의 임시 폴더를 살아있는 임시 폴더로 되살림."""
        path = (body.get("path") or "").strip()
        if not path or not os.path.isdir(path):
            return {"ok": False, "error": "경로 없음"}
        norm = os.path.abspath(path)
        if not norm.startswith(os.path.abspath(TRASH_ROOT)):
            return {"ok": False, "error": "휴지통 항목이 아님"}
        # __trashed-xxx 접미사 제거한 원래 이름으로 복원
        base = os.path.basename(path)
        orig = re.sub(r"__trashed-\d{8}_\d{6}$", "", base)
        target = os.path.join(TEMP_ROOT, orig)
        if os.path.exists(target):
            target = os.path.join(TEMP_ROOT, f"{orig}__restored-{datetime.now().strftime('%H%M%S')}")
        try:
            os.makedirs(TEMP_ROOT, exist_ok=True)
            shutil.move(path, target)
            # DB 휴지통 레코드 제거
            try:
                db_exec("DELETE FROM trash WHERE original_table='temps_folder' AND data LIKE ?",
                        (f'%"path": "{path}"%',))
            except Exception: pass
            # mtime 새로고침 (TTL 재시작)
            os.utime(target, None)
            log(f"[TRASH→TEMP] {path} → {target}")
            return {"ok": True, "path": target}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _temp_folder_promote(self, body):
        """임시/휴지통 폴더를 정식 프로젝트로 승격.
        body: {path, name?, date?}
        """
        path = (body.get("path") or "").strip()
        if not path or not os.path.isdir(path):
            return {"ok": False, "error": "경로 없음"}
        norm = os.path.abspath(path)
        if not (norm.startswith(os.path.abspath(TEMP_ROOT)) or
                norm.startswith(os.path.abspath(TRASH_ROOT))):
            return {"ok": False, "error": "접근 거부"}
        # snapshot.json 에서 캔버스 state 복원
        canvas_state = None
        snap_path = os.path.join(path, "snapshot.json")
        if os.path.exists(snap_path):
            try:
                with open(snap_path, "r", encoding="utf-8") as f:
                    j = json.load(f)
                canvas_state = j.get("canvas")
            except Exception: pass
        # 이름 / 날짜 결정
        name = (body.get("name") or "").strip()
        if not name and canvas_state:
            name = canvas_state.get("wfName") or canvas_state.get("name") or ""
        if not name:
            base = os.path.basename(path)
            base = re.sub(r"__trashed-\d{8}_\d{6}$", "", base)
            base = re.sub(r"__[a-zA-Z0-9]{1,8}$", "", base)
            name = base or "Untitled"
        date_override = None
        date_str = (body.get("date") or "").strip()
        if date_str:
            for fmt in ("%Y-%m-%d", "%Y%m%d", "%Y/%m/%d"):
                try:
                    date_override = datetime.strptime(date_str, fmt); break
                except ValueError: continue
        target = compute_work_dir(name, date_override=date_override)
        # 최종 저장 (이동 + DB 레코드 + project.json)
        payload = dict(canvas_state or {})
        payload["name"] = name
        save_body = {"name": name, "saveDate": date_str or None,
                     "promoteFromTemp": path, **payload}
        return self._project_save(save_body)

    def _temp_folder_delete_now(self, body):
        """임시/휴지통 폴더 즉시 영구 삭제."""
        path = (body.get("path") or "").strip()
        if not path or not os.path.isdir(path):
            return {"ok": False, "error": "경로 없음"}
        norm = os.path.abspath(path)
        if not (norm.startswith(os.path.abspath(TEMP_ROOT)) or
                norm.startswith(os.path.abspath(TRASH_ROOT))):
            return {"ok": False, "error": "접근 거부"}
        try:
            shutil.rmtree(path)
            try:
                db_exec("DELETE FROM trash WHERE original_table='temps_folder' AND data LIKE ?",
                        (f'%"path": "{path}"%',))
            except Exception: pass
            log(f"[TEMP_DELETE] {path}")
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _temp_cleanup_now(self, body):
        """수동으로 자동정리 한 번 돌림."""
        return {"ok": True, **cleanup_temp_and_trash()}

    # ── File Upload (이미지/시스템프롬프트 파일) ──

    def _upload_file(self, body):
        """base64로 인코딩된 파일을 WSL에 저장하고 경로 반환"""
        import base64
        filename = body.get("filename", "")
        b64data = body.get("data", "")
        purpose = body.get("purpose", "image")  # image | sysprompt
        if not filename or not b64data:
            return {"ok": False, "error": "filename and data required"}

        # 디렉토리: ~/tmux-controller/uploads/(images|sysprompts)/
        upload_dir = os.path.join(UPLOADS_DIR, purpose + "s")
        os.makedirs(upload_dir, exist_ok=True)

        # 파일명 sanitize + timestamp 추가 (덮어쓰기 방지)
        safe_name = re.sub(r'[^\w\s.\-_가-힣]', '_', filename)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        final_name = f"{ts}_{safe_name}"
        filepath = os.path.join(upload_dir, final_name)

        try:
            # base64 디코딩 (data:image/png;base64,xxx 형식 지원)
            if ',' in b64data:
                b64data = b64data.split(',', 1)[1]
            data = base64.b64decode(b64data)
            with open(filepath, "wb") as f:
                f.write(data)
            log(f"UPLOAD [{purpose}] {filename} → {filepath} ({len(data)} bytes)")
            # url 필드 = 브라우저가 <img src=> 로 쓸 수 있는 상대 URL
            # path 필드 = 파일시스템 절대경로 (Claude/Gemini CLI 에 첨부용)
            url = f"/uploads/{purpose}s/{final_name}"
            return {"ok": True, "path": filepath, "url": url, "filename": final_name}
        except Exception as e:
            log(f"UPLOAD ERROR: {e}")
            return {"ok": False, "error": str(e)}

    def _list_pdf_images(self, body):
        """이미지 경로에서 부모 PDF 디렉토리의 모든 이미지 목록 반환"""
        img_path = body.get("path", "")
        if not img_path or not os.path.exists(img_path):
            return {"ok": False, "error": "path not found"}
        pdf_dir = os.path.dirname(img_path)
        if not os.path.basename(pdf_dir).startswith(("20", "19")):
            # not a pdf split dir
            return {"ok": True, "images": []}
        try:
            files = sorted([f for f in os.listdir(pdf_dir) if f.startswith("page_") and f.endswith(".png")])
            images = [{"path": os.path.join(pdf_dir, f), "name": f} for f in files]
            return {"ok": True, "images": images, "dir": pdf_dir}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    # ── PDF Split → 이미지 변환 ──

    def _parse_page_range(self, spec, total):
        """페이지 범위 파싱 '1,5,7-10' → [1,5,7,8,9,10]"""
        if not spec or spec.strip() == "":
            return list(range(1, total + 1))
        pages = []
        for part in spec.replace(" ", "").split(","):
            if "-" in part:
                a, b = part.split("-", 1)
                a = int(a) if a else 1
                b = int(b) if b else total
                pages.extend(range(a, min(b, total) + 1))
            else:
                p = int(part)
                if 1 <= p <= total: pages.append(p)
        return sorted(set(pages))

    def _pdf_split(self, body):
        """PDF base64 → 페이지별 PNG 이미지 변환 (600 DPI)"""
        import base64
        filename = body.get("filename", "doc.pdf")
        b64data = body.get("data", "")
        page_spec = body.get("pages", "")  # "1,5,7-10" or empty for all
        dpi = int(body.get("dpi", 600))

        if not b64data:
            return {"ok": False, "error": "data required"}

        try:
            import sys
            sys.path.insert(0, os.path.expanduser("~/.local/lib/python3.12/site-packages"))
            import fitz
        except Exception as e:
            log(f"PDF_SPLIT pymupdf import FAIL: {e}")
            return {"ok": False, "error": "pymupdf not installed"}

        # 디코딩
        try:
            if "," in b64data:
                b64data = b64data.split(",", 1)[1]
            pdf_bytes = base64.b64decode(b64data)
        except Exception as e:
            return {"ok": False, "error": f"decode failed: {e}"}

        # 임시 PDF 저장 + 변환
        upload_dir = os.path.join(UPLOADS_DIR, "pdfs")
        os.makedirs(upload_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = re.sub(r'[^\w\s.\-_가-힣]', '_', filename).replace(".pdf", "")
        pdf_dir = os.path.join(upload_dir, f"{ts}_{safe_name}")
        os.makedirs(pdf_dir, exist_ok=True)
        pdf_path = os.path.join(pdf_dir, "source.pdf")
        with open(pdf_path, "wb") as f:
            f.write(pdf_bytes)

        try:
            doc = fitz.open(pdf_path)
            total_pages = len(doc)
            target_pages = self._parse_page_range(page_spec, total_pages)
            log(f"PDF_SPLIT {filename} total={total_pages} target={target_pages} dpi={dpi}")

            results = []
            for pnum in target_pages:
                page = doc[pnum - 1]
                # DPI 설정
                mat = fitz.Matrix(dpi / 72, dpi / 72)
                pix = page.get_pixmap(matrix=mat)
                img_path = os.path.join(pdf_dir, f"page_{pnum:03d}.png")
                pix.save(img_path)
                results.append({"page": pnum, "path": img_path})
            doc.close()

            log(f"PDF_SPLIT done: {len(results)} pages")
            return {"ok": True, "totalPages": total_pages, "pages": results, "pdfPath": pdf_path}
        except Exception as e:
            log(f"PDF_SPLIT ERROR: {e}")
            return {"ok": False, "error": str(e)}

    # ── Auto State (항상 최신 상태를 DB에 보관) ──

    # ── Scratchpad ──

    def _scratchpad_load(self):
        row = db_exec("SELECT content FROM memos WHERE name='__scratchpad__' LIMIT 1", fetchone=True)
        return {"ok": True, "content": row["content"] if row else ""}

    def _scratchpad_save(self, body):
        content = body.get("content", "")
        now = datetime.now().isoformat()
        existing = db_exec("SELECT id FROM memos WHERE name='__scratchpad__'", fetchone=True)
        if existing:
            db_exec("UPDATE memos SET content=?, modified=? WHERE name='__scratchpad__'", (content, now))
        else:
            db_exec("INSERT INTO memos (name, content, is_temp, created, modified) VALUES (?,?,0,?,?)",
                    ("__scratchpad__", content, now, now))
        return {"ok": True}

    # ── Memo Pin ──

    def _memo_pinned(self):
        rows = db_exec("SELECT id, name, substr(content,1,40) as preview FROM memos WHERE pinned=1 AND name!='__scratchpad__' ORDER BY modified DESC LIMIT 5", fetch=True)
        return {"ok": True, "memos": rows}

    def _memo_pin(self, body):
        mid = body.get("id", "")
        pinned = 1 if body.get("pinned", True) else 0
        if mid:
            db_exec("UPDATE memos SET pinned=? WHERE id=?", (pinned, mid))
        return {"ok": True}

    # ── 즐겨찾기 폴더 (메모와 별개 카테고리, 같은 메모가 여러 폴더에 들어갈 수 있음) ──

    def _fav_folders(self, params):
        parent_id = params.get("parentId", [""])[0]
        kind = params.get("kind", ["project"])[0]  # 기본: 프로젝트 즐겨찾기
        if parent_id and parent_id != "null":
            rows = db_exec("SELECT id, name, parent_id, color, icon, sort_order, kind FROM fav_folders WHERE parent_id=? AND (kind=? OR kind IS NULL) ORDER BY sort_order, id", (int(parent_id), kind), fetch=True)
        else:
            rows = db_exec("SELECT id, name, parent_id, color, icon, sort_order, kind FROM fav_folders WHERE parent_id IS NULL AND (kind=? OR kind IS NULL) ORDER BY sort_order, id", (kind,), fetch=True)
        return {"ok": True, "folders": rows or []}

    def _fav_folder_save(self, body):
        fid = body.get("id")
        now = datetime.now().isoformat()
        if fid:
            updates = []
            params = []
            if "name" in body:
                updates.append("name=?"); params.append(body.get("name") or "새 폴더")
            if "color" in body:
                updates.append("color=?"); params.append(body.get("color") or "")
            if "icon" in body:
                updates.append("icon=?"); params.append(body.get("icon") or "⭐")
            if "parentId" in body:
                pid = body.get("parentId")
                updates.append("parent_id=?"); params.append(pid if pid else None)
            if "sortOrder" in body:
                updates.append("sort_order=?"); params.append(int(body.get("sortOrder", 0)))
            if updates:
                params.append(fid)
                db_exec(f"UPDATE fav_folders SET {', '.join(updates)} WHERE id=?", tuple(params))
            return {"ok": True, "id": fid}
        else:
            name = body.get("name", "새 폴더")
            color = body.get("color", "")
            icon = body.get("icon", "⭐")
            kind = body.get("kind", "project")
            parent_id = body.get("parentId")
            if parent_id in ("", 0, None): parent_id = None
            new_id = db_exec("INSERT INTO fav_folders (name, parent_id, color, icon, kind, created) VALUES (?,?,?,?,?,?)", (name, parent_id, color, icon, kind, now))
            return {"ok": True, "id": new_id}

    def _fav_folder_delete(self, body):
        fid = body.get("id")
        if not fid: return {"ok": False, "error": "id required"}
        # 하위 폴더는 부모 NULL로 (최상단으로 이동), 항목은 삭제
        db_exec("UPDATE fav_folders SET parent_id=NULL WHERE parent_id=?", (fid,))
        db_exec("DELETE FROM fav_items WHERE folder_id=?", (fid,))
        db_exec("DELETE FROM fav_folders WHERE id=?", (fid,))
        return {"ok": True}

    def _fav_items(self, params):
        fid = params.get("folderId", [""])[0]
        if not fid: return {"ok": True, "items": []}
        # 양쪽 호환: 신규는 target_id+kind, 레거시는 memo_id (kind='memo'로 간주)
        rows = db_exec("""
            SELECT id, memo_id, target_id, kind, sort_order
            FROM fav_items WHERE folder_id=? ORDER BY sort_order, id
        """, (int(fid),), fetch=True) or []
        out = []
        for r in rows:
            kind = r.get("kind") or ('memo' if r.get("memo_id") else 'project')
            tid = r.get("target_id") or (str(r.get("memo_id")) if r.get("memo_id") else None)
            if not tid: continue
            item = {"id": r["id"], "kind": kind, "targetId": tid, "sortOrder": r["sort_order"]}
            try:
                if kind == 'memo':
                    m = db_exec("SELECT name, substr(content,1,80) as preview FROM memos WHERE id=?", (int(tid),), fetchone=True)
                    if m: item["title"] = m.get("name") or "무제"; item["preview"] = m.get("preview") or ""
                elif kind == 'project':
                    p = db_exec("SELECT name, modified FROM projects WHERE id=?", (tid,), fetchone=True)
                    if p: item["title"] = p.get("name") or "무제"; item["preview"] = (p.get("modified") or "")[:16]
                elif kind == 'conversation' or kind == 'chat':
                    c = db_exec("SELECT title, node_name, created FROM conversations WHERE id=?", (tid,), fetchone=True)
                    if c:
                        item["title"] = c.get("title") or "(제목 없는 대화)"
                        item["preview"] = f"노드: {c.get('node_name', '')} · {(c.get('created') or '')[:16]}"
                        item["nodeName"] = c.get("node_name", "")
            except Exception:
                pass
            out.append(item)
        return {"ok": True, "items": out}

    def _fav_item_add(self, body):
        fid = body.get("folderId")
        kind = body.get("kind", "project")
        target_id = body.get("targetId") or body.get("memoId") or body.get("projectId")
        if not fid or target_id is None: return {"ok": False, "error": "folderId and targetId required"}
        target_id = str(target_id)
        # 중복 방지
        existing = db_exec("SELECT id FROM fav_items WHERE folder_id=? AND kind=? AND target_id=?",
                           (int(fid), kind, target_id), fetchone=True)
        if existing: return {"ok": True, "id": existing["id"], "duplicate": True}
        # ⚠ memo_id가 NOT NULL 제약이라 0 placeholder 사용 (project/non-memo 항목)
        # 실제 식별은 target_id+kind 사용
        memo_id = int(target_id) if kind == 'memo' and str(target_id).isdigit() else 0
        try:
            new_id = db_exec("INSERT INTO fav_items (folder_id, memo_id, target_id, kind, created) VALUES (?,?,?,?,?)",
                             (int(fid), memo_id, target_id, kind, datetime.now().isoformat()))
            log(f"[FAV] item added: folder={fid} kind={kind} target={target_id} → id={new_id}")
            return {"ok": True, "id": new_id}
        except Exception as e:
            log(f"[FAV] INSERT failed: {e}")
            # NOT NULL 제약 시도 우회 — 컬럼이 nullable로 안 바뀐 케이스
            try:
                # NOT NULL을 nullable로 ALTER (PG)
                if _is_pg():
                    db_exec("ALTER TABLE fav_items ALTER COLUMN memo_id DROP NOT NULL")
                # 재시도
                new_id = db_exec("INSERT INTO fav_items (folder_id, memo_id, target_id, kind, created) VALUES (?,?,?,?,?)",
                                 (int(fid), memo_id, target_id, kind, datetime.now().isoformat()))
                return {"ok": True, "id": new_id}
            except Exception as e2:
                return {"ok": False, "error": f"INSERT failed: {e2}"}

    def _fav_item_remove(self, body):
        iid = body.get("id")
        fid = body.get("folderId")
        kind = body.get("kind")
        target_id = body.get("targetId") or body.get("memoId")
        if iid:
            db_exec("DELETE FROM fav_items WHERE id=?", (int(iid),))
        elif fid and target_id is not None:
            if kind:
                db_exec("DELETE FROM fav_items WHERE folder_id=? AND kind=? AND target_id=?", (int(fid), kind, str(target_id)))
            else:
                db_exec("DELETE FROM fav_items WHERE folder_id=? AND (target_id=? OR memo_id=?)", (int(fid), str(target_id), int(target_id) if str(target_id).isdigit() else -1))
        return {"ok": True}

    def _state_save(self, body):
        """자동 저장 — workflowId 있으면 해당 프로젝트 갱신, 없으면 탭별 임시 슬롯
        (멀티 탭 분리: 각 탭이 sessionStorage tabId를 보내서 서로 안 섞임)"""
        now = datetime.now().isoformat()
        data = json.dumps(body, ensure_ascii=False)
        wf_id = body.get("workflowId")
        tab_id = body.get("tabId", "default")
        if wf_id:
            existing = db_exec("SELECT id FROM projects WHERE id=?", (wf_id,), fetchone=True)
            if existing:
                db_exec("UPDATE projects SET data=?, modified=? WHERE id=?", (data, now, wf_id))
                return {"ok": True, "savedTo": wf_id}
        recovery_id = f"__current_{tab_id}__"
        existing = db_exec("SELECT id FROM projects WHERE id=?", (recovery_id,), fetchone=True)
        if existing:
            db_exec("UPDATE projects SET data=?, modified=? WHERE id=?", (data, now, recovery_id))
        else:
            db_exec("INSERT INTO projects (id, name, data, created, modified) VALUES (?,?,?,?,?)",
                    (recovery_id, "__current__", data, now, now))
        return {"ok": True, "savedTo": recovery_id}

    def _state_load(self, params=None):
        """탭별 복구 슬롯에서 상태 복원 (없으면 legacy __current__ fallback)"""
        tab_id = "default"
        if params:
            v = params.get("tabId", ["default"])
            tab_id = v[0] if v else "default"
        recovery_id = f"__current_{tab_id}__"
        row = db_exec("SELECT data FROM projects WHERE id=?", (recovery_id,), fetchone=True)
        if not row:
            row = db_exec("SELECT data FROM projects WHERE id='__current__'", fetchone=True)
        if row:
            return {"ok": True, "state": json.loads(row["data"])}
        return {"ok": True, "state": None}

    # ── tmux Controls ──

    def _send_to_pane(self, target, command):
        cmd_len = len(command)
        if '\n' in command or cmd_len > 200:
            CHUNK = 500
            for i in range(0, cmd_len, CHUNK):
                r = run_tmux("send-keys", "-l", "-t", target, command[i:i+CHUNK])
                if not r["ok"]: return r
        else:
            r = run_tmux("send-keys", "-l", "-t", target, command)
            if not r["ok"]: return r
        return run_tmux("send-keys", "-t", target, "Enter")

    def _send_command(self, body):
        pane = body.get("pane", "all")
        command = body.get("command", "")
        if not command: return {"ok": False, "error": "no command"}
        if pane == "all":
            r = run_tmux("list-panes", "-t", SESSION_NAME, "-F", "#{pane_index}")
            if not r["ok"] or not r["stdout"]: return r
            results = [self._send_to_pane(f"{SESSION_NAME}:.{idx.strip()}", command) for idx in r["stdout"].strip().split("\n")]
            failed = [r for r in results if not r["ok"]]
            return {"ok": len(failed) == 0, "sent": len(results), "failed": len(failed)}
        return self._send_to_pane(f"{SESSION_NAME}:.{pane}", command)

    def _split(self, body):
        d = body.get("direction", "h")
        return run_tmux("split-window", "-h" if d == "h" else "-v", "-t", body.get("target", SESSION_NAME))

    def _layout(self, body):
        return run_tmux("select-layout", "-t", SESSION_NAME, body.get("layout", "tiled"))

    def _kill_pane(self, body):
        return run_tmux("kill-pane", "-t", f"{SESSION_NAME}:.{body.get('pane', '')}")

    def _new_window(self, body):
        args = ["new-window", "-t", SESSION_NAME]
        name = body.get("name", "")
        if name: args += ["-n", name]
        return run_tmux(*args)

    def _select_window(self, body):
        return run_tmux("select-window", "-t", f"{SESSION_NAME}:{body.get('index', 0)}")

    def _preset(self, body):
        count = body.get("count", 2)
        layout = body.get("layout", "tiled")
        r = run_tmux("list-panes", "-t", SESSION_NAME, "-F", "#{pane_index}")
        if not r["ok"]: return r
        current = len(r["stdout"].strip().split("\n")) if r["stdout"].strip() else 1
        for _ in range(count - current):
            run_tmux("split-window", "-t", SESSION_NAME)
            run_tmux("select-layout", "-t", SESSION_NAME, "tiled")
        while current > count:
            run_tmux("kill-pane", "-t", f"{SESSION_NAME}:.{current - 1}")
            current -= 1
        return run_tmux("select-layout", "-t", SESSION_NAME, layout)

    def _add_pane(self, body):
        r = run_tmux("split-window", "-t", SESSION_NAME)
        if r["ok"]:
            run_tmux("select-layout", "-t", SESSION_NAME, "tiled")
            r2 = run_tmux("list-panes", "-t", SESSION_NAME, "-F", "#{pane_index}")
            if r2["ok"]:
                indices = [int(x) for x in r2["stdout"].strip().split("\n")]
                return {"ok": True, "paneIndex": max(indices), "total": len(indices)}
        return r

    def _setup_session(self, body):
        """N개 패널 생성 + 각 패널에 claude 자동 실행"""
        count = int(body.get("count", 1))
        launch_claude = body.get("launchClaude", True)
        log(f"SETUP_SESSION count={count} launchClaude={launch_claude}")

        # 1. 세션 리셋 (main 세션만 삭제, 서버 세션은 유지)
        run_tmux("kill-session", "-t", SESSION_NAME)
        time.sleep(1)
        r = run_tmux("new-session", "-d", "-s", SESSION_NAME, "-c", os.path.expanduser("~"))
        if not r["ok"]:
            return {"ok": False, "error": "session create failed"}

        # 2. count - 1개 패널 추가 (이미 1개 있음)
        for _ in range(count - 1):
            run_tmux("split-window", "-t", SESSION_NAME)
            run_tmux("select-layout", "-t", SESSION_NAME, "tiled")
        run_tmux("select-layout", "-t", SESSION_NAME, "tiled")

        # 3. 각 패널에 claude 명령 전송
        if launch_claude:
            # -p 한번 실행으로 trust 디렉토리 사전 등록
            try:
                env = get_claude_env()
                subprocess.run(
                    ["claude", "-p", "ok", "--dangerously-skip-permissions"],
                    capture_output=True, text=True, timeout=15, env=env, cwd=os.path.expanduser("~")
                )
                log("SETUP_SESSION trust pre-registered via -p")
            except Exception as e:
                log(f"SETUP_SESSION trust pre-register failed: {e}")

            time.sleep(0.5)
            r = run_tmux("list-panes", "-t", SESSION_NAME, "-F", "#{pane_index}")
            if r["ok"]:
                pane_indices = [idx.strip() for idx in r["stdout"].strip().split("\n")]
                for idx in pane_indices:
                    target = f"{SESSION_NAME}:.{idx}"
                    cmd = "source ~/.nvm/nvm.sh 2>/dev/null; claude --dangerously-skip-permissions"
                    run_tmux("send-keys", "-t", target, cmd, "Enter")
                # 백그라운드: trust 다이얼로그 적극 자동 승인
                def auto_trust():
                    # 60초 동안 모든 패널에서 trust/permission 프롬프트 감지 → Enter로 자동 승인
                    trust_keywords = ["trust this folder", "yes, i trust", "do you trust",
                                      "safety check", "trust this project"]
                    done_keywords = ["bypass permissions"]  # claude에 진입 완료된 상태 (프롬프트 "❯" 가 보이는 상태)
                    done = set()
                    for sec in range(60):
                        time.sleep(1)
                        if len(done) >= len(pane_indices):
                            log(f"AUTO_TRUST all {len(done)} panes done")
                            break
                        for idx in pane_indices:
                            if idx in done:
                                continue
                            target = f"{SESSION_NAME}:.{idx}"
                            r = run_tmux("capture-pane", "-t", target, "-p")
                            if not r["ok"]:
                                continue
                            out = r["stdout"].lower()
                            # 이미 claude에 진입 완료?
                            if any(kw in out for kw in done_keywords):
                                done.add(idx)
                                log(f"AUTO_TRUST pane={idx} already in claude (sec={sec})")
                                continue
                            # trust 프롬프트 감지 → Enter 전송
                            if any(kw in out for kw in trust_keywords):
                                run_tmux("send-keys", "-t", target, "Enter")
                                log(f"AUTO_TRUST pane={idx} Enter sent (sec={sec})")
                                time.sleep(0.5)
                threading.Thread(target=auto_trust, daemon=True).start()
                log(f"SETUP_SESSION launched claude in {len(pane_indices)} panes + auto-trust")
                return {"ok": True, "panes": pane_indices, "launched": len(pane_indices)}

        return {"ok": True}

    def _reset_session(self, body):
        log("RESET SESSION")
        run_tmux("kill-session", "-t", SESSION_NAME)
        time.sleep(1)
        return run_tmux("new-session", "-d", "-s", SESSION_NAME, "-c", os.path.expanduser("~"))

    def _pane_content(self, params):
        pane = params.get("pane", ["1"])[0]
        r = run_tmux("capture-pane", "-t", f"{SESSION_NAME}:.{pane}", "-p", "-S", "-")
        if not r["ok"]: return r
        lines = r["stdout"].split("\n") if r["stdout"] else []
        return {"ok": True, "content": r["stdout"], "lines": lines, "lineCount": len(lines)}

    def _pane_prompt(self, params):
        pane = params.get("pane", ["1"])[0]
        r = run_tmux("capture-pane", "-t", f"{SESSION_NAME}:.{pane}", "-p")
        if not r["ok"]: return r
        content_lines = [l for l in r["stdout"].split("\n") if l.strip()
                         and 'bypass permissions' not in l and 'shift+tab' not in l and '⏵⏵' not in l
                         and all(c not in '─━═' for c in l.strip()[:3] if l.strip())]
        # 구분선 필터 보정
        content_lines = [l for l in content_lines if not (l.strip() and all(c in '─━═' for c in l.strip()))]
        idle = False
        for cl in content_lines[-3:]:
            if re.search(r'❯|~\$\s*$|\$\s*$', cl): idle = True; break
        return {"ok": True, "idle": idle, "lastLines": content_lines[-3:] if content_lines else []}

    # ── Claude Accounts (multi-account manager) ──

    def _claude_accounts_list(self):
        try:
            rows = db_exec(
                "SELECT id, name, active, priority, created, rate_limited_until, rate_limit_reason FROM claude_accounts ORDER BY priority ASC, created DESC",
                fetch=True
            ) or []
            # 만료된 cooldown 자동 정리
            now = datetime.now()
            for r in rows:
                until = r.get("rate_limited_until")
                if until:
                    try:
                        u = datetime.fromisoformat(until)
                        if now >= u:
                            db_exec("UPDATE claude_accounts SET rate_limited_until=NULL, rate_limit_reason=NULL WHERE id=?", (r["id"],))
                            r["rate_limited_until"] = None
                            r["rate_limit_reason"] = None
                    except: pass
            return {"ok": True, "accounts": rows}
        except Exception as e:
            return {"ok": False, "error": str(e), "accounts": []}

    def _claude_account_unlock(self, body):
        """수동으로 rate-limit cooldown 해제"""
        aid = body.get("id")
        if not aid: return {"ok": False, "error": "id required"}
        db_exec("UPDATE claude_accounts SET rate_limited_until=NULL, rate_limit_reason=NULL WHERE id=?", (int(aid),))
        return {"ok": True}

    def _claude_account_save(self, body):
        aid = body.get("id")
        name = (body.get("name", "") or "").strip()
        credentials = (body.get("credentials", "") or "").strip()
        if not name:
            return {"ok": False, "error": "이름이 필요합니다"}
        if not credentials:
            return {"ok": False, "error": "credentials가 비어있습니다"}
        try:
            json.loads(credentials)
        except Exception as e:
            return {"ok": False, "error": f"유효한 JSON이 아닙니다: {e}"}
        now = datetime.now().isoformat()
        try:
            if aid:
                db_exec(
                    "UPDATE claude_accounts SET name=?, credentials=? WHERE id=?",
                    (name, credentials, aid)
                )
                try: _sync_account_to_dir(aid, credentials)
                except Exception as e: log(f"sync acct dir failed: {e}")
                return {"ok": True, "id": aid}
            else:
                new_id = db_exec(
                    "INSERT INTO claude_accounts (name, credentials, active, created) VALUES (?,?,?,?)",
                    (name, credentials, 0, now)
                )
                try: _sync_account_to_dir(new_id, credentials)
                except Exception as e: log(f"sync acct dir failed: {e}")
                return {"ok": True, "id": new_id}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _claude_account_delete(self, body):
        aid = body.get("id")
        if not aid:
            return {"ok": False, "error": "id required"}
        try:
            db_exec("DELETE FROM claude_accounts WHERE id=?", (aid,))
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _claude_account_activate(self, body):
        aid = body.get("id")
        if not aid:
            return {"ok": False, "error": "id required"}
        try:
            row = db_exec("SELECT * FROM claude_accounts WHERE id=?", (aid,), fetchone=True)
            if not row:
                return {"ok": False, "error": "not found"}
            db_exec("UPDATE claude_accounts SET active=0")
            db_exec("UPDATE claude_accounts SET active=1 WHERE id=?", (aid,))

            # setup-token 여부 체크 (refreshToken 빈 값)
            is_setup_token = False
            try:
                parsed = json.loads(row["credentials"])
                oauth = parsed.get("claudeAiOauth", {})
                if oauth.get("accessToken") and not oauth.get("refreshToken"):
                    is_setup_token = True
            except Exception:
                pass

            claude_dir = os.path.expanduser("~/.claude")
            os.makedirs(claude_dir, exist_ok=True)
            creds_path = os.path.join(claude_dir, ".credentials.json")

            if is_setup_token:
                # setup-token 계정은 global credentials.json 안 만듦 (CLI가 혼동)
                # 오히려 기존 전체 OAuth 파일이 있다면 보존
                log(f"Claude account activated (setup-token, env var only): {row['name']}")
            else:
                # 전체 OAuth 계정이면 global에 저장
                with open(creds_path, "w", encoding="utf-8") as f:
                    f.write(row["credentials"])
            if not is_setup_token:
                try:
                    os.chmod(creds_path, 0o600)
                except Exception:
                    pass
            try: _sync_account_to_dir(row["id"], row["credentials"])
            except Exception as e: log(f"sync acct dir failed: {e}")
            if not is_setup_token:
                log(f"Claude account activated: {row['name']}")
            return {"ok": True, "active": row["name"], "setup_token": is_setup_token}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _claude_account_test(self, body):
        """계정으로 Claude 호출 테스트. 성공/실패 즉시 반환."""
        aid = body.get("id")
        if not aid: return {"ok": False, "error": "id required"}
        try:
            env = get_claude_env_for_account(int(aid))
            # 짧은 프롬프트로 빠르게 테스트
            result = subprocess.run(
                ["claude", "-p", "--dangerously-skip-permissions"],
                input="Reply only with: OK",
                capture_output=True, text=True, timeout=30, env=env
            )
            output = (result.stdout or "").strip()
            stderr = (result.stderr or "").strip()
            # 성공 판단: returncode 0 + 응답 있음 + "Not logged in" 아님
            if result.returncode == 0 and output and "Not logged in" not in output and "Please run" not in output:
                return {
                    "ok": True,
                    "authenticated": True,
                    "response": output[:100],
                    "msg": f"✅ 인증 성공! 응답: \"{output[:50]}\""
                }
            else:
                err = output or stderr or "(응답 없음)"
                return {
                    "ok": True,
                    "authenticated": False,
                    "error": err[:300],
                    "msg": f"❌ 인증 실패\n{err[:300]}"
                }
        except subprocess.TimeoutExpired:
            return {"ok": True, "authenticated": False, "msg": "❌ 타임아웃 (30초) — 인증 문제 가능성"}
        except Exception as e:
            return {"ok": True, "authenticated": False, "msg": f"❌ 오류: {e}"}

    # ═══════════════════════════════════════
    # Gemini accounts (API key / OAuth JSON)
    # ═══════════════════════════════════════
    def _gemini_accounts_list(self):
        try:
            rows = db_exec(
                "SELECT id, name, auth_type, active, priority, created FROM gemini_accounts ORDER BY priority ASC, created DESC",
                fetch=True,
            ) or []
            return {"ok": True, "accounts": [dict(r) for r in rows]}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _gemini_account_save(self, body):
        """name + auth_type('apikey'|'oauth') + credentials(문자열). id 있으면 업데이트."""
        name = (body.get("name") or "").strip()
        auth_type = (body.get("auth_type") or "apikey").strip()
        credentials = (body.get("credentials") or "").strip()
        aid = body.get("id")
        if auth_type not in ("apikey", "oauth"):
            return {"ok": False, "error": "auth_type은 apikey 또는 oauth"}
        if not name or not credentials:
            return {"ok": False, "error": "name/credentials 필수"}
        # 간단 검증
        if auth_type == "oauth":
            try:
                json.loads(credentials)
            except Exception:
                return {"ok": False, "error": "oauth_creds.json 형식이 올바르지 않음"}
        try:
            if aid:
                db_exec(
                    "UPDATE gemini_accounts SET name=?, auth_type=?, credentials=? WHERE id=?",
                    (name, auth_type, credentials, int(aid)),
                )
                new_id = int(aid)
            else:
                new_id = db_exec(
                    "INSERT INTO gemini_accounts (name, auth_type, credentials, active, created) VALUES (?,?,?,?,?)",
                    (name, auth_type, credentials, 0, datetime.now().isoformat()),
                )
            _sync_gemini_account_to_dir(new_id, auth_type, credentials)
            # active 계정이면 전역 ~/.gemini도 갱신
            active = db_exec("SELECT id FROM gemini_accounts WHERE active=1 LIMIT 1", fetchone=True)
            if active and int(active["id"]) == int(new_id):
                _sync_gemini_credentials()
            return {"ok": True, "id": new_id}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _gemini_account_delete(self, body):
        aid = body.get("id")
        if not aid: return {"ok": False, "error": "id required"}
        try:
            db_exec("DELETE FROM gemini_accounts WHERE id=?", (int(aid),))
            # 디렉토리 정리 (best-effort)
            try:
                import shutil
                shutil.rmtree(_get_gemini_account_dir(int(aid)), ignore_errors=True)
            except Exception:
                pass
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _gemini_account_activate(self, body):
        aid = body.get("id")
        if not aid: return {"ok": False, "error": "id required"}
        try:
            db_exec("UPDATE gemini_accounts SET active=0")
            db_exec("UPDATE gemini_accounts SET active=1 WHERE id=?", (int(aid),))
            _sync_gemini_credentials()
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _gemini_account_test(self, body):
        """계정으로 Gemini CLI 간단 호출 테스트."""
        aid = body.get("id")
        if not aid: return {"ok": False, "error": "id required"}
        try:
            env = get_gemini_env_for_account(int(aid))
            result = subprocess.run(
                ["gemini", "-m", "gemini-2.5-flash", "-p", "Reply only with: OK"],
                capture_output=True, text=True, timeout=30, env=env,
                encoding="utf-8", errors="replace",
            )
            output = (result.stdout or "").strip()
            stderr = (result.stderr or "").strip()
            if result.returncode == 0 and output:
                return {
                    "ok": True, "authenticated": True,
                    "response": output[:200],
                    "msg": f"✅ 인증 성공! 응답: \"{output[:80]}\"",
                }
            err = stderr or output or "(응답 없음)"
            return {
                "ok": True, "authenticated": False,
                "error": err[:300],
                "msg": f"❌ 인증 실패\n{err[:300]}",
            }
        except subprocess.TimeoutExpired:
            return {"ok": True, "authenticated": False, "msg": "❌ 타임아웃 (30초)"}
        except FileNotFoundError:
            return {"ok": True, "authenticated": False, "msg": "❌ gemini CLI 미설치"}
        except Exception as e:
            return {"ok": True, "authenticated": False, "msg": f"❌ 오류: {e}"}

    # ═══════════════════════════════════════
    # Auth health check — 로그인 후 자동 점검 + 상단 배지용
    # ═══════════════════════════════════════
    def _auth_health(self, params=None):
        """등록된 모든 Claude/Gemini 계정에 대해 짧은 인증 테스트를 병렬 실행.
        파라미터: ?quick=1 → credentials 파싱만 (CLI 호출 X, 매우 빠름).
        기본: CLI 호출 (30초 timeout, 계정별 병렬)."""
        quick = bool((params or {}).get("quick", ["0"])[0] == "1") if params else False

        def _check_claude(acc):
            aid = acc.get("id")
            name = acc.get("name") or f"claude#{aid}"
            # 1) credentials JSON 파싱 가능?
            try:
                row = db_exec("SELECT credentials FROM claude_accounts WHERE id=?", (int(aid),), fetchone=True)
                if not row or not row.get("credentials"):
                    return {"provider": "claude", "id": aid, "name": name, "ok": False,
                            "reason": "credentials 없음"}
                try:
                    data = json.loads(row["credentials"])
                except Exception:
                    return {"provider": "claude", "id": aid, "name": name, "ok": False,
                            "reason": "credentials JSON 파싱 실패"}
                oauth = data.get("claudeAiOauth") or {}
                expires_at = oauth.get("expiresAt") or 0
                has_refresh = bool(oauth.get("refreshToken"))
                has_access = bool(oauth.get("accessToken"))
                if not has_access:
                    return {"provider": "claude", "id": aid, "name": name, "ok": False,
                            "reason": "accessToken 없음"}
                expired = False
                if expires_at:
                    try:
                        expired = (expires_at/1000 if expires_at > 1e12 else expires_at) < time.time()
                    except Exception:
                        pass
                # rate-limited 체크
                rl = acc.get("rate_limited_until")
                rate_limited = False
                if rl:
                    try:
                        rate_limited = datetime.fromisoformat(rl) > datetime.now()
                    except Exception: pass
            except Exception as e:
                return {"provider": "claude", "id": aid, "name": name, "ok": False,
                        "reason": f"체크 오류: {e}"}
            # quick 모드는 여기까지
            if quick:
                if rate_limited:
                    return {"provider": "claude", "id": aid, "name": name, "ok": False,
                            "reason": "rate-limited 쿨다운 중", "rateLimited": True}
                if expired and not has_refresh:
                    return {"provider": "claude", "id": aid, "name": name, "ok": False,
                            "reason": "토큰 만료 (refresh 없음)"}
                return {"provider": "claude", "id": aid, "name": name, "ok": True,
                        "reason": "credentials OK (CLI 호출 생략)", "rateLimited": rate_limited}
            # 실제 CLI 호출 (짧게)
            try:
                env = get_claude_env_for_account(int(aid))
                result = subprocess.run(
                    ["claude", "-p", "--dangerously-skip-permissions"],
                    input="OK", capture_output=True, text=True, timeout=25, env=env,
                )
                out = (result.stdout or "").strip()
                err = (result.stderr or "").strip()
                if result.returncode == 0 and out and "Not logged in" not in out and "Please run" not in out:
                    return {"provider": "claude", "id": aid, "name": name, "ok": True,
                            "reason": "OK", "rateLimited": rate_limited}
                return {"provider": "claude", "id": aid, "name": name, "ok": False,
                        "reason": (out or err or "응답 없음")[:160]}
            except subprocess.TimeoutExpired:
                return {"provider": "claude", "id": aid, "name": name, "ok": False,
                        "reason": "타임아웃 (25초)"}
            except Exception as e:
                return {"provider": "claude", "id": aid, "name": name, "ok": False,
                        "reason": f"{e}"}

        def _check_gemini(acc):
            aid = acc.get("id")
            name = acc.get("name") or f"gemini#{aid}"
            try:
                row = db_exec("SELECT auth_type, credentials FROM gemini_accounts WHERE id=?",
                              (int(aid),), fetchone=True)
                if not row or not row.get("credentials"):
                    return {"provider": "gemini", "id": aid, "name": name, "ok": False,
                            "reason": "credentials 없음"}
                if row.get("auth_type") == "apikey":
                    key_ok = bool((row.get("credentials") or "").strip())
                else:
                    try:
                        json.loads(row["credentials"]); key_ok = True
                    except Exception:
                        return {"provider": "gemini", "id": aid, "name": name, "ok": False,
                                "reason": "OAuth credentials JSON 파싱 실패"}
            except Exception as e:
                return {"provider": "gemini", "id": aid, "name": name, "ok": False,
                        "reason": f"체크 오류: {e}"}
            if quick:
                return {"provider": "gemini", "id": aid, "name": name, "ok": bool(key_ok),
                        "reason": "credentials OK (CLI 호출 생략)" if key_ok else "credentials 비어있음"}
            try:
                env = get_gemini_env_for_account(int(aid))
                result = subprocess.run(
                    ["gemini", "-m", "gemini-2.5-flash", "-p", "OK"],
                    capture_output=True, text=True, timeout=25, env=env,
                    encoding="utf-8", errors="replace",
                )
                out = (result.stdout or "").strip()
                err = (result.stderr or "").strip()
                if result.returncode == 0 and out:
                    return {"provider": "gemini", "id": aid, "name": name, "ok": True, "reason": "OK"}
                return {"provider": "gemini", "id": aid, "name": name, "ok": False,
                        "reason": (err or out or "응답 없음")[:160]}
            except subprocess.TimeoutExpired:
                return {"provider": "gemini", "id": aid, "name": name, "ok": False,
                        "reason": "타임아웃 (25초)"}
            except FileNotFoundError:
                return {"provider": "gemini", "id": aid, "name": name, "ok": False,
                        "reason": "gemini CLI 미설치"}
            except Exception as e:
                return {"provider": "gemini", "id": aid, "name": name, "ok": False,
                        "reason": f"{e}"}

        try:
            claude_rows = db_exec(
                "SELECT id, name, rate_limited_until FROM claude_accounts ORDER BY priority ASC, id ASC",
                fetch=True) or []
        except Exception: claude_rows = []
        try:
            gemini_rows = db_exec(
                "SELECT id, name FROM gemini_accounts ORDER BY priority ASC, id ASC",
                fetch=True) or []
        except Exception: gemini_rows = []

        # 병렬 실행 (계정마다 thread)
        from concurrent.futures import ThreadPoolExecutor
        results = []
        pool_size = max(1, len(claude_rows) + len(gemini_rows))
        with ThreadPoolExecutor(max_workers=min(pool_size, 8)) as ex:
            futs = [ex.submit(_check_claude, dict(a)) for a in claude_rows]
            futs += [ex.submit(_check_gemini, dict(a)) for a in gemini_rows]
            for f in futs:
                try:
                    results.append(f.result(timeout=35))
                except Exception as e:
                    results.append({"provider": "?", "ok": False, "reason": f"{e}"})

        claude_results = [r for r in results if r.get("provider") == "claude"]
        gemini_results = [r for r in results if r.get("provider") == "gemini"]
        # 전체 요약 — 각 provider 최소 1개 계정이 OK 면 provider OK
        claude_ok = any(r.get("ok") for r in claude_results)
        gemini_ok = any(r.get("ok") for r in gemini_results)
        return {
            "ok": True,
            "checkedAt": datetime.now().isoformat(),
            "quick": quick,
            "claude": {"ok": claude_ok, "count": len(claude_results),
                       "okCount": sum(1 for r in claude_results if r.get("ok")),
                       "accounts": claude_results},
            "gemini": {"ok": gemini_ok, "count": len(gemini_results),
                       "okCount": sum(1 for r in gemini_results if r.get("ok")),
                       "accounts": gemini_results},
        }

    def _claude_next_account(self, body):
        """회전 전략에 따라 다음 account_id 반환."""
        try:
            mode_row = db_exec("SELECT value FROM system_settings WHERE key='claude_rotation_mode'", fetchone=True)
            mode = mode_row["value"] if mode_row and mode_row.get("value") else "round-robin"

            accounts = db_exec("SELECT id, name, priority FROM claude_accounts ORDER BY priority ASC, id ASC", fetch=True) or []
            if not accounts:
                return {"ok": False, "error": "계정이 없습니다"}

            if mode == "manual":
                active = db_exec("SELECT id FROM claude_accounts WHERE active=1 LIMIT 1", fetchone=True)
                return {"ok": True, "accountId": active["id"] if active else accounts[0]["id"], "mode": mode}

            elif mode == "sequential":
                return {"ok": True, "accountId": accounts[0]["id"], "mode": mode}

            else:  # round-robin — rate-limited 계정 자동 스킵
                last_row = db_exec("SELECT value FROM system_settings WHERE key='claude_last_used_id'", fetchone=True)
                last_id = 0
                try:
                    if last_row and last_row.get("value"):
                        last_id = int(last_row["value"])
                except Exception:
                    last_id = 0
                ids = [a["id"] for a in accounts]
                try:
                    idx = ids.index(last_id)
                    next_idx = (idx + 1) % len(ids)
                except ValueError:
                    next_idx = 0
                # 한 바퀴 돌면서 사용 가능한 첫 계정 찾기
                next_id = None
                for _ in range(len(ids)):
                    cand = ids[next_idx]
                    if is_account_available(cand):
                        next_id = cand
                        break
                    next_idx = (next_idx + 1) % len(ids)
                if next_id is None:
                    return {"ok": False, "error": "모든 계정이 한도 초과 상태"}
                now = datetime.now().isoformat()
                existing = db_exec("SELECT key FROM system_settings WHERE key='claude_last_used_id'", fetchone=True)
                if existing:
                    db_exec("UPDATE system_settings SET value=?, updated=? WHERE key='claude_last_used_id'", (str(next_id), now))
                else:
                    db_exec("INSERT INTO system_settings (key, value, updated) VALUES (?,?,?)", ("claude_last_used_id", str(next_id), now))
                return {"ok": True, "accountId": next_id, "mode": mode}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _claude_next_preview(self, body):
        """다음에 사용될 계정만 반환 (counter 증가 안 함, UI 표시용 — rate-limited 스킵)"""
        try:
            nid = pick_available_account()
            return {"ok": True, "accountId": nid} if nid else {"ok": False, "error": "사용 가능한 계정 없음"}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _claude_accounts_reorder(self, body):
        ids = body.get("ids", [])
        try:
            for priority, aid in enumerate(ids):
                db_exec("UPDATE claude_accounts SET priority=? WHERE id=?", (priority, aid))
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    # ── Claude Web-based Login (OAuth via subprocess) ──

    def _claude_login_start(self, body):
        global _claude_login_proc, _claude_login_url, _claude_login_output, _claude_login_master_fd
        with _claude_login_lock:
            # Kill existing process if any
            try:
                if _claude_login_proc:
                    try:
                        os.kill(_claude_login_proc, 9)
                    except Exception:
                        pass
            except Exception:
                pass
            _claude_login_url = None
            _claude_login_output = []
            _claude_login_master_fd = None

            # PTY로 실행 - claude CLI가 터미널 크기를 자체 감지하므로
            # COLUMNS 환경변수로는 안 되고 PTY의 winsize를 크게 설정해야 함
            import pty, fcntl, termios, struct, tty
            try:
                master, slave = pty.openpty()
                # 터미널 크기 9999칸 (URL 줄바꿈 방지)
                ws = struct.pack('HHHH', 50, 9999, 0, 0)
                fcntl.ioctl(slave, termios.TIOCSWINSZ, ws)
                # ECHO만 끄기 (ICANON 유지 — CLI가 line 모드 기대)
                attrs = termios.tcgetattr(slave)
                attrs[3] = attrs[3] & ~termios.ECHO  # lflags: ECHO off
                termios.tcsetattr(slave, termios.TCSANOW, attrs)
            except Exception as e:
                return {"ok": False, "error": f"PTY 생성 실패: {e}"}

            env = get_claude_env()
            pid = os.fork()
            if pid == 0:
                # child process
                os.setsid()
                try:
                    fcntl.ioctl(slave, termios.TIOCSCTTY, 0)
                except Exception:
                    pass
                os.close(master)
                os.dup2(slave, 0)
                os.dup2(slave, 1)
                os.dup2(slave, 2)
                os.close(slave)
                try:
                    # auth login 사용 (전체 권한: profile, inference, sessions:claude_code, mcp_servers, api_key, file_upload)
                    os.execvpe("claude", ["claude", "auth", "login"], env)
                except Exception:
                    os._exit(1)
            else:
                # parent
                os.close(slave)
                _claude_login_proc = pid
                _claude_login_master_fd = master

            def reader():
                global _claude_login_url, _claude_login_output
                import select as sel
                ansi_re = re.compile(rb'\x1b\[[0-9;?]*[a-zA-Z]')
                buf = b""
                last_url_update = 0  # URL 발견 후 경과 시간
                tentative_url = None
                try:
                    while True:
                        r, _, _ = sel.select([master], [], [], 0.3)
                        if r:
                            try:
                                chunk = os.read(master, 4096)
                            except OSError:
                                break
                            if not chunk:
                                break
                            buf += chunk
                            # 디버그: 원본 바이트를 파일에 덤프
                            try:
                                with open('/tmp/claude_login_raw.log', 'ab') as df:
                                    df.write(chunk)
                            except: pass
                            clean = ansi_re.sub(b'', buf).decode('utf-8', errors='replace')
                            _claude_login_output = clean.split('\n')

                            if _claude_login_url:
                                continue

                            # URL 추출: 줄바꿈+공백 제거 후 매칭
                            joined = re.sub(r'\n+\s*', '', clean)
                            best = None
                            for m in re.finditer(r'https?://[^\s]+', joined):
                                c = m.group(0).rstrip(').,"\'')
                                if 'oauth' in c.lower() or 'authorize' in c.lower():
                                    if not best or len(c) > len(best):
                                        best = c
                            if best:
                                # 조건: 새 URL이 이전과 같고 (더 이상 길어지지 않음)
                                #      그리고 최소 길이 200자 이상 (완전한 OAuth URL)
                                #      그리고 state= 포함 (URL 끝부분 파라미터)
                                if tentative_url == best and len(best) > 200 and 'state=' in best:
                                    _claude_login_url = best
                                else:
                                    tentative_url = best
                        else:
                            # 입력 없음: 잠깐 기다린 후 현재 tentative URL이 충분히 완성되었으면 commit
                            if tentative_url and not _claude_login_url:
                                if len(tentative_url) > 200 and 'state=' in tentative_url and 'redirect_uri' in tentative_url:
                                    _claude_login_url = tentative_url
                            # 프로세스 체크
                            try:
                                wpid, _ = os.waitpid(_claude_login_proc, os.WNOHANG)
                                if wpid != 0:
                                    break
                            except Exception:
                                break
                except Exception as e:
                    log(f"claude login PTY reader error: {e}")
            threading.Thread(target=reader, daemon=True).start()

            # URL 대기 (최대 25초) — 완전한 URL 확보 후 반환
            for _ in range(125):
                if _claude_login_url and 'redirect_uri' in _claude_login_url and 'state=' in _claude_login_url:
                    break
                time.sleep(0.2)

            if _claude_login_url:
                return {"ok": True, "url": _claude_login_url}
            # 프로세스 상태 확인
            exited = not _is_claude_proc_alive()
            return {
                "ok": False,
                "error": "URL 추출 실패 — 파일 업로드/붙여넣기를 사용하세요.",
                "exited": exited,
                "output": "\n".join(_claude_login_output[-40:])
            }

    def _claude_login_submit(self, body):
        global _claude_login_proc, _claude_login_output, _claude_login_master_fd
        code = (body.get("code", "") or "").strip()
        if not code:
            return {"ok": False, "error": "code required"}
        if not _claude_login_proc or not _is_claude_proc_alive():
            return {"ok": False, "error": "로그인 세션이 없습니다 (프로세스 종료됨)"}

        # claude auth login은 .credentials.json 파일에 저장
        creds_path = os.path.expanduser("~/.claude/.credentials.json")
        old_mtime = os.path.getmtime(creds_path) if os.path.exists(creds_path) else 0

        try:
            os.write(_claude_login_master_fd, (code + "\r").encode())
            log(f"CLAUDE_LOGIN submit code len={len(code)} has_hash={'#' in code}")
        except Exception as e:
            return {"ok": False, "error": f"PTY write 실패: {e}"}

        # .credentials.json 파일 갱신 대기 (최대 60초)
        cli_error = None
        creds = None
        for _iter in range(300):
            if os.path.exists(creds_path):
                new_mtime = os.path.getmtime(creds_path)
                if new_mtime > old_mtime:
                    time.sleep(0.5)  # 파일 완전 쓰기 대기
                    try:
                        with open(creds_path, "r", encoding="utf-8") as f:
                            creds = f.read()
                        json.loads(creds)  # 검증
                        log(f"CLAUDE_LOGIN creds updated size={len(creds)}")
                        break
                    except Exception as e:
                        log(f"creds read error: {e}")
            recent = "\n".join(_claude_login_output[-10:])
            if "Invalid code" in recent or "OAuth error" in recent or "authentication failed" in recent.lower():
                cli_error = recent
                log(f"CLAUDE_LOGIN CLI error: {recent[-300:]}")
                break
            if not _is_claude_proc_alive():
                log(f"CLAUDE_LOGIN process exited")
                # 마지막 확인
                if os.path.exists(creds_path) and os.path.getmtime(creds_path) > old_mtime:
                    try:
                        with open(creds_path, "r", encoding="utf-8") as f:
                            creds = f.read()
                        json.loads(creds)
                    except: pass
                break
            time.sleep(0.2)

        # 프로세스 정리
        if _is_claude_proc_alive():
            try: os.kill(_claude_login_proc, 9)
            except Exception: pass

        if creds:
            now = datetime.now().isoformat()
            name = (body.get("name") or "").strip() or f"Account {datetime.now().strftime('%m-%d %H:%M')}"
            try:
                aid = db_exec(
                    "INSERT INTO claude_accounts (name, credentials, active, created) VALUES (?,?,?,?)",
                    (name, creds, 0, now)
                )
                try: _sync_account_to_dir(aid, creds)
                except Exception as e: log(f"sync acct dir failed: {e}")
                return {"ok": True, "accountId": aid, "name": name, "scope": "full"}
            except Exception as e:
                return {"ok": False, "error": f"DB 저장 실패: {e}"}
        else:
            err_msg = "로그인 실패 — "
            if cli_error:
                if "Invalid code" in cli_error:
                    err_msg += "❌ 유효하지 않은 코드입니다.\n\n1. 취소 후 새 URL로 다시 시작\n2. Anthropic 승인 후 나오는 코드 복사 (# 포함 전체)\n3. 붙여넣고 완료"
                else:
                    err_msg += cli_error[-200:]
            else:
                err_msg += "credentials.json이 갱신되지 않았습니다. 코드가 만료되었을 수 있습니다."
            return {
                "ok": False,
                "error": err_msg,
                "output": "\n".join(_claude_login_output[-60:])
            }

    def _claude_login_status(self):
        if not _claude_login_proc:
            return {"ok": True, "running": False}
        running = _is_claude_proc_alive()
        return {
            "ok": True,
            "running": running,
            "url": _claude_login_url,
            "output": "\n".join(_claude_login_output[-10:])
        }

    def _claude_login_cancel(self, body):
        global _claude_login_proc
        if _claude_login_proc and _is_claude_proc_alive():
            try:
                os.kill(_claude_login_proc, 9)
            except Exception:
                pass
        return {"ok": True}

    # ── System Settings ──

    def _settings_get(self):
        rows = db_exec("SELECT key, value FROM system_settings", fetch=True)
        settings = {r["key"]: r["value"] for r in rows}
        # YouTube API key: DB에 없으면 env 를 대체 노출 (읽기 전용 안내용)
        if not (settings.get("youtube_api_key") or "").strip():
            env_key = os.environ.get("YOUTUBE_API_KEY", "").strip()
            if env_key:
                settings["youtube_api_key"] = env_key
                settings["youtube_api_key_source"] = "env"
            else:
                settings["youtube_api_key_source"] = "none"
        else:
            settings["youtube_api_key_source"] = "db"
        return {"ok": True, "settings": settings}

    def _settings_set(self, body):
        key = body.get("key", "")
        value = body.get("value", "")
        if not key:
            return {"ok": False, "error": "key required"}
        now = datetime.now().isoformat()
        existing = db_exec("SELECT key FROM system_settings WHERE key=?", (key,), fetchone=True)
        if existing:
            db_exec("UPDATE system_settings SET value=?, updated=? WHERE key=?", (value, now, key))
        else:
            db_exec("INSERT INTO system_settings (key, value, updated) VALUES (?,?,?)", (key, value, now))

        # Special handling: claude_credentials → write to ~/.claude/.credentials.json
        if key == "claude_credentials" and value:
            try:
                json.loads(value)  # Validate JSON
                claude_dir = os.path.expanduser("~/.claude")
                os.makedirs(claude_dir, exist_ok=True)
                creds_path = os.path.join(claude_dir, ".credentials.json")
                with open(creds_path, "w", encoding="utf-8") as f:
                    f.write(value)
                try:
                    os.chmod(creds_path, 0o600)
                except Exception:
                    pass
                log(f"Claude credentials saved to {creds_path}")
            except json.JSONDecodeError as e:
                log(f"Failed to save Claude credentials: invalid JSON")
                return {"ok": False, "error": f"Invalid JSON: {e}"}
            except Exception as e:
                log(f"Failed to save Claude credentials: {e}")
                return {"ok": False, "error": str(e)}

        return {"ok": True}

    def _settings_delete(self, body):
        key = body.get("key", "")
        if key:
            db_exec("DELETE FROM system_settings WHERE key=?", (key,))
        return {"ok": True}

    # ── File Browser ──

    def _browse_path(self, params):
        """Browse filesystem from workspace root."""
        row = db_exec("SELECT value FROM system_settings WHERE key='workspace_root'", fetchone=True)
        root = row["value"] if row and row.get("value") else os.path.expanduser("~")
        root = os.path.normpath(root)

        rel_path = params.get("path", [""])[0]
        # Security: only allow paths within root
        target = os.path.normpath(os.path.join(root, rel_path.lstrip("/").lstrip("\\")))
        if not target.startswith(root):
            return {"ok": False, "error": "접근 거부 (root 밖)"}

        if not os.path.isdir(target):
            return {"ok": False, "error": "디렉토리 아님"}

        items = []
        try:
            for name in sorted(os.listdir(target)):
                if name.startswith("."):
                    continue
                # .trash 폴더는 일반 브라우징에서 숨김 (루트에서만 발생 가능)
                if name == ".trash":
                    continue
                full = os.path.join(target, name)
                try:
                    is_dir = os.path.isdir(full)
                    items.append({
                        "name": name,
                        "path": os.path.relpath(full, root).replace("\\", "/"),
                        "isDir": is_dir,
                        "size": os.path.getsize(full) if not is_dir else None
                    })
                except Exception:
                    continue
        except Exception as e:
            return {"ok": False, "error": str(e)}

        return {"ok": True, "root": root.replace("\\", "/"), "path": rel_path, "items": items}

    def _fs_mkdir(self, body):
        """workspace root 내에 폴더 생성"""
        row = db_exec("SELECT value FROM system_settings WHERE key='workspace_root'", fetchone=True)
        root = row["value"] if row and row.get("value") else os.path.expanduser("~")

        rel_path = body.get("path", "").strip()
        name = body.get("name", "").strip()
        if not name:
            return {"ok": False, "error": "폴더 이름이 필요합니다"}
        # Sanitize name
        if any(c in name for c in "/\\:*?\"<>|"):
            return {"ok": False, "error": "폴더 이름에 사용할 수 없는 문자가 있습니다"}

        target = os.path.normpath(os.path.join(root, rel_path.lstrip("/").lstrip("\\"), name))
        # Security: must be within root
        root_abs = os.path.abspath(root)
        target_abs = os.path.abspath(target)
        if not target_abs.startswith(root_abs):
            return {"ok": False, "error": "접근 거부 (root 밖)"}

        try:
            os.makedirs(target_abs, exist_ok=True)
            return {"ok": True, "path": os.path.relpath(target_abs, root_abs).replace("\\", "/")}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _fs_get_root(self):
        row = db_exec("SELECT value FROM system_settings WHERE key='workspace_root'", fetchone=True)
        root = row["value"] if row and row.get("value") else os.path.expanduser("~")
        return os.path.normpath(os.path.abspath(root))

    def _fs_delete(self, body):
        """파일/폴더를 휴지통으로 이동"""
        import shutil
        root = self._fs_get_root()
        rel = body.get("path", "").strip()
        if not rel: return {"ok": False, "error": "path required"}
        target = os.path.normpath(os.path.abspath(os.path.join(root, rel.lstrip("/").lstrip("\\"))))
        if not target.startswith(root):
            return {"ok": False, "error": "접근 거부"}
        if not os.path.exists(target):
            return {"ok": False, "error": "파일/폴더 없음"}
        # .trash 디렉토리 자체는 삭제 못함
        trash_dir = os.path.join(root, ".trash")
        if target == trash_dir or target.startswith(trash_dir + os.sep):
            return {"ok": False, "error": "휴지통 내부 항목은 휴지통 삭제/복원 기능 사용"}

        os.makedirs(trash_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        basename = os.path.basename(target.rstrip(os.sep))
        trash_name = f"{ts}_{basename}"
        trash_path = os.path.join(trash_dir, trash_name)
        # 중복 시 suffix
        i = 1
        while os.path.exists(trash_path):
            trash_path = os.path.join(trash_dir, f"{ts}_{i}_{basename}")
            i += 1
        try:
            shutil.move(target, trash_path)
            # 메타 저장
            meta = {
                "original_path": rel.replace("\\", "/"),
                "original_name": basename,
                "deleted_at": datetime.now().isoformat(),
                "was_dir": os.path.isdir(trash_path)
            }
            with open(trash_path + ".meta.json", "w", encoding="utf-8") as f:
                json.dump(meta, f, ensure_ascii=False)
            return {"ok": True, "trashPath": os.path.basename(trash_path)}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _fs_trash_list(self, body):
        root = self._fs_get_root()
        trash_dir = os.path.join(root, ".trash")
        if not os.path.isdir(trash_dir):
            return {"ok": True, "items": []}
        items = []
        for name in sorted(os.listdir(trash_dir), reverse=True):
            if name.endswith(".meta.json"): continue
            full = os.path.join(trash_dir, name)
            meta_path = full + ".meta.json"
            meta = {}
            if os.path.exists(meta_path):
                try:
                    with open(meta_path, "r", encoding="utf-8") as f:
                        meta = json.load(f)
                except: pass
            try:
                size = os.path.getsize(full) if os.path.isfile(full) else None
            except: size = None
            items.append({
                "trashName": name,
                "originalPath": meta.get("original_path", ""),
                "originalName": meta.get("original_name", name),
                "deletedAt": meta.get("deleted_at", ""),
                "wasDir": os.path.isdir(full),
                "size": size
            })
        return {"ok": True, "items": items}

    def _fs_trash_restore(self, body):
        import shutil
        root = self._fs_get_root()
        trash_dir = os.path.join(root, ".trash")
        trash_name = body.get("trashName", "").strip()
        if not trash_name or "/" in trash_name or "\\" in trash_name:
            return {"ok": False, "error": "잘못된 이름"}
        trash_path = os.path.join(trash_dir, trash_name)
        if not os.path.exists(trash_path):
            return {"ok": False, "error": "파일 없음"}
        # 메타 읽기
        meta_path = trash_path + ".meta.json"
        original_path = ""
        try:
            with open(meta_path, "r", encoding="utf-8") as f:
                meta = json.load(f)
            original_path = meta.get("original_path", "")
        except: pass
        if not original_path:
            # 메타 없으면 루트에 복원
            original_path = body.get("originalName") or trash_name
        target = os.path.normpath(os.path.abspath(os.path.join(root, original_path.lstrip("/"))))
        if not target.startswith(root):
            return {"ok": False, "error": "경로 오류"}
        # 대상 이미 존재 시 suffix
        if os.path.exists(target):
            base, ext = os.path.splitext(target)
            i = 1
            while os.path.exists(f"{base}_restored{i}{ext}"):
                i += 1
            target = f"{base}_restored{i}{ext}"
        try:
            os.makedirs(os.path.dirname(target), exist_ok=True)
            shutil.move(trash_path, target)
            if os.path.exists(meta_path):
                try: os.remove(meta_path)
                except: pass
            return {"ok": True, "restoredTo": os.path.relpath(target, root).replace("\\", "/")}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _fs_trash_delete(self, body):
        import shutil
        root = self._fs_get_root()
        trash_dir = os.path.join(root, ".trash")
        trash_name = body.get("trashName", "").strip()
        if not trash_name or "/" in trash_name or "\\" in trash_name:
            return {"ok": False, "error": "잘못된 이름"}
        trash_path = os.path.join(trash_dir, trash_name)
        try:
            if os.path.isdir(trash_path):
                shutil.rmtree(trash_path)
            elif os.path.isfile(trash_path):
                os.remove(trash_path)
            meta_path = trash_path + ".meta.json"
            if os.path.exists(meta_path):
                os.remove(meta_path)
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _fs_trash_empty(self, body):
        import shutil
        root = self._fs_get_root()
        trash_dir = os.path.join(root, ".trash")
        if not os.path.isdir(trash_dir):
            return {"ok": True}
        try:
            shutil.rmtree(trash_dir)
            os.makedirs(trash_dir, exist_ok=True)
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _fs_download(self, params):
        """파일 다운로드 - workspace root 내만 허용"""
        row = db_exec("SELECT value FROM system_settings WHERE key='workspace_root'", fetchone=True)
        root = row["value"] if row and row.get("value") else os.path.expanduser("~")
        root = os.path.normpath(os.path.abspath(root))
        rel_path = params.get("path", [""])[0]
        target = os.path.normpath(os.path.abspath(os.path.join(root, rel_path.lstrip("/").lstrip("\\"))))
        if not target.startswith(root):
            self.send_error(403, "Access denied")
            return
        if not os.path.isfile(target):
            self.send_error(404, "Not a file")
            return
        try:
            import mimetypes, urllib.parse
            mime = mimetypes.guess_type(target)[0] or 'application/octet-stream'
            fname = os.path.basename(target)
            encoded = urllib.parse.quote(fname)
            self.send_response(200)
            self.send_header("Content-Type", mime)
            self.send_header("Content-Length", os.path.getsize(target))
            # 브라우저 미리보기 vs 다운로드 선택
            as_download = params.get("download", ["0"])[0] == "1"
            if as_download:
                self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{encoded}")
            else:
                self.send_header("Content-Disposition", f"inline; filename*=UTF-8''{encoded}")
            self.end_headers()
            with open(target, "rb") as f:
                while True:
                    chunk = f.read(65536)
                    if not chunk: break
                    self.wfile.write(chunk)
        except Exception as e:
            self.send_error(500, str(e))

    def _fs_browse_system(self, params):
        """시스템 전체 파일 탐색 (workspace_root 선택용). 절대경로 기준.
        권한 에러가 있어도 진입은 허용하고 경고만 표시 (listdir 부분 실패 시 읽을 수 있는 것만 반환)."""
        path = params.get("path", ["/"])[0] or "/"
        try:
            path = os.path.abspath(path)
        except:
            path = "/"
        if not os.path.isdir(path):
            return {"ok": False, "error": "디렉토리가 아닙니다", "path": path}
        items = []
        warn = None
        try:
            names = sorted(os.listdir(path))
        except PermissionError:
            # 권한 에러여도 진입은 성공으로 처리 (프론트에서 경고 배너 표시)
            names = []
            try: import os as _os; uid=_os.geteuid();
            except: uid='?'
            warn = (f"⚠ 폴더 읽기 권한 없음 (listdir 실패) — 컨테이너 uid={uid} 에게 읽기 권한 필요\n"
                    f"→ 시놀로지 SSH에서: chmod -R o+rX {path.replace('/synology','/volume1/00_Gils_Project')}\n"
                    f"또는 DSM 파일스테이션에서 공유폴더 권한에 'everyone read' 추가")
        except Exception as e:
            names = []
            warn = f"⚠ 읽기 실패: {e}"
        for name in names:
            if name.startswith("."): continue
            full = os.path.join(path, name)
            try:
                if os.path.isdir(full):
                    items.append({"name": name, "path": full, "isDir": True})
            except PermissionError:
                continue
            except Exception:
                continue
        # 추천 시작 경로들 (root일 때만)
        suggestions = []
        if path == "/":
            for p in ["/synology", "/workspace", "/app", "/mnt", "/home", "/volume1"]:
                if os.path.isdir(p):
                    suggestions.append(p)
        resp = {"ok": True, "path": path, "items": items, "suggestions": suggestions}
        if warn: resp["warn"] = warn
        return resp

    def _fs_mkdir_system(self, body):
        """절대 경로에 폴더 생성 (workspace_root 설정용)."""
        parent = body.get("path", "").strip()
        name = body.get("name", "").strip()
        if not name:
            return {"ok": False, "error": "이름이 필요합니다"}
        if any(c in name for c in "/\\:*?\"<>|"):
            return {"ok": False, "error": "사용할 수 없는 문자가 있습니다"}
        if not parent or not os.path.isdir(parent):
            return {"ok": False, "error": "부모 경로가 올바르지 않습니다"}
        target = os.path.join(parent, name)
        try:
            os.makedirs(target, exist_ok=True)
            return {"ok": True, "path": target}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    # ── Auth ──

    def _check_auth(self):
        cookie_header = self.headers.get("Cookie", "")
        token = None
        for part in cookie_header.split(";"):
            part = part.strip()
            if part.startswith("session="):
                token = part[8:]
                break
        if not token:
            return False
        session = db_exec("SELECT * FROM sessions WHERE token=?", (token,), fetchone=True)
        if not session:
            return False
        if datetime.fromisoformat(session["expires"]) < datetime.now():
            db_exec("DELETE FROM sessions WHERE token=?", (token,))
            return False
        return True

    def _auth_check(self):
        if self._check_auth():
            return {"ok": True, "authenticated": True}
        return {"ok": True, "authenticated": False}

    def _handle_auth_login(self, body):
        username = body.get("username", "")
        password = body.get("password", "")
        if not username or not password:
            self._json({"ok": False, "error": "아이디와 비밀번호를 입력하세요"})
            return

        user = db_exec("SELECT id, username, password_hash, salt FROM users WHERE username=?", (username,), fetchone=True)
        if not user:
            time.sleep(1)
            self._json({"ok": False, "error": "아이디 또는 비밀번호가 잘못되었습니다"})
            return

        pw_hash, _ = hash_password(password, user["salt"])
        if pw_hash != user["password_hash"]:
            time.sleep(1)
            self._json({"ok": False, "error": "아이디 또는 비밀번호가 잘못되었습니다"})
            return

        token = secrets.token_hex(32)
        expires = (datetime.now() + timedelta(days=7)).isoformat()
        db_exec("INSERT INTO sessions (token, user_id, username, created, expires) VALUES (?,?,?,?,?)",
                (token, user["id"], username, datetime.now().isoformat(), expires))

        result = {"ok": True, "token": token, "username": username}
        self.send_response(200)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        origin = self._get_cors_origin()
        self.send_header("Access-Control-Allow-Origin", origin)
        if origin != "*":
            self.send_header("Access-Control-Allow-Credentials", "true")
        self.send_header("Set-Cookie", f"session={token}; Path=/; HttpOnly; Max-Age=604800")
        self.end_headers()
        self.wfile.write(json.dumps(result, ensure_ascii=False).encode("utf-8"))
        log(f"AUTH login: {username}")

    def _handle_auth_logout(self, body):
        cookie_header = self.headers.get("Cookie", "")
        token = None
        for part in cookie_header.split(";"):
            part = part.strip()
            if part.startswith("session="):
                token = part[8:]
                break
        if token:
            db_exec("DELETE FROM sessions WHERE token=?", (token,))

        self.send_response(200)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        origin = self._get_cors_origin()
        self.send_header("Access-Control-Allow-Origin", origin)
        if origin != "*":
            self.send_header("Access-Control-Allow-Credentials", "true")
        self.send_header("Set-Cookie", "session=; Path=/; HttpOnly; Max-Age=0")
        self.end_headers()
        self.wfile.write(json.dumps({"ok": True}, ensure_ascii=False).encode("utf-8"))
        log("AUTH logout")

    # ── Response ──

    def _get_cors_origin(self):
        origin = self.headers.get("Origin", "")
        return origin if origin else "*"

    def _json(self, data, code=200):
        self.send_response(code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        origin = self._get_cors_origin()
        self.send_header("Access-Control-Allow-Origin", origin)
        if origin != "*":
            self.send_header("Access-Control-Allow-Credentials", "true")
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode("utf-8"))

    def log_message(self, format, *args):
        pass  # 조용한 HTTP 로그


class ThreadedServer(ThreadingMixIn, HTTPServer):
    allow_reuse_address = True
    daemon_threads = True

if __name__ == "__main__":
    os.chdir(BASE_DIR)
    acquire_lock()
    server = ThreadedServer(("0.0.0.0", PORT), TmuxHandler)
    if _is_pg():
        db_info = f"postgresql://{CONFIG.get('db_user','canvas')}@{CONFIG.get('db_host','127.0.0.1')}:{CONFIG.get('db_port',5432)}/{CONFIG.get('db_name','canvas_db')}"
    else:
        db_info = DB_PATH
    print(f"╔══════════════════════════════════════════════════╗")
    print(f"║  Gil's FlowDesk on ::{PORT}                    ║")
    print(f"║  DB: {db_info}")
    print(f"║  User: {_get_machine_id()}")
    print(f"╚══════════════════════════════════════════════════╝")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nShutdown")
        _remove_lock()
        server.server_close()
